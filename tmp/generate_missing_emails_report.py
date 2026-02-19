from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent.parent

INV_PATH = ROOT / "tmp" / "missing_emails_investors.json"
INTRO_PATH = ROOT / "tmp" / "missing_emails_introducers.json"
BY_VEHICLE_PATH = ROOT / "tmp" / "missing_emails_by_vehicle.json"

OUT_PATH = ROOT / "verso_capital_2_data" / "Missing_Emails_Report.xlsx"


def autosize_columns(ws, max_width: int = 60) -> None:
    for idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            v = str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        width = min(max_len + 2, max_width)
        ws.column_dimensions[get_column_letter(idx)].width = max(10, width)


def style_header_row(ws, row: int = 1) -> None:
    bold = Font(bold=True)
    for cell in ws[row]:
        cell.font = bold
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def main() -> None:
    investors = json.loads(INV_PATH.read_text(encoding="utf-8"))
    introducers = json.loads(INTRO_PATH.read_text(encoding="utf-8"))
    by_vehicle = json.loads(BY_VEHICLE_PATH.read_text(encoding="utf-8"))

    wb = Workbook()
    wb.remove(wb.active)

    generated_at = datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M:%S %Z")

    # Summary
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Missing Emails Report (Production)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generated at: {generated_at}"

    ws["A4"] = "Unique Investors (with subscriptions) missing email"
    ws["B4"] = len(investors)
    ws["A5"] = "Unique Introducers (with commissions) missing email"
    ws["B5"] = len(introducers)

    ws["A7"] = "Per-Vehicle Counts"
    ws["A7"].font = Font(bold=True)
    ws["A8"] = "Note: introducer count is per-vehicle distinct; the same introducer can appear in multiple vehicles."

    start_row = 10
    ws.cell(start_row, 1, "Vehicle")
    ws.cell(start_row, 2, "Investors Missing Email")
    ws.cell(start_row, 3, "Introducers Missing Email")
    style_header_row(ws, start_row)

    for r, row in enumerate(by_vehicle, start=start_row + 1):
        ws.cell(r, 1, row.get("entity_code"))
        ws.cell(r, 2, row.get("investors_missing_email"))
        ws.cell(r, 3, row.get("introducers_missing_email"))

    ws.freeze_panes = "A10"
    autosize_columns(ws)

    # Investors
    ws = wb.create_sheet("Investors")
    inv_headers = [
        "Legal Name",
        "Display Name",
        "Type",
        "First Name",
        "Middle Name",
        "Last Name",
        "Representative Name",
        "Email (TO FILL)",
        "Entity Codes",
        "Subscription Count",
    ]
    ws.append(inv_headers)
    style_header_row(ws, 1)

    for inv in investors:
        ws.append(
            [
                inv.get("legal_name"),
                inv.get("display_name"),
                inv.get("type"),
                inv.get("first_name"),
                inv.get("middle_name"),
                inv.get("last_name"),
                inv.get("representative_name"),
                "",  # email missing
                inv.get("entity_codes"),
                inv.get("subscription_count"),
            ]
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    autosize_columns(ws)

    # Introducers
    ws = wb.create_sheet("Introducers")
    intro_headers = [
        "Legal Name",
        "Display Name",
        "Email (TO FILL)",
        "Entity Codes",
        "Commission Rows",
    ]
    ws.append(intro_headers)
    style_header_row(ws, 1)

    for intro in introducers:
        ws.append(
            [
                intro.get("legal_name"),
                intro.get("display_name"),
                "",  # email missing
                intro.get("entity_codes"),
                intro.get("commission_rows"),
            ]
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    autosize_columns(ws)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)


if __name__ == "__main__":
    main()
