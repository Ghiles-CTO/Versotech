#!/usr/bin/env python3
"""
Generate Introducer Commissions Excel Report
============================================
Generates 05_Introducer_Commissions.xlsx with Entity Code column from database.

Columns:
- Entity Code (vehicle code)
- Introducer
- Investor
- Fee Type (basis_type)
- Rate (bps)
- Rate (%)
- Commission Amount
- Currency
"""

import json
from datetime import datetime
import os
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# File paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_PATH = os.path.join(SCRIPT_DIR, "commission_data_export.json")
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "05_Introducer_Commissions.xlsx")
COMMENTS_PATH = os.path.join(SCRIPT_DIR, "..", "VERSO", "datafixing", "05_Introducer_Commissions_FD comments.xlsx")
ACTIONS_LOG_PATH = os.path.join(SCRIPT_DIR, "..", "VERSO", "datafixing", "05_commission_actions_log.md")

STANDARD_COLUMNS = [
    "Entity Code",
    "Introducer",
    "Investor",
    "Fee Type",
    "Rate (bps)",
    "Rate (%)",
    "Commission Amount",
    "Currency",
]


def _normalize_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (int, float)):
        if abs(value - int(value)) < 1e-9:
            return str(int(value))
        return f"{value:.2f}"
    return str(value).strip()


def _row_key(values: Iterable) -> tuple[str, ...]:
    return tuple(_normalize_value(v) for v in values)


def _is_red_color(color) -> bool:
    if color is None:
        return False
    rgb = getattr(color, "rgb", None)
    if not rgb:
        return False
    if hasattr(rgb, "value"):
        rgb = rgb.value
    rgb = str(rgb).upper()
    return rgb in {"FF0000", "FFFF0000", "00FF0000"} or rgb.endswith("FF0000")


def _cell_marked_red(cell) -> bool:
    if _is_red_color(cell.fill.fgColor):
        return True
    return _is_red_color(cell.font.color)


def load_comment_changes(comment_path: str) -> tuple[set[tuple[str, ...]], set[tuple[str, ...]]]:
    """Return (updated_keys, deleted_keys) from the commented Excel file."""
    if not os.path.exists(comment_path):
        return set(), set()

    wb = load_workbook(comment_path, data_only=True)
    ws = wb.active
    if "Introducer" in wb.sheetnames:
        ws = wb["Introducer"]
    elif "Introducer Commissions" in wb.sheetnames:
        ws = wb["Introducer Commissions"]

    header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {str(h).strip().lower(): idx for idx, h in enumerate(header_row) if h}

    col_indices = []
    for name in STANDARD_COLUMNS:
        idx = header_map.get(name.lower())
        col_indices.append(idx)

    if any(idx is None for idx in col_indices):
        # Can't reliably map; skip highlighting
        return set(), set()

    updated_keys: set[tuple[str, ...]] = set()
    deleted_keys: set[tuple[str, ...]] = set()

    for row in ws.iter_rows(min_row=2):
        values = [row[i].value for i in col_indices]
        key = _row_key(values)
        if not any(key):
            continue
        row_has_strike = any(getattr(cell.font, "strike", False) for cell in row if cell.value is not None)
        row_has_red = any(_cell_marked_red(cell) for cell in row if cell.value is not None)
        if row_has_strike:
            deleted_keys.add(key)
        elif row_has_red:
            updated_keys.add(key)

    return updated_keys, deleted_keys


def load_commission_data(input_path: str) -> list:
    """Load commission data from JSON export file."""
    with open(input_path, 'r') as f:
        data = json.load(f)
    return data


def load_actions_log_lines(log_path: str) -> list[str]:
    if not os.path.exists(log_path):
        return []
    lines: list[str] = []
    for raw in Path(log_path).read_text().splitlines():
        line = raw.strip()
        if not line:
            continue
        if line.startswith("#"):
            lines.append(line.lstrip("#").strip().upper())
            continue
        if line.startswith("-"):
            lines.append(line.lstrip("-").strip())
        else:
            lines.append(line)
    return lines

def generate_excel(data: list, output_path: str, updated_keys: set[tuple[str, ...]], deleted_keys: set[tuple[str, ...]]):
    """Generate Excel file from data."""
    # Create DataFrame with proper column names
    df = pd.DataFrame(data)

    # Rename columns for Excel headers
    df = df.rename(columns={
        'entity_code': 'Entity Code',
        'introducer': 'Introducer',
        'investor': 'Investor',
        'fee_type': 'Fee Type',
        'rate_bps': 'Rate (bps)',
        'rate_pct': 'Rate (%)',
        'commission_amount': 'Commission Amount',
        'currency': 'Currency'
    })

    # Reorder columns
    columns = ['Entity Code', 'Introducer', 'Investor', 'Fee Type', 'Rate (bps)', 'Rate (%)', 'Commission Amount', 'Currency']
    df = df[columns]

    # Write to Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Introducer Commissions', index=False)

        ws = writer.sheets['Introducer Commissions']
        header = [cell.value for cell in ws[1]]
        header_map = {str(h).strip().lower(): idx for idx, h in enumerate(header) if h}
        col_indices = [header_map.get(name.lower()) for name in STANDARD_COLUMNS]

        highlight_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        if all(idx is not None for idx in col_indices):
            for row_idx in range(2, ws.max_row + 1):
                row_values = [ws.cell(row=row_idx, column=idx + 1).value for idx in col_indices]
                key = _row_key(row_values)
                if key in updated_keys:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col).fill = highlight_fill

        # Add change log sheet
        log_ws = writer.book.create_sheet("Change Log")
        log_ws.append(["Change Type"] + STANDARD_COLUMNS)
        for key in sorted(updated_keys):
            log_ws.append(["Updated (red in comments)"] + list(key))
        for key in sorted(deleted_keys):
            log_ws.append(["Deleted (strikethrough)"] + list(key))

        # Add applied changes summary sheet
        summary_lines = load_actions_log_lines(ACTIONS_LOG_PATH)
        summary_ws = writer.book.create_sheet("Applied Changes Summary")
        summary_ws.append(["Source", ACTIONS_LOG_PATH])
        summary_ws.append([])
        for line in summary_lines:
            summary_ws.append([line])
        summary_ws.column_dimensions["A"].width = 120

    print(f"Generated: {output_path}")
    print(f"Total records: {len(df)}")
    print(f"\nRecords by Entity Code:")
    print(df.groupby('Entity Code').size().to_string())

def main():
    """Main function."""
    print("=" * 60)
    print("Generating Introducer Commissions Excel Report")
    print(f"Timestamp: {datetime.now().isoformat()}")
    print("=" * 60)

    try:
        # Load data from JSON
        print(f"\nLoading data from: {INPUT_PATH}")
        data = load_commission_data(INPUT_PATH)
        print(f"Loaded {len(data)} records")

        # Generate Excel
        print("\nGenerating Excel file...")
        updated_keys, deleted_keys = load_comment_changes(COMMENTS_PATH)
        generate_excel(data, OUTPUT_PATH, updated_keys, deleted_keys)

        print("\n" + "=" * 60)
        print("SUCCESS: Excel file generated with Entity Code column")
        print("=" * 60)

    except Exception as e:
        print(f"\nERROR: {e}")
        raise

if __name__ == "__main__":
    main()
