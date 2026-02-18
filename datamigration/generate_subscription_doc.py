#!/usr/bin/env python3
"""
Generate 06_Full_Subscription_Data.xlsx from database.
Includes subscription fields plus current position per investor/vehicle.
"""

import json
import pandas as pd
from datetime import datetime
import os

def main():
    # Load data from the extracted JSON
    with open("/tmp/subscription_data.json", "r") as f:
        data = json.load(f)

    print(f"Loaded {len(data)} subscription records")

    # Create DataFrame
    df = pd.DataFrame(data)

    column_order = [
        "Entity Code",
        "Investor Name",
        "Investor Type",
        "Currency",
        "Commitment",
        "Funded Amount",
        "Shares",
        "Current Position",
        "Price Per Share",
        "Cost Per Share",
        "Subscription Fee (%)",
        "Subscription Fee Amount",
        "Performance Fee Tier 1 (%)",
        "Performance Fee Tier 1 Threshold",
        "Performance Fee Tier 2 (%)",
        "Performance Fee Tier 2 Threshold",
        "Spread Per Share",
        "Spread Fee Amount",
        "Management Fee (%)",
        "BD Fee (%)",
        "BD Fee Amount",
        "FINRA Fee Shares",
        "FINRA Fee Amount",
        "Discount Rate",
        "Interest Rate",
        "Valuation Cap",
        "Opportunity Name",
        "Sourcing Contract Ref",
        "Contract Date",
    ]

    df = df[[col for col in column_order if col in df.columns]]

    # Convert numeric columns
    numeric_columns = [
        "Commitment",
        "Funded Amount",
        "Shares",
        "Current Position",
        "Price Per Share",
        "Cost Per Share",
        "Subscription Fee (%)",
        "Subscription Fee Amount",
        "Performance Fee Tier 1 (%)",
        "Performance Fee Tier 1 Threshold",
        "Performance Fee Tier 2 (%)",
        "Performance Fee Tier 2 Threshold",
        "Spread Per Share",
        "Spread Fee Amount",
        "Management Fee (%)",
        "BD Fee (%)",
        "BD Fee Amount",
        "FINRA Fee Shares",
        "FINRA Fee Amount",
        "Discount Rate",
        "Interest Rate",
        "Valuation Cap",
    ]
    for col in numeric_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Convert Contract Date to datetime, handle nulls
    df["Contract Date"] = pd.to_datetime(df["Contract Date"], errors="coerce")

    # Sort by Entity Code, then Investor Name, then Contract Date
    df = df.sort_values(["Entity Code", "Investor Name", "Contract Date"])

    # Calculate totals for TOTALS row
    totals = {col: None for col in df.columns}
    totals["Entity Code"] = "TOTALS"
    totals["Investor Name"] = ""
    totals["Investor Type"] = ""
    for col in ["Commitment", "Funded Amount", "Shares", "Current Position", "Subscription Fee Amount", "Spread Fee Amount", "BD Fee Amount", "FINRA Fee Amount"]:
        if col in df.columns:
            totals[col] = df[col].sum()

    # Append totals row
    df = pd.concat([df, pd.DataFrame([totals])], ignore_index=True)

    print(f"\nDocument Summary:")
    print(f"  Total records: {len(df) - 1} (plus TOTALS row)")
    print(f"  Total Commitment: ${totals['Commitment']:,.2f}")
    print(f"  Total Funded Amount: ${totals['Funded Amount']:,.2f}")
    print(f"  Total Shares: {totals['Shares']:,.2f}")

    # Save to Excel
    output_path = "/Users/ghilesmoussaoui/Desktop/Versotech/datamigration/06_Full_Subscription_Data_REGENERATED.xlsx"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Subscription Data", index=False)

        # Get workbook and worksheet for formatting
        workbook = writer.book
        worksheet = writer.sheets["Subscription Data"]

        # Adjust column widths
        width_by_name = {
            "Entity Code": 12,
            "Investor Name": 40,
            "Investor Type": 14,
            "Currency": 10,
            "Commitment": 16,
            "Funded Amount": 16,
            "Shares": 12,
            "Current Position": 16,
            "Price Per Share": 14,
            "Cost Per Share": 14,
            "Subscription Fee (%)": 16,
            "Subscription Fee Amount": 18,
            "Performance Fee Tier 1 (%)": 20,
            "Performance Fee Tier 1 Threshold": 22,
            "Performance Fee Tier 2 (%)": 20,
            "Performance Fee Tier 2 Threshold": 22,
            "Spread Per Share": 16,
            "Spread Fee Amount": 18,
            "Management Fee (%)": 18,
            "BD Fee (%)": 14,
            "BD Fee Amount": 16,
            "FINRA Fee Shares": 16,
            "FINRA Fee Amount": 16,
            "Discount Rate": 14,
            "Interest Rate": 14,
            "Valuation Cap": 16,
            "Opportunity Name": 32,
            "Sourcing Contract Ref": 24,
            "Contract Date": 14,
        }

        from openpyxl.utils import get_column_letter

        headers = [cell.value for cell in worksheet[1]]
        for idx, name in enumerate(headers, start=1):
            width = width_by_name.get(name, 14)
            worksheet.column_dimensions[get_column_letter(idx)].width = width

        # Apply number formatting
        from openpyxl.styles import numbers

        currency_cols = {
            "Commitment",
            "Funded Amount",
            "Subscription Fee Amount",
            "Spread Fee Amount",
            "BD Fee Amount",
            "FINRA Fee Amount",
            "Valuation Cap",
        }
        decimal_cols = {
            "Shares",
            "Current Position",
            "Price Per Share",
            "Cost Per Share",
            "Spread Per Share",
            "FINRA Fee Shares",
            "Performance Fee Tier 1 Threshold",
            "Performance Fee Tier 2 Threshold",
        }
        percent_cols = {
            "Subscription Fee (%)",
            "Performance Fee Tier 1 (%)",
            "Performance Fee Tier 2 (%)",
            "Management Fee (%)",
            "BD Fee (%)",
            "Discount Rate",
            "Interest Rate",
        }

        header_map = {name: idx + 1 for idx, name in enumerate(headers)}
        for row in range(2, len(df) + 2):  # Skip header row
            for col_name in headers:
                col_idx = header_map[col_name]
                cell = worksheet.cell(row=row, column=col_idx)
                if col_name in currency_cols:
                    cell.number_format = '#,##0.00'
                elif col_name in decimal_cols:
                    if col_name in {"Price Per Share", "Cost Per Share"}:
                        cell.number_format = '#,##0.000000'
                    else:
                        cell.number_format = '#,##0.00'
                elif col_name in percent_cols:
                    cell.number_format = '0.00%'
                elif col_name == "Contract Date" and cell.value:
                    cell.number_format = 'YYYY-MM-DD'

    print(f"\nSaved to: {output_path}")

    # Print verification by entity
    print("\n" + "=" * 70)
    print("VERIFICATION BY ENTITY CODE")
    print("=" * 70)

    df_data = df[df["Entity Code"] != "TOTALS"]
    by_entity = df_data.groupby("Entity Code").agg({
        "Commitment": "sum",
        "Funded Amount": "sum",
        "Shares": "sum",
        "Current Position": "sum",
        "Investor Name": "count"
    }).rename(columns={"Investor Name": "Record Count"})

    print(by_entity.to_string())
    print("=" * 70)

if __name__ == "__main__":
    main()
