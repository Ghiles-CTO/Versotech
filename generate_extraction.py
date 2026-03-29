#!/usr/bin/env python3
"""
Generate Investor_Extraction_Production.xlsx from 6 JSON data dumps.

Data sources:
  - extraction_data/investors.json
  - extraction_data/subscriptions.json
  - extraction_data/positions.json
  - extraction_data/entity_investors.json
  - extraction_data/introductions.json
  - extraction_data/commissions.json

Each JSON file is a flat array of objects (rows).
"""

import json
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "extraction_data")
OUTPUT_FILE = os.path.join(BASE_DIR, "Investor_Extraction_Production.xlsx")

# ── Sheet config ──────────────────────────────────────────────────────────
# (filename, sheet_name, name_columns_to_put_first)
SHEETS = [
    ("investors.json",        "Investors",              ["investor_name"]),
    ("subscriptions.json",    "Subscriptions",          ["investor_name", "vehicle_name", "deal_name", "introducer_name"]),
    ("positions.json",        "Positions",              ["investor_name", "vehicle_name"]),
    ("entity_investors.json", "Entity Investors",       ["investor_name", "vehicle_name"]),
    ("introductions.json",    "Introductions",          ["investor_name", "introducer_name", "deal_name"]),
    ("commissions.json",      "Introducer Commissions", ["investor_name", "introducer_name", "deal_name"]),
]

# ── Styling ───────────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
NAME_COL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
DATA_FONT = Font(name="Calibri", size=10)
STRIPE_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")


def snake_to_title(s: str) -> str:
    """Convert snake_case to Title Case, with common abbreviation handling."""
    abbrevs = {
        "id": "ID", "kyc": "KYC", "aml": "AML", "rm": "RM", "pep": "PEP",
        "url": "URL", "nav": "NAV", "bps": "BPS", "pdf": "PDF", "us": "US",
        "spf": "SPF", "bd": "BD", "finra": "FINRA",
    }
    words = s.split("_")
    result = []
    for w in words:
        if w.lower() in abbrevs:
            result.append(abbrevs[w.lower()])
        else:
            result.append(w.capitalize())
    return " ".join(result)


def load_json(filename: str) -> list[dict]:
    """Load a JSON file from the data directory."""
    path = os.path.join(DATA_DIR, filename)
    with open(path, "r") as f:
        data = json.load(f)
    if not data:
        return []
    # Handle json_agg wrapper: [{"json_agg": [...]}]
    if isinstance(data, list) and len(data) == 1 and isinstance(data[0], dict) and "json_agg" in data[0]:
        return data[0]["json_agg"] or []
    return data


def get_non_null_columns(rows: list[dict]) -> list[str]:
    """Return columns that have at least one non-null value across all rows."""
    if not rows:
        return []
    all_cols = list(rows[0].keys())
    non_null = []
    for col in all_cols:
        for row in rows:
            val = row.get(col)
            if val is not None:
                non_null.append(col)
                break
    return non_null


def order_columns(non_null_cols: list[str], name_cols_first: list[str]) -> list[str]:
    """Put resolved name columns first, then the rest in original order."""
    first = [c for c in name_cols_first if c in non_null_cols]
    rest = [c for c in non_null_cols if c not in first]
    return first + rest


def auto_width(ws, col_idx: int, header: str, rows: list[dict], col_key: str) -> float:
    """Calculate a reasonable column width."""
    max_len = len(header)
    sample = rows[:50]  # Check first 50 rows for performance
    for row in sample:
        val = row.get(col_key)
        if val is not None:
            cell_len = len(str(val))
            if cell_len > max_len:
                max_len = cell_len
    # Cap width and add padding
    return min(max_len + 3, 50)


def write_sheet(wb: Workbook, sheet_name: str, rows: list[dict], name_cols_first: list[str]):
    """Write one sheet with data, formatting, filters, and frozen panes."""
    ws = wb.create_sheet(title=sheet_name)

    if not rows:
        ws.cell(row=1, column=1, value="No data")
        return

    # Determine columns
    non_null_cols = get_non_null_columns(rows)
    ordered_cols = order_columns(non_null_cols, name_cols_first)
    headers = [snake_to_title(c) for c in ordered_cols]

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Write data rows
    for row_idx, row_data in enumerate(rows, 2):
        is_name_col_set = set(name_cols_first)
        for col_idx, col_key in enumerate(ordered_cols, 1):
            val = row_data.get(col_key)
            # Convert booleans to readable strings
            if isinstance(val, bool):
                val = "Yes" if val else "No"
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

            # Stripe alternate rows
            if row_idx % 2 == 0:
                cell.fill = STRIPE_FILL

            # Highlight name columns with light blue
            if col_key in is_name_col_set and row_idx % 2 != 0:
                cell.fill = NAME_COL_FILL

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(ordered_cols))}{len(rows) + 1}"

    # Freeze top row
    ws.freeze_panes = "A2"

    # Set column widths
    for col_idx, col_key in enumerate(ordered_cols, 1):
        width = auto_width(ws, col_idx, headers[col_idx - 1], rows, col_key)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    total_rows = 0
    for filename, sheet_name, name_cols in SHEETS:
        print(f"Loading {filename}...", end=" ")
        rows = load_json(filename)
        print(f"{len(rows)} rows, ", end="")

        non_null = get_non_null_columns(rows)
        all_cols = list(rows[0].keys()) if rows else []
        skipped = len(all_cols) - len(non_null)
        print(f"{len(non_null)} columns with data (skipped {skipped} all-null)")

        write_sheet(wb, sheet_name, rows, name_cols)
        total_rows += len(rows)

    wb.save(OUTPUT_FILE)
    print(f"\nSaved to: {OUTPUT_FILE}")
    print(f"Total: {len(SHEETS)} sheets, {total_rows} rows")


if __name__ == "__main__":
    main()
