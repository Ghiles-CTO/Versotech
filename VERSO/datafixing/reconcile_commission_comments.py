#!/usr/bin/env python3
"""Reconcile client commission comments with DB commissions.

Outputs:
- VERSO/datafixing/05_commission_updates_to_apply.csv
- VERSO/datafixing/05_commission_unmatched_needs_review.csv
- VERSO/datafixing/05_commission_match_summary.md
"""
from __future__ import annotations

import csv
import json
import math
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import urlencode
from urllib.request import Request, urlopen

import openpyxl

ENV_PATH = Path('.env.local')
COMMENTS_XLSX = Path('VERSO/datafixing/05_Introducer_Commissions_FD comments.xlsx')
INTRODUCER_RENAME_CSV = Path('VERSO/datafixing/introducers name change.csv')

OUT_MATCHED = Path('VERSO/datafixing/05_commission_updates_to_apply.csv')
OUT_UNMATCHED = Path('VERSO/datafixing/05_commission_unmatched_needs_review.csv')
OUT_SUMMARY = Path('VERSO/datafixing/05_commission_match_summary.md')


@dataclass
class CommentRow:
    row_index: int
    entity_code: Optional[str]
    introducer: Optional[str]
    investor: Optional[str]
    fee_type: Optional[str]
    rate_bps: Optional[float]
    rate_pct: Optional[float]
    amount: Optional[float]
    currency: Optional[str]
    is_strike: bool
    is_red: bool
    raw: Dict[str, Any]


@dataclass
class CommissionRow:
    id: str
    entity_code: Optional[str]
    introducer: Optional[str]
    investor: Optional[str]
    fee_type: Optional[str]
    rate_bps: Optional[float]
    amount: Optional[float]
    currency: Optional[str]
    tier_number: Optional[int]
    created_at: Optional[str]
    introduction_id: Optional[str]
    deal_id: Optional[str]
    investor_id: Optional[str]
    introducer_id: Optional[str]
    base_amount: Optional[float]


@dataclass
class SubscriptionRow:
    id: str
    introduction_id: Optional[str]
    investor_id: Optional[str]
    deal_id: Optional[str]
    contract_date: Optional[str]
    commitment: Optional[float]
    num_shares: Optional[float]
    price_per_share: Optional[float]
    cost_per_share: Optional[float]
    subscription_fee_amount: Optional[float]
    spread_fee_amount: Optional[float]
    funded_amount: Optional[float]
    units: Optional[float]


def load_env(path: Path) -> dict:
    env: dict[str, str] = {}
    if not path.exists():
        return env
    for raw in path.read_text().splitlines():
        line = raw.strip()
        if not line or line.startswith('#') or '=' not in line:
            continue
        key, value = line.split('=', 1)
        env[key.strip()] = value.strip().strip('"')
    return env


def fetch_all(base: dict, table: str, select: str, page_size: int = 1000) -> List[dict]:
    headers = {
        'apikey': base['key'],
        'Authorization': f"Bearer {base['key']}",
        'Accept': 'application/json',
    }
    results: List[dict] = []
    offset = 0
    while True:
        params = {
            'select': select,
            'order': 'id',
            'limit': page_size,
            'offset': offset,
        }
        url = f"{base['url']}/rest/v1/{table}?{urlencode(params)}"
        req = Request(url, headers=headers)
        try:
            with urlopen(req) as resp:
                payload = resp.read()
        except Exception as exc:
            if hasattr(exc, 'read'):
                error_body = exc.read().decode('utf-8', errors='ignore')
                raise RuntimeError(f"Request failed: {url}\n{error_body}") from exc
            raise
        data = json.loads(payload.decode('utf-8'))
        if not data:
            break
        results.extend(data)
        if len(data) < page_size:
            break
        offset += page_size
    return results


def norm_text(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip().lower()
    if not text:
        return None
    text = text.replace('&', 'and')
    text = re.sub(r"\b(limited)\b", "ltd", text)
    text = re.sub(r"\b(incorporated)\b", "inc", text)
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text or None


def norm_investor(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip().lower()
    if not text:
        return None
    # remove parenthetical notes like "(on behalf of ...)"
    text = re.sub(r"\([^\)]*\)", " ", text)
    text = text.replace('&', ' and ')
    text = re.sub(r"\b(limited)\b", "ltd", text)
    text = re.sub(r"\b(incorporated)\b", "inc", text)
    tokens = re.split(r"[^a-z0-9]+", text)
    stopwords = {'mr', 'mrs', 'ms', 'dr', 'and', 'of', 'the', 'on', 'behalf'}
    tokens = [t for t in tokens if t and t not in stopwords]
    if not tokens:
        return None
    tokens = sorted(tokens)
    return ''.join(tokens)


def to_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except Exception:
        return None


def amount_key(value: Optional[float]) -> Optional[str]:
    if value is None:
        return None
    return f"{value:.2f}"


def rate_bps_from_pct(rate_pct: Optional[float]) -> Optional[int]:
    if rate_pct is None:
        return None
    try:
        return int(round(rate_pct * 100))
    except Exception:
        return None


def load_introducer_mapping(path: Path) -> dict[str, str]:
    mapping: dict[str, str] = {}
    if not path.exists():
        return mapping
    with path.open(newline='') as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            current = row.get('Current Name') or row.get('Current') or row.get('CurrentName')
            full = row.get('Full Name') or row.get('Full') or row.get('FullName')
            if not current or not full:
                continue
            mapping[current.strip().lower()] = full.strip()
    # manual overrides
    manual = {
        'denis': 'Denis Matthey',
        'robin': 'Robin Doble',
        'sandro': 'GEMERA Consulting Pte Ltd',
        'simone': 'Manna Capital',
        'rick': 'Altras Capital Financing Broker',
        'elevation+rick': 'Altras Capital Financing Broker',
        'gemera': 'GEMERA Consulting Pte Ltd',
    }
    for key, value in manual.items():
        mapping.setdefault(key, value)
    return mapping


def map_introducer(name: Optional[str], mapping: dict[str, str]) -> Optional[str]:
    if name is None:
        return None
    raw = str(name).strip()
    if not raw:
        return None
    lowered = raw.lower()
    lookup = mapping.get(lowered)
    if lookup:
        return lookup
    # handle combined labels like \"Sandro GEMERA Consulting Pte Ltd\" or \"Set Cap\"
    if 'gemera' in lowered or 'sandro' in lowered:
        return 'GEMERA Consulting Pte Ltd'
    if 'set cap' in lowered or 'setcap' in lowered:
        return 'Setcap'
    return raw


def detect_red_fill(cell) -> bool:
    fill = cell.fill
    if fill is None or fill.patternType is None:
        return False
    color = fill.fgColor
    if color is None:
        return False
    rgb = color.rgb
    if not rgb:
        return False
    # normalize rgb like 'FFFF0000' or 'FF0000'
    rgb = rgb.upper()
    return rgb.endswith('FF0000') or rgb in {'FFFF0000', '00FF0000'}


def detect_red_font(cell) -> bool:
    font = cell.font
    if font is None:
        return False
    color = font.color
    if color is None:
        return False
    rgb = color.rgb or color.value
    if not rgb:
        return False
    rgb = str(rgb).upper()
    return rgb.endswith('FF0000') or rgb in {'FFFF0000', '00FF0000'}


def detect_strike(cell) -> bool:
    font = cell.font
    if font is None:
        return False
    return bool(font.strike)


def load_comment_rows(path: Path) -> List[CommentRow]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, 10)]
    rows: List[CommentRow] = []
    for r in range(2, ws.max_row + 1):
        values = [ws.cell(r, c).value for c in range(1, 10)]
        row_data = dict(zip(headers, values))
        entity = row_data.get('Entity Code')
        if entity is None and all(v in (None, '') for v in values):
            continue
        strike = False
        red = False
        for c in range(1, 9):
            cell = ws.cell(r, c)
            if detect_strike(cell):
                strike = True
            if detect_red_fill(cell) or detect_red_font(cell):
                red = True
        rows.append(
            CommentRow(
                row_index=r,
                entity_code=str(entity).strip() if entity else None,
                introducer=row_data.get('Introducer'),
                investor=row_data.get('Investor'),
                fee_type=row_data.get('Fee Type'),
                rate_bps=to_float(row_data.get('Rate (bps)')),
                rate_pct=to_float(row_data.get('Rate (%)')),
                amount=to_float(row_data.get('Commission Amount')),
                currency=row_data.get('Currency'),
                is_strike=strike,
                is_red=red,
                raw=row_data,
            )
        )
    return rows


def load_commissions(base: dict) -> List[CommissionRow]:
    select = ','.join([
        'id', 'basis_type', 'rate_bps', 'accrual_amount', 'currency', 'tier_number', 'created_at',
        'introduction_id', 'base_amount', 'deal_id', 'investor_id', 'introducer_id',
        'deal:deals(id,vehicle:vehicles(entity_code))',
        'introducer:introducers(id,display_name,legal_name)',
        'investor:investors(id,display_name,legal_name)'
    ])
    raw = fetch_all(base, 'introducer_commissions', select)
    rows: List[CommissionRow] = []
    for row in raw:
        deal = row.get('deal') or {}
        vehicle = deal.get('vehicle') or {}
        introducer = row.get('introducer') or {}
        investor = row.get('investor') or {}
        rows.append(
            CommissionRow(
                id=row.get('id'),
                entity_code=vehicle.get('entity_code'),
                introducer=introducer.get('legal_name') or introducer.get('display_name'),
                investor=investor.get('legal_name') or investor.get('display_name'),
                fee_type=row.get('basis_type'),
                rate_bps=to_float(row.get('rate_bps')),
                amount=to_float(row.get('accrual_amount')),
                currency=row.get('currency'),
                tier_number=row.get('tier_number'),
                created_at=row.get('created_at'),
                introduction_id=row.get('introduction_id'),
                deal_id=row.get('deal_id'),
                investor_id=row.get('investor_id'),
                introducer_id=row.get('introducer_id'),
                base_amount=to_float(row.get('base_amount')),
            )
        )
    return rows


def load_subscriptions(base: dict) -> List[SubscriptionRow]:
    select = ','.join([
        'id', 'introduction_id', 'investor_id', 'deal_id', 'contract_date',
        'commitment', 'num_shares', 'price_per_share', 'cost_per_share',
        'subscription_fee_amount', 'spread_fee_amount', 'funded_amount', 'units'
    ])
    raw = fetch_all(base, 'subscriptions', select)
    rows: List[SubscriptionRow] = []
    for row in raw:
        rows.append(
            SubscriptionRow(
                id=row.get('id'),
                introduction_id=row.get('introduction_id'),
                investor_id=row.get('investor_id'),
                deal_id=row.get('deal_id'),
                contract_date=row.get('contract_date'),
                commitment=to_float(row.get('commitment')),
                num_shares=to_float(row.get('num_shares')),
                price_per_share=to_float(row.get('price_per_share')),
                cost_per_share=to_float(row.get('cost_per_share')),
                subscription_fee_amount=to_float(row.get('subscription_fee_amount')),
                spread_fee_amount=to_float(row.get('spread_fee_amount')),
                funded_amount=to_float(row.get('funded_amount')),
                units=to_float(row.get('units')),
            )
        )
    return rows


def build_subscription_index(subs: List[SubscriptionRow]):
    by_intro: dict[str, List[SubscriptionRow]] = defaultdict(list)
    by_investor_deal: dict[Tuple[str, str], List[SubscriptionRow]] = defaultdict(list)
    for row in subs:
        if row.introduction_id:
            by_intro[row.introduction_id].append(row)
        if row.investor_id and row.deal_id:
            by_investor_deal[(row.investor_id, row.deal_id)].append(row)
    return by_intro, by_investor_deal


def commission_match_key(row: CommissionRow) -> Tuple:
    return (
        row.entity_code,
        norm_text(row.introducer),
        norm_investor(row.investor),
        (row.fee_type or '').strip().lower() if row.fee_type else None,
    )


def match_comment_row(
    comment: CommentRow,
    comm_by_key: dict,
    comm_rows: List[CommissionRow],
    action: str,
) -> Tuple[Optional[CommissionRow], str, List[CommissionRow]]:
    fee_type = (comment.fee_type or '').strip().lower() if comment.fee_type else None
    intro_norm = norm_text(comment.introducer)
    inv_norm = norm_investor(comment.investor)
    key = (comment.entity_code, intro_norm, inv_norm, fee_type)
    candidates = comm_by_key.get(key, []) if intro_norm and inv_norm else []

    # fallback: filter by entity+fee_type then name(s)
    if not candidates:
        candidates = [r for r in comm_rows if r.entity_code == comment.entity_code and (r.fee_type or '').lower() == fee_type]
        if intro_norm:
            candidates = [r for r in candidates if norm_text(r.introducer) == intro_norm]
        if inv_norm:
            candidates = [r for r in candidates if norm_investor(r.investor) == inv_norm]

    if not candidates:
        if action == 'update':
            # Relax introducer match for updates (introducer might be corrected in comments)
            relaxed = [r for r in comm_rows if r.entity_code == comment.entity_code and (r.fee_type or '').lower() == fee_type]
            if inv_norm:
                relaxed = [r for r in relaxed if norm_investor(r.investor) == inv_norm]
            if len(relaxed) == 1:
                return relaxed[0], 'MATCH_RELAXED_INVESTOR_ONLY', relaxed
            if len(relaxed) > 1:
                # Try to break ties with amount or rate if they still match any
                amount = amount_key(comment.amount)
                rate_bps = comment.rate_bps
                if rate_bps is None and comment.rate_pct is not None:
                    rate_bps = rate_bps_from_pct(comment.rate_pct)
                if amount is not None:
                    amt = [r for r in relaxed if amount_key(r.amount) == amount]
                    if len(amt) == 1:
                        return amt[0], 'MATCH_RELAXED_INVESTOR_AMOUNT', relaxed
                if rate_bps is not None:
                    rate = [r for r in relaxed if r.rate_bps is not None and int(r.rate_bps) == int(rate_bps)]
                    if len(rate) == 1:
                        return rate[0], 'MATCH_RELAXED_INVESTOR_RATE', relaxed
                return None, 'AMBIGUOUS_RELAXED_INVESTOR', relaxed
        return None, 'NO_CANDIDATES', []

    amount = amount_key(comment.amount)
    rate_bps = comment.rate_bps
    if rate_bps is None and comment.rate_pct is not None:
        rate_bps = rate_bps_from_pct(comment.rate_pct)

    def match_amount(row: CommissionRow) -> bool:
        if amount is None:
            return False
        return amount_key(row.amount) == amount

    def match_rate(row: CommissionRow) -> bool:
        if rate_bps is None:
            return False
        return row.rate_bps is not None and int(row.rate_bps) == int(rate_bps)

    # Strict match: amount + rate if available
    strict = [r for r in candidates if (match_amount(r) if amount is not None else True) and (match_rate(r) if rate_bps is not None else True)]
    if len(strict) == 1:
        return strict[0], 'MATCH_STRICT', candidates

    # Amount-only
    if amount is not None:
        amt = [r for r in candidates if match_amount(r)]
        if len(amt) == 1:
            return amt[0], 'MATCH_AMOUNT', candidates

    # Rate-only
    if rate_bps is not None:
        rate = [r for r in candidates if match_rate(r)]
        if len(rate) == 1:
            return rate[0], 'MATCH_RATE', candidates

    # Name-only
    if len(candidates) == 1:
        return candidates[0], 'MATCH_NAMES_ONLY', candidates

    if action == 'update':
        # If still ambiguous, attempt to relax introducer and rely on investor
        relaxed = [r for r in comm_rows if r.entity_code == comment.entity_code and (r.fee_type or '').lower() == fee_type]
        if inv_norm:
            relaxed = [r for r in relaxed if norm_investor(r.investor) == inv_norm]
        if len(relaxed) == 1:
            return relaxed[0], 'MATCH_RELAXED_INVESTOR_ONLY', relaxed
        if len(relaxed) > 1:
            return None, 'AMBIGUOUS_RELAXED_INVESTOR', relaxed
    return None, 'AMBIGUOUS', candidates


def main() -> None:
    env = load_env(ENV_PATH)
    supabase_url = env.get('NEXT_PUBLIC_SUPABASE_URL') or env.get('SUPABASE_URL')
    supabase_key = env.get('SUPABASE_SERVICE_ROLE_KEY') or env.get('SUPABASE_SERVICE_KEY')
    if not supabase_url or not supabase_key:
        raise SystemExit('Missing Supabase URL/service key in .env.local')

    base = {'url': supabase_url.rstrip('/'), 'key': supabase_key}

    comment_rows = load_comment_rows(COMMENTS_XLSX)
    mapping = load_introducer_mapping(INTRODUCER_RENAME_CSV)

    # apply mapping to comment introducers
    for row in comment_rows:
        row.introducer = map_introducer(row.introducer, mapping)

    commissions = load_commissions(base)
    subs = load_subscriptions(base)
    _ = build_subscription_index(subs)  # reserved for later, not used in matching

    comm_by_key: dict[Tuple, List[CommissionRow]] = defaultdict(list)
    for row in commissions:
        comm_by_key[commission_match_key(row)].append(row)

    matched_rows = []
    unmatched_rows = []

    for row in comment_rows:
        action = 'delete' if row.is_strike else 'update' if row.is_red else 'none'
        if action == 'none':
            continue
        match, reason, candidates = match_comment_row(row, comm_by_key, commissions, action)
        if match:
            matched_rows.append(
                {
                    'excel_row': row.row_index,
                    'action': action,
                    'commission_id': match.id,
                    'entity_code': row.entity_code,
                    'introducer_comment': row.introducer,
                    'investor_comment': row.investor,
                    'fee_type_comment': row.fee_type,
                    'rate_bps_comment': row.rate_bps,
                    'rate_pct_comment': row.rate_pct,
                    'amount_comment': row.amount,
                    'currency_comment': row.currency,
                    'introducer_db': match.introducer,
                    'investor_db': match.investor,
                    'fee_type_db': match.fee_type,
                    'rate_bps_db': match.rate_bps,
                    'amount_db': match.amount,
                    'currency_db': match.currency,
                    'match_reason': reason,
                }
            )
        else:
            unmatched_rows.append(
                {
                    'excel_row': row.row_index,
                    'action': action,
                    'entity_code': row.entity_code,
                    'introducer': row.introducer,
                    'investor': row.investor,
                    'fee_type': row.fee_type,
                    'rate_bps': row.rate_bps,
                    'rate_pct': row.rate_pct,
                    'amount': row.amount,
                    'currency': row.currency,
                    'match_reason': reason,
                    'candidate_count': len(candidates),
                    'candidate_ids': ','.join([c.id for c in candidates[:5]]),
                }
            )

    # write outputs
    if matched_rows:
        with OUT_MATCHED.open('w', newline='') as handle:
            writer = csv.DictWriter(handle, fieldnames=list(matched_rows[0].keys()))
            writer.writeheader()
            writer.writerows(matched_rows)
    else:
        OUT_MATCHED.write_text('')

    if unmatched_rows:
        with OUT_UNMATCHED.open('w', newline='') as handle:
            writer = csv.DictWriter(handle, fieldnames=list(unmatched_rows[0].keys()))
            writer.writeheader()
            writer.writerows(unmatched_rows)
    else:
        OUT_UNMATCHED.write_text('')

    summary_lines = [
        '# 05 commission comment reconciliation (auto match)',
        f'- Total comment rows: {len(comment_rows)}',
        f'- Marked delete (strike): {sum(1 for r in comment_rows if r.is_strike)}',
        f'- Marked update (red): {sum(1 for r in comment_rows if r.is_red and not r.is_strike)}',
        f'- Matched: {len(matched_rows)}',
        f'- Unmatched: {len(unmatched_rows)}',
        '',
        'Outputs:',
        f'- {OUT_MATCHED}',
        f'- {OUT_UNMATCHED}',
    ]
    OUT_SUMMARY.write_text('\n'.join(summary_lines))


if __name__ == '__main__':
    main()
