import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import load_workbook

ROOT = Path("/Users/ghilesmoussaoui/Desktop/Versotech")
DATA = ROOT / "verso_capital_2_data"
ENV = ROOT / ".env.local"

EXPORT_XLSX = DATA / "VC2_Client_Data_Export.xlsx"
DASH_XLSX = DATA / "VERSO DASHBOARD_V1.0.xlsx"

SCOPE_VC = ["VC201", "VC202", "VC203", "VC206", "VC207", "VC209", "VC210", "VC211", "VC215"]
SCOPE_WITH_VCL = set(SCOPE_VC + ["VCL001", "VCL002"])


def to_float(v):
    if v in (None, ""):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if not s:
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0


def map_vc(vc):
    # VC203 dashboard sheet contains rows whose Vehicle is VCL001/VCL002, but these are
    # distinct vehicles in the DB (VERSO Capital LLC), not part of VC203 (VERSO Capital 2 SCSp).
    return vc


def load_key():
    for ln in ENV.read_text().splitlines():
        if ln.startswith("SUPABASE_SERVICE_ROLE_KEY="):
            return ln.split("=", 1)[1].strip()
    raise RuntimeError("SUPABASE_SERVICE_ROLE_KEY missing")


def api_get(path: str, params: dict, key: str):
    url = f"https://kagzryotbbnusdcyvqei.supabase.co/rest/v1/{path}?{urlencode(params)}"
    req = Request(url, headers={"apikey": key, "Authorization": f"Bearer {key}"})
    with urlopen(req, timeout=120) as r:
        return json.loads(r.read().decode("utf-8"))


def find_col(headers, label):
    for idx, h in enumerate(headers, start=1):
        if isinstance(h, str) and h.strip().lower() == label.strip().lower():
            return idx
    return None


def collect_dashboard_totals():
    wb_dash = load_workbook(DASH_XLSX, data_only=True, read_only=True)
    totals = defaultdict(lambda: defaultdict(float))

    for sheet in wb_dash.sheetnames:
        if sheet not in set(SCOPE_VC):
            continue
        ws = wb_dash[sheet]
        headers = [ws.cell(2, c).value for c in range(1, 90)]
        col = {
            "vehicle": find_col(headers, "Vehicle"),
            "currency": find_col(headers, "Currency"),
            "amount": find_col(headers, "Amount invested"),
            "shares": find_col(headers, "Number of shares invested"),
            "ownership": find_col(headers, "OWNERSHIP POSITION"),
            "spread_fee": find_col(headers, "Spread PPS Fees"),
            "sub_fee": find_col(headers, "Subscription fees"),
            "bd_fee": find_col(headers, "BD fees"),
            "finra_fee": find_col(headers, "FINRA fees"),
        }
        required = ["vehicle", "amount", "shares", "ownership", "spread_fee", "sub_fee", "bd_fee", "finra_fee"]
        if not all(col[k] for k in required):
            continue

        for row in ws.iter_rows(min_row=3, max_col=90, values_only=True):
            vehicle_raw = row[col["vehicle"] - 1] if len(row) >= col["vehicle"] else None
            if vehicle_raw is None:
                continue
            vehicle_raw = str(vehicle_raw).strip()
            if vehicle_raw not in SCOPE_WITH_VCL:
                continue

            ownership = to_float(row[col["ownership"] - 1])
            if ownership <= 0:
                continue

            vc = map_vc(vehicle_raw)
            cur = "USD"
            if col["currency"] and len(row) >= col["currency"]:
                raw_cur = row[col["currency"] - 1]
                if raw_cur not in (None, ""):
                    cur = str(raw_cur).strip()

            key = (vc, cur)
            totals[key]["count"] += 1
            totals[key]["commitment"] += to_float(row[col["amount"] - 1])
            totals[key]["shares"] += to_float(row[col["shares"] - 1])
            totals[key]["ownership"] += ownership
            totals[key]["spread_fee"] += to_float(row[col["spread_fee"] - 1])
            totals[key]["subscription_fee"] += to_float(row[col["sub_fee"] - 1])
            totals[key]["bd_fee"] += to_float(row[col["bd_fee"] - 1])
            totals[key]["finra_fee"] += to_float(row[col["finra_fee"] - 1])
    return totals


def collect_db_totals():
    key = load_key()
    vehicles = api_get(
        "vehicles",
        {"select": "id,entity_code", "entity_code": "in.(" + ",".join([f'"{x}"' for x in SCOPE_VC]) + ")"},
        key,
    )
    vid_to_code = {v["id"]: v["entity_code"] for v in vehicles}
    vid_in = "(" + ",".join([f'"{x}"' for x in vid_to_code.keys()]) + ")"
    subs = api_get(
        "subscriptions",
        {
            "select": "vehicle_id,currency,commitment,num_shares,units,spread_fee_amount,subscription_fee_amount,bd_fee_amount,finra_fee_amount",
            "vehicle_id": "in." + vid_in,
        },
        key,
    )
    totals = defaultdict(lambda: defaultdict(float))
    for s in subs:
        vc = vid_to_code.get(s["vehicle_id"])
        if not vc:
            continue
        cur = (s.get("currency") or "USD").strip() if isinstance(s.get("currency"), str) else (s.get("currency") or "USD")
        key2 = (vc, cur)
        totals[key2]["count"] += 1
        totals[key2]["commitment"] += to_float(s.get("commitment"))
        totals[key2]["shares"] += to_float(s.get("num_shares"))
        totals[key2]["ownership"] += to_float(s.get("units"))
        totals[key2]["spread_fee"] += to_float(s.get("spread_fee_amount"))
        totals[key2]["subscription_fee"] += to_float(s.get("subscription_fee_amount"))
        totals[key2]["bd_fee"] += to_float(s.get("bd_fee_amount"))
        totals[key2]["finra_fee"] += to_float(s.get("finra_fee_amount"))
    return totals


def write_validation_sheet(dash_totals, db_totals):
    wb = load_workbook(EXPORT_XLSX)
    if "Validation" in wb.sheetnames:
        ws = wb["Validation"]
        wb.remove(ws)
    ws = wb.create_sheet("Validation")

    headers = [
        "vehicle",
        "currency",
        "dashboard_count",
        "db_count",
        "delta_count",
        "dashboard_commitment",
        "db_commitment",
        "delta_commitment",
        "dashboard_shares",
        "db_shares",
        "delta_shares",
        "dashboard_ownership",
        "db_ownership",
        "delta_ownership",
        "dashboard_spread_fee",
        "db_spread_fee",
        "delta_spread_fee",
        "dashboard_subscription_fee",
        "db_subscription_fee",
        "delta_subscription_fee",
        "dashboard_bd_fee",
        "db_bd_fee",
        "delta_bd_fee",
        "dashboard_finra_fee",
        "db_finra_fee",
        "delta_finra_fee",
    ]
    ws.append(headers)

    all_keys = sorted(set(dash_totals.keys()) | set(db_totals.keys()), key=lambda x: (x[0], x[1]))
    for vc, cur in all_keys:
        d = dash_totals[(vc, cur)]
        b = db_totals[(vc, cur)]
        ws.append(
            [
                vc,
                cur,
                int(d["count"]),
                int(b["count"]),
                int(b["count"] - d["count"]),
                d["commitment"],
                b["commitment"],
                b["commitment"] - d["commitment"],
                d["shares"],
                b["shares"],
                b["shares"] - d["shares"],
                d["ownership"],
                b["ownership"],
                b["ownership"] - d["ownership"],
                d["spread_fee"],
                b["spread_fee"],
                b["spread_fee"] - d["spread_fee"],
                d["subscription_fee"],
                b["subscription_fee"],
                b["subscription_fee"] - d["subscription_fee"],
                d["bd_fee"],
                b["bd_fee"],
                b["bd_fee"] - d["bd_fee"],
                d["finra_fee"],
                b["finra_fee"],
                b["finra_fee"] - d["finra_fee"],
            ]
        )

    for i, h in enumerate(headers, start=1):
        ws.cell(1, i).value = h
    ws.freeze_panes = "A2"
    wb.save(EXPORT_XLSX)


def print_summary(dash_totals, db_totals):
    all_keys = sorted(set(dash_totals.keys()) | set(db_totals.keys()), key=lambda x: (x[0], x[1]))
    mismatches = []
    for key in all_keys:
        d = dash_totals[key]
        b = db_totals[key]
        fields = ["count", "commitment", "shares", "ownership", "spread_fee", "subscription_fee", "bd_fee", "finra_fee"]
        deltas = {f: (b[f] - d[f]) for f in fields}
        if any(abs(v) > 1e-6 for v in deltas.values()):
            mismatches.append((key, deltas))
    print(f"validation_rows={len(all_keys)}")
    print(f"mismatches={len(mismatches)}")
    for (vc, cur), d in mismatches:
        print(vc, cur, d)


if __name__ == "__main__":
    dash = collect_dashboard_totals()
    db = collect_db_totals()
    write_validation_sheet(dash, db)
    print_summary(dash, db)
