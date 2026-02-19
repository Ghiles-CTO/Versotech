#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_RULES = ROOT / "verso_capital_2_data" / "audit_engine" / "rules_vc2.json"
DEFAULT_OUTDIR = ROOT / "verso_capital_2_data" / "audit_engine" / "output"
ENV_PATH = ROOT / ".env.local"
SUPABASE_URL = "https://kagzryotbbnusdcyvqei.supabase.co"


def to_float(v: Any) -> float:
    if v in (None, ""):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if not s:
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0


def parse_date(v: Any) -> str | None:
    if v in (None, ""):
        return None
    if hasattr(v, "date"):
        try:
            return v.date().isoformat()
        except Exception:
            pass
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            pass
    return s


def normalize_text(s: str) -> str:
    s = s.lower().strip()
    s = s.replace("&", " and ")
    s = s.replace("–", "-").replace("—", "-")
    s = re.sub(r"\b(mr|mrs|ms|dr|sir|madam|mme|m)\.?\b", " ", s)
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def name_key(s: str, aliases: dict[str, str]) -> str:
    n = normalize_text(s)
    parts = []
    for p in n.split(" "):
        if not p or p in {"and", "or"}:
            continue
        if len(p) > 4 and p.endswith("s"):
            p = p[:-1]
        parts.append(p)
    parts.sort()
    k = "".join(parts)
    return aliases.get(k, k)


def load_env(path: Path) -> dict[str, str]:
    out: dict[str, str] = {}
    if not path.exists():
        return out
    for raw in path.read_text().splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        out[k.strip()] = v.strip().strip('"')
    return out


def api_get_all(table: str, select: str, key: str, filters: dict[str, str] | None = None) -> list[dict[str, Any]]:
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Accept": "application/json",
    }
    out: list[dict[str, Any]] = []
    offset = 0
    limit = 1000
    while True:
        params = {"select": select, "order": "id", "limit": str(limit), "offset": str(offset)}
        if filters:
            params.update(filters)
        url = f"{SUPABASE_URL}/rest/v1/{table}?{urlencode(params)}"
        req = Request(url, headers=headers)
        try:
            with urlopen(req, timeout=120) as r:
                payload = r.read().decode("utf-8")
        except Exception as exc:
            body = ""
            if hasattr(exc, "read"):
                try:
                    body = exc.read().decode("utf-8", errors="ignore")
                except Exception:
                    body = ""
            raise RuntimeError(f"API error for {table} offset={offset}: {body}") from exc
        rows = json.loads(payload)
        if not rows:
            break
        out.extend(rows)
        if len(rows) < limit:
            break
        offset += limit
    return out


def find_col(headers: list[Any], candidates: list[str]) -> int | None:
    # Pass 1: strict raw text match (preserves '%' distinction)
    raw_headers = [str(h).strip().lower() if isinstance(h, str) else "" for h in headers]
    raw_candidates = [str(c).strip().lower() for c in candidates]
    for i, h in enumerate(raw_headers, start=1):
        if h in raw_candidates:
            return i

    # Pass 2: normalized fallback for minor punctuation/spacing drift
    norm_headers = [normalize_text(h) if isinstance(h, str) else "" for h in headers]
    norm_candidates = [normalize_text(c) for c in candidates]
    for i, h in enumerate(norm_headers, start=1):
        if h in norm_candidates:
            return i
    return None


@dataclass
class DashRow:
    vehicle: str
    sheet: str
    row_num: int
    investor_name: str
    investor_key: str
    commitment: float
    shares: float
    ownership: float
    contract_date: str | None
    spread_fee: float
    sub_fee: float
    bd_fee: float
    finra_fee: float


@dataclass
class DbSub:
    id: str
    vehicle: str
    investor_id: str
    investor_name: str
    investor_key: str
    commitment: float
    shares: float
    units: float
    contract_date: str | None
    spread_fee: float
    sub_fee: float
    bd_fee: float
    finra_fee: float
    status: str
    funded_amount: float


class Auditor:
    def __init__(self, rules: dict[str, Any], outdir: Path):
        self.rules = rules
        self.outdir = outdir
        self.aliases = rules.get("name_aliases", {})
        self.failures: list[dict[str, str]] = []
        self.warnings: list[dict[str, str]] = []

    def add_failure(self, check: str, vehicle: str, details: str, row_ref: str = ""):
        self.failures.append(
            {"severity": "fail", "check": check, "vehicle": vehicle, "row_ref": row_ref, "details": details}
        )

    def add_warning(self, check: str, vehicle: str, details: str, row_ref: str = ""):
        self.warnings.append(
            {"severity": "warn", "check": check, "vehicle": vehicle, "row_ref": row_ref, "details": details}
        )

    def load_dashboard_rows(self) -> tuple[list[DashRow], list[DashRow], dict[str, dict[str, float]]]:
        dash_path = ROOT / self.rules["dashboard_files"]["main"]
        scope_sheets = set(self.rules["scope_dashboard_sheets"])
        scope_codes = set(self.rules["scope_vehicle_codes"])
        alias_map = self.rules.get("dashboard_vehicle_alias", {})

        wb = load_workbook(dash_path, data_only=True, read_only=True)
        active: list[DashRow] = []
        zero: list[DashRow] = []
        totals = defaultdict(lambda: defaultdict(float))

        for sheet in wb.sheetnames:
            if sheet not in scope_sheets:
                continue
            ws = wb[sheet]
            headers = [ws.cell(2, c).value for c in range(1, 120)]
            col = {
                "inv_last": find_col(headers, ["Investor Last Name"]),
                "inv_mid": find_col(headers, ["Investor Middle Name"]),
                "inv_first": find_col(headers, ["Investor First Name"]),
                "inv_entity": find_col(headers, ["Investor Entity"]),
                "vehicle": find_col(headers, ["Vehicle"]),
                "amount": find_col(headers, ["Amount invested"]),
                "shares": find_col(headers, ["Number of shares invested"]),
                "ownership": find_col(headers, ["OWNERSHIP POSITION"]),
                "date": find_col(headers, ["Contract Date"]),
                "spread_fee": find_col(headers, ["Spread PPS Fees"]),
                "sub_fee": find_col(headers, ["Subscription fees"]),
                "bd_fee": find_col(headers, ["BD fees"]),
                "finra_fee": find_col(headers, ["FINRA fees"])
            }
            required = ["vehicle", "amount", "shares", "ownership", "date", "spread_fee", "sub_fee", "bd_fee", "finra_fee"]
            if not all(col[k] for k in required):
                self.add_failure("dashboard_columns_missing", sheet, f"missing columns in sheet {sheet}")
                continue

            for r_idx, row in enumerate(ws.iter_rows(min_row=3, max_col=120, values_only=True), start=3):
                raw_v = row[col["vehicle"] - 1] if len(row) >= col["vehicle"] else None
                if raw_v in (None, ""):
                    continue
                raw_v = str(raw_v).strip()
                mapped_v = alias_map.get(raw_v, raw_v)
                if mapped_v not in scope_codes:
                    continue

                entity = row[col["inv_entity"] - 1] if col["inv_entity"] else None
                last = row[col["inv_last"] - 1] if col["inv_last"] else None
                middle = row[col["inv_mid"] - 1] if col["inv_mid"] else None
                first = row[col["inv_first"] - 1] if col["inv_first"] else None
                if entity not in (None, ""):
                    investor_name = str(entity).strip()
                else:
                    investor_name = " ".join([str(x).strip() for x in (first, middle, last) if x not in (None, "")]).strip()
                if not investor_name:
                    continue

                rec = DashRow(
                    vehicle=mapped_v,
                    sheet=sheet,
                    row_num=r_idx,
                    investor_name=investor_name,
                    investor_key=name_key(investor_name, self.aliases),
                    commitment=to_float(row[col["amount"] - 1]),
                    shares=to_float(row[col["shares"] - 1]),
                    ownership=to_float(row[col["ownership"] - 1]),
                    contract_date=parse_date(row[col["date"] - 1]),
                    spread_fee=to_float(row[col["spread_fee"] - 1]),
                    sub_fee=to_float(row[col["sub_fee"] - 1]),
                    bd_fee=to_float(row[col["bd_fee"] - 1]),
                    finra_fee=to_float(row[col["finra_fee"] - 1]),
                )
                if rec.ownership <= 0:
                    zero.append(rec)
                else:
                    active.append(rec)
                    t = totals[rec.vehicle]
                    t["count"] += 1
                    t["commitment"] += rec.commitment
                    t["shares"] += rec.shares
                    t["ownership"] += rec.ownership
                    t["spread_fee"] += rec.spread_fee
                    t["sub_fee"] += rec.sub_fee
                    t["bd_fee"] += rec.bd_fee
                    t["finra_fee"] += rec.finra_fee
        return active, zero, totals

    def load_db_data(self) -> dict[str, Any]:
        env = load_env(ENV_PATH)
        key = env.get("SUPABASE_SERVICE_ROLE_KEY")
        if not key:
            raise RuntimeError("SUPABASE_SERVICE_ROLE_KEY missing in .env.local")

        scope_codes = sorted(self.rules["scope_vehicle_codes"])
        v_filter = "in.(" + ",".join([f'"{x}"' for x in scope_codes]) + ")"
        vehicles = api_get_all("vehicles", "id,entity_code,name", key, {"entity_code": v_filter})
        vid_to_code = {v["id"]: v["entity_code"] for v in vehicles}
        vehicle_ids = list(vid_to_code.keys())
        if not vehicle_ids:
            raise RuntimeError("No scope vehicles found in DB")

        vid_filter = "in.(" + ",".join([f'"{x}"' for x in vehicle_ids]) + ")"
        subs_raw = api_get_all(
            "subscriptions",
            "id,investor_id,vehicle_id,commitment,num_shares,units,contract_date,spread_fee_amount,subscription_fee_amount,bd_fee_amount,finra_fee_amount,status,funded_amount",
            key,
            {"vehicle_id": vid_filter},
        )
        pos_raw = api_get_all("positions", "id,investor_id,vehicle_id,units,cost_basis", key, {"vehicle_id": vid_filter})
        deals_raw = api_get_all("deals", "id,vehicle_id,name", key, {"vehicle_id": vid_filter})
        deal_ids = [d["id"] for d in deals_raw]
        did_to_vehicle = {d["id"]: vid_to_code.get(d["vehicle_id"]) for d in deals_raw}

        intros = []
        comms = []
        if deal_ids:
            did_filter = "in.(" + ",".join([f'"{x}"' for x in deal_ids]) + ")"
            intros = api_get_all(
                "introductions",
                "id,introducer_id,prospect_investor_id,deal_id,status",
                key,
                {"deal_id": did_filter},
            )
            comms = api_get_all(
                "introducer_commissions",
                "id,introducer_id,deal_id,investor_id,introduction_id,basis_type,rate_bps,base_amount,accrual_amount,status,tier_number",
                key,
                {"deal_id": did_filter},
            )

        inv_ids = sorted(
            {
                x["investor_id"]
                for x in subs_raw
                if x.get("investor_id")
            }
            | {
                x["investor_id"]
                for x in comms
                if x.get("investor_id")
            }
            | {
                x["prospect_investor_id"]
                for x in intros
                if x.get("prospect_investor_id")
            }
        )
        investors = []
        if inv_ids:
            investors = api_get_all(
                "investors",
                "id,legal_name,display_name",
                key,
                {"id": "in.(" + ",".join([f'"{x}"' for x in inv_ids]) + ")"},
            )
        iid_to_name = {i["id"]: (i.get("legal_name") or i.get("display_name") or "") for i in investors}

        introducers = api_get_all("introducers", "id,legal_name,display_name,status,email", key)
        brokers = api_get_all("brokers", "id,legal_name,display_name,status,email,contact_name", key)

        subs: list[DbSub] = []
        sub_totals = defaultdict(lambda: defaultdict(float))
        sub_key_counts = Counter()
        for s in subs_raw:
            vc = vid_to_code.get(s["vehicle_id"], "")
            nm = iid_to_name.get(s["investor_id"], "")
            rec = DbSub(
                id=s["id"],
                vehicle=vc,
                investor_id=s["investor_id"],
                investor_name=nm,
                investor_key=name_key(nm, self.aliases),
                commitment=to_float(s.get("commitment")),
                shares=to_float(s.get("num_shares")),
                units=to_float(s.get("units")),
                contract_date=parse_date(s.get("contract_date")),
                spread_fee=to_float(s.get("spread_fee_amount")),
                sub_fee=to_float(s.get("subscription_fee_amount")),
                bd_fee=to_float(s.get("bd_fee_amount")),
                finra_fee=to_float(s.get("finra_fee_amount")),
                status=str(s.get("status") or ""),
                funded_amount=to_float(s.get("funded_amount")),
            )
            subs.append(rec)
            k = (rec.vehicle, round(rec.commitment, 2), round(rec.shares, 6), rec.contract_date or "")
            sub_key_counts[k] += 1
            t = sub_totals[rec.vehicle]
            t["count"] += 1
            t["commitment"] += rec.commitment
            t["shares"] += rec.shares
            t["ownership"] += rec.units
            t["spread_fee"] += rec.spread_fee
            t["sub_fee"] += rec.sub_fee
            t["bd_fee"] += rec.bd_fee
            t["finra_fee"] += rec.finra_fee

        pos_units_by_vehicle = defaultdict(float)
        pos_dup = Counter()
        zero_pos_rows = []
        for p in pos_raw:
            vc = vid_to_code.get(p["vehicle_id"], "")
            units = to_float(p.get("units"))
            pos_units_by_vehicle[vc] += units
            pos_dup[(p["investor_id"], p["vehicle_id"])] += 1
            if abs(units) < 1e-9:
                zero_pos_rows.append(p["id"])

        return {
            "vehicles": vehicles,
            "vid_to_code": vid_to_code,
            "did_to_vehicle": did_to_vehicle,
            "subs": subs,
            "sub_totals": sub_totals,
            "sub_key_counts": sub_key_counts,
            "positions_raw": pos_raw,
            "pos_units_by_vehicle": pos_units_by_vehicle,
            "pos_dup": pos_dup,
            "zero_pos_rows": zero_pos_rows,
            "introductions": intros,
            "commissions": comms,
            "introducers": introducers,
            "brokers": brokers,
            "iid_to_name": iid_to_name,
        }

    def run_checks(self):
        dash_active, dash_zero, dash_totals = self.load_dashboard_rows()
        db = self.load_db_data()

        scope_codes = self.rules["scope_vehicle_codes"]
        db_totals = db["sub_totals"]

        # vehicle totals parity
        metrics = ["count", "commitment", "shares", "ownership", "spread_fee", "sub_fee", "bd_fee", "finra_fee"]
        for vc in scope_codes:
            for m in metrics:
                dv = dash_totals[vc].get(m, 0.0)
                bv = db_totals[vc].get(m, 0.0)
                tol = 0 if m == "count" else 0.01
                if abs(bv - dv) > tol:
                    self.add_failure("vehicle_totals_mismatch", vc, f"{m}: dashboard={dv} db={bv} delta={bv-dv}")

        # row multiset parity by key
        dash_keys = Counter((r.vehicle, round(r.commitment, 2), round(r.shares, 6), r.contract_date or "") for r in dash_active)
        db_keys = db["sub_key_counts"]
        for k in sorted(set(dash_keys) | set(db_keys)):
            if dash_keys.get(k, 0) != db_keys.get(k, 0):
                self.add_failure("row_key_mismatch", k[0], f"key={k} dashboard={dash_keys.get(k,0)} db={db_keys.get(k,0)}")

        # row-level numeric parity by (vehicle, commitment, shares, contract_date)
        dash_by_key = defaultdict(lambda: {"spread_fee": 0.0, "sub_fee": 0.0, "bd_fee": 0.0, "finra_fee": 0.0, "ownership": 0.0})
        db_by_key = defaultdict(lambda: {"spread_fee": 0.0, "sub_fee": 0.0, "bd_fee": 0.0, "finra_fee": 0.0, "ownership": 0.0})
        for r in dash_active:
            k = (r.vehicle, round(r.commitment, 2), round(r.shares, 6), r.contract_date or "")
            dash_by_key[k]["spread_fee"] += r.spread_fee
            dash_by_key[k]["sub_fee"] += r.sub_fee
            dash_by_key[k]["bd_fee"] += r.bd_fee
            dash_by_key[k]["finra_fee"] += r.finra_fee
            dash_by_key[k]["ownership"] += r.ownership
        for s in db["subs"]:
            k = (s.vehicle, round(s.commitment, 2), round(s.shares, 6), s.contract_date or "")
            db_by_key[k]["spread_fee"] += s.spread_fee
            db_by_key[k]["sub_fee"] += s.sub_fee
            db_by_key[k]["bd_fee"] += s.bd_fee
            db_by_key[k]["finra_fee"] += s.finra_fee
            db_by_key[k]["ownership"] += s.units
        for k in sorted(set(dash_by_key) | set(db_by_key)):
            for m in ("spread_fee", "sub_fee", "bd_fee", "finra_fee", "ownership"):
                dv = dash_by_key.get(k, {}).get(m, 0.0)
                bv = db_by_key.get(k, {}).get(m, 0.0)
                if abs(bv - dv) > 0.01:
                    self.add_failure(
                        "row_numeric_mismatch",
                        k[0],
                        f"key={k} metric={m} dashboard={round(dv,6)} db={round(bv,6)} delta={round(bv-dv,6)}",
                    )

        # zero ownership exclusion
        if self.rules["checks"].get("zero_ownership_must_not_be_loaded", False):
            zero_keys = Counter((r.vehicle, round(r.commitment, 2), round(r.shares, 6), r.contract_date or "") for r in dash_zero)
            for k, cnt in zero_keys.items():
                db_cnt = db_keys.get(k, 0)
                if db_cnt > 0:
                    self.add_failure("zero_ownership_loaded", k[0], f"zero_key={k} dashboard_zero={cnt} db_rows={db_cnt}")

        # status/funded
        if self.rules["checks"].get("status_must_be_funded", False):
            for s in db["subs"]:
                if s.status.lower() != "funded":
                    self.add_failure("status_not_funded", s.vehicle, f"sub={s.id} status={s.status}")
        if self.rules["checks"].get("funded_amount_equals_commitment", False):
            for s in db["subs"]:
                if abs(s.funded_amount - s.commitment) > 0.01:
                    self.add_failure(
                        "funded_amount_mismatch",
                        s.vehicle,
                        f"sub={s.id} commitment={s.commitment} funded={s.funded_amount}",
                    )

        # positions
        if self.rules["checks"].get("positions_units_no_zero_rows", False):
            for pid in db["zero_pos_rows"]:
                self.add_failure("position_zero_units", "", f"position_id={pid}")
        if self.rules["checks"].get("position_unique_per_investor_vehicle", False):
            for (inv_id, veh_id), c in db["pos_dup"].items():
                if c > 1:
                    vc = db["vid_to_code"].get(veh_id, "")
                    self.add_failure("position_duplicate", vc, f"investor_id={inv_id} vehicle_id={veh_id} count={c}")
        for vc in scope_codes:
            db_units = db["pos_units_by_vehicle"].get(vc, 0.0)
            dash_units = dash_totals[vc].get("ownership", 0.0)
            if abs(db_units - dash_units) > 0.01:
                self.add_failure("position_vs_dashboard_ownership", vc, f"positions={db_units} dashboard={dash_units}")

        # commission duplicate exact
        if self.rules["checks"].get("commission_exact_duplicate_check", False):
            dup = Counter(
                (
                    c.get("introducer_id"),
                    c.get("deal_id"),
                    c.get("investor_id"),
                    c.get("introduction_id"),
                    c.get("basis_type"),
                    c.get("tier_number"),
                    c.get("rate_bps"),
                    str(c.get("base_amount")),
                    str(c.get("accrual_amount")),
                )
                for c in db["commissions"]
            )
            for k, c in dup.items():
                if c > 1:
                    did = k[1]
                    vc = db["did_to_vehicle"].get(did, "")
                    self.add_failure("commission_duplicate_exact", vc, f"count={c} key={k}")

        # commission -> introduction FK
        if self.rules["checks"].get("commission_fk_introduction_check", False):
            intro_ids = {x["id"] for x in db["introductions"]}
            for c in db["commissions"]:
                iid = c.get("introduction_id")
                if iid and iid not in intro_ids:
                    vc = db["did_to_vehicle"].get(c.get("deal_id"), "")
                    self.add_failure("commission_broken_introduction_fk", vc, f"commission_id={c.get('id')} intro_id={iid}")

        max_intro = int(self.rules["checks"].get("max_introducers_per_subscription", 0) or 0)
        if max_intro > 0:
            intro_set = defaultdict(set)
            for i in db["introductions"]:
                key_i = (i.get("deal_id"), i.get("prospect_investor_id"))
                intro_set[key_i].add(i.get("introducer_id"))
            for (deal_id, investor_id), ids in intro_set.items():
                if len(ids) > max_intro:
                    vc = db["did_to_vehicle"].get(deal_id, "")
                    inv_name = db["iid_to_name"].get(investor_id, investor_id)
                    self.add_failure(
                        "max_introducers_per_subscription_exceeded",
                        vc,
                        f"investor={inv_name} deal_id={deal_id} introducers={len(ids)} max={max_intro}",
                    )

        # broker rules
        broker_expected = set(self.rules.get("brokers_table_expected", []))
        broker_actual = {str(b.get("legal_name") or "").strip() for b in db["brokers"]}
        missing_brokers = sorted(broker_expected - broker_actual)
        extra_brokers = sorted(broker_actual - broker_expected)
        if missing_brokers:
            self.add_failure("brokers_missing", "", f"{missing_brokers}")
        if extra_brokers:
            self.add_warning("brokers_unexpected", "", f"{extra_brokers}")

        intro_by_id = {i["id"]: str(i.get("legal_name") or i.get("display_name") or "").strip() for i in db["introducers"]}
        forbidden_global = set(self.rules.get("broker_like_introducers_forbidden_global", []))
        forbidden_by_vehicle = self.rules.get("broker_like_introducers_forbidden_by_vehicle", {})
        warnings_only = set(self.rules.get("warnings_only_introducers", []))
        forbidden_master = set(self.rules.get("forbidden_in_introducers_master", []))
        forbidden_in_introductions = set(self.rules.get("forbidden_in_introductions_global", []))

        for i in db["introducers"]:
            iname = str(i.get("legal_name") or i.get("display_name") or "").strip()
            if iname in forbidden_master:
                self.add_failure("forbidden_name_in_introducers_master", "", iname)

        for i in db["introductions"]:
            iname = intro_by_id.get(i.get("introducer_id"), "")
            if iname in forbidden_in_introductions:
                vc = db["did_to_vehicle"].get(i.get("deal_id"), "")
                inv = db["iid_to_name"].get(i.get("prospect_investor_id"), i.get("prospect_investor_id"))
                self.add_failure("forbidden_name_in_introductions", vc, f"introducer={iname} investor={inv}")

        # comm aggregates by introducer + vehicle
        agg = defaultdict(lambda: {"rows": 0, "amount": 0.0})
        for c in db["commissions"]:
            iname = intro_by_id.get(c.get("introducer_id"), "")
            did = c.get("deal_id")
            vc = db["did_to_vehicle"].get(did, "")
            key = (vc, iname)
            agg[key]["rows"] += 1
            agg[key]["amount"] += to_float(c.get("accrual_amount"))

        for (vc, iname), v in sorted(agg.items()):
            if iname in warnings_only:
                self.add_warning(
                    "introducer_warning_only_present",
                    vc,
                    f"{iname} rows={v['rows']} amount={round(v['amount'],2)}",
                )
                continue
            if iname in forbidden_global:
                self.add_failure(
                    "forbidden_broker_name_in_introducer_commissions",
                    vc,
                    f"{iname} rows={v['rows']} amount={round(v['amount'],2)}",
                )
            if vc in forbidden_by_vehicle and iname in set(forbidden_by_vehicle[vc]):
                self.add_failure(
                    "forbidden_vehicle_broker_name_in_introducer_commissions",
                    vc,
                    f"{iname} rows={v['rows']} amount={round(v['amount'],2)}",
                )

        # Explicit forbidden pairs
        forbidden_pairs = self.rules.get("forbidden_commission_pairs", [])
        if forbidden_pairs:
            pair_set = {(x.get("vehicle"), x.get("introducer")) for x in forbidden_pairs}
            for (vc, iname), v in sorted(agg.items()):
                if (vc, iname) in pair_set and v["rows"] > 0:
                    self.add_failure(
                        "forbidden_commission_pair_present",
                        vc,
                        f"introducer={iname} rows={v['rows']} amount={round(v['amount'],2)}",
                    )

        # Specific documented amount/rate checks
        if self.rules["checks"].get("specific_rule_assertions", False):
            comm_by_tuple = defaultdict(lambda: {"amount": 0.0, "rates": set(), "rows": 0})
            for c in db["commissions"]:
                vc = db["did_to_vehicle"].get(c.get("deal_id"), "")
                iid = c.get("investor_id")
                inv_name = db["iid_to_name"].get(iid, "")
                inv_key = name_key(inv_name, self.aliases) if inv_name else ""
                iname = intro_by_id.get(c.get("introducer_id"), "")
                btype = str(c.get("basis_type") or "")
                k = (vc, inv_key, iname, btype)
                comm_by_tuple[k]["amount"] += to_float(c.get("accrual_amount"))
                comm_by_tuple[k]["rows"] += 1
                rb = c.get("rate_bps")
                if rb not in (None, ""):
                    try:
                        comm_by_tuple[k]["rates"].add(int(float(rb)))
                    except Exception:
                        pass

            for rule in self.rules.get("specific_commission_expectations", []):
                vc = rule["vehicle"]
                inv_key = name_key(rule["investor"], self.aliases)
                iname = rule["introducer"]
                btype = rule["basis_type"]
                expected_amount = float(rule["amount"])
                expected_rate = int(rule["rate_bps"])
                got = comm_by_tuple.get((vc, inv_key, iname, btype))
                if not got:
                    self.add_failure(
                        "specific_commission_missing",
                        vc,
                        f"investor={rule['investor']} introducer={iname} basis={btype}",
                    )
                    continue
                got_amount = round(got["amount"], 2)
                if abs(got_amount - round(expected_amount, 2)) > 0.01:
                    self.add_failure(
                        "specific_commission_amount_mismatch",
                        vc,
                        f"investor={rule['investor']} introducer={iname} expected={expected_amount} got={got_amount}",
                    )
                if expected_rate not in got["rates"]:
                    self.add_failure(
                        "specific_commission_rate_mismatch",
                        vc,
                        f"investor={rule['investor']} introducer={iname} expected_rate={expected_rate} got_rates={sorted(got['rates'])}",
                    )

            set_cap_rule = self.rules.get("set_cap_vc215_spread_only_check", {})
            if set_cap_rule:
                vc = set_cap_rule.get("vehicle", "")
                iname = set_cap_rule.get("introducer", "")
                if set_cap_rule.get("invested_amount_must_be_zero"):
                    # agg is across all basis types, so check directly from comm rows for invested_amount
                    invested_total = 0.0
                    for c in db["commissions"]:
                        c_vc = db["did_to_vehicle"].get(c.get("deal_id"), "")
                        c_iname = intro_by_id.get(c.get("introducer_id"), "")
                        if c_vc == vc and c_iname == iname and str(c.get("basis_type") or "") == "invested_amount":
                            invested_total += to_float(c.get("accrual_amount"))
                    if abs(invested_total) > 0.01:
                        self.add_failure(
                            "set_cap_vc215_invested_amount_not_zero",
                            vc,
                            f"introducer={iname} invested_amount_total={round(invested_total,2)}",
                        )

        # contacts file sanity checks (non-red VC2/VCL rows only)
        contacts_path = ROOT / self.rules["dashboard_files"]["contacts"]
        wb = load_workbook(contacts_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows_checked = 0
        red_rows = []
        scope_set = set(self.rules["scope_vehicle_codes"])
        intro_names = {normalize_text(str(i.get("legal_name") or "")) for i in db["introducers"]}
        broker_names = {normalize_text(str(b.get("legal_name") or "")) for b in db["brokers"]}

        for r in range(2, ws.max_row + 1):
            series = ws.cell(r, 1).value
            legal = ws.cell(r, 2).value
            role = ws.cell(r, 7).value
            intro1 = ws.cell(r, 8).value
            if series in (None, "") and legal in (None, ""):
                continue
            # simple red row detection on first 2 cells
            is_red = False
            for c in (ws.cell(r, 1), ws.cell(r, 2)):
                rgb = getattr(getattr(c.fill, "fgColor", None), "rgb", None)
                if isinstance(rgb, str) and rgb.upper().endswith("FF0000"):
                    is_red = True
            if is_red:
                red_rows.append(r)
                continue
            s = str(series).strip() if series not in (None, "") else ""
            if s and s not in scope_set:
                continue
            rows_checked += 1
            if intro1 not in (None, "", "-"):
                n = normalize_text(str(intro1))
                if n not in intro_names and n not in broker_names:
                    self.add_failure(
                        "contacts_intro_not_found_in_db_master",
                        s,
                        f"row={r} intro={intro1}",
                        f"contacts:{r}",
                    )
            if role not in (None, "") and str(role).strip().lower() == "broker" and intro1 not in (None, "", "-"):
                n = normalize_text(str(intro1))
                if n not in broker_names:
                    self.add_failure(
                        "contacts_role_broker_missing_in_brokers_table",
                        s,
                        f"row={r} broker={intro1}",
                        f"contacts:{r}",
                    )

        summary = {
            "rules_version": self.rules.get("version"),
            "scope_vehicle_codes": scope_codes,
            "dashboard_active_rows": len(dash_active),
            "dashboard_zero_rows": len(dash_zero),
            "db_subscriptions": len(db["subs"]),
            "db_positions": len(db["positions_raw"]),
            "db_introductions": len(db["introductions"]),
            "db_commissions": len(db["commissions"]),
            "contacts_checked_rows": rows_checked,
            "contacts_red_rows_excluded": red_rows,
            "fail_count": len(self.failures),
            "warning_count": len(self.warnings),
            "fail_by_check": dict(Counter(x["check"] for x in self.failures)),
            "warn_by_check": dict(Counter(x["check"] for x in self.warnings)),
        }
        return summary

    def write_outputs(self, summary: dict[str, Any]):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        run_dir = self.outdir / f"run_{ts}"
        run_dir.mkdir(parents=True, exist_ok=True)

        out_json = run_dir / "audit_report.json"
        out_csv = run_dir / "audit_issues.csv"
        out_md = run_dir / "audit_summary.md"

        report = {
            "summary": summary,
            "failures": self.failures,
            "warnings": self.warnings,
        }
        out_json.write_text(json.dumps(report, indent=2))

        with out_csv.open("w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["severity", "check", "vehicle", "row_ref", "details"])
            w.writeheader()
            for r in self.failures + self.warnings:
                w.writerow(r)

        lines = []
        lines.append("# VC2 Audit Summary")
        lines.append("")
        lines.append(f"- Rules version: `{summary['rules_version']}`")
        lines.append(f"- Dashboard active rows: `{summary['dashboard_active_rows']}`")
        lines.append(f"- Dashboard zero rows: `{summary['dashboard_zero_rows']}`")
        lines.append(f"- DB subscriptions: `{summary['db_subscriptions']}`")
        lines.append(f"- DB positions: `{summary['db_positions']}`")
        lines.append(f"- DB introductions: `{summary['db_introductions']}`")
        lines.append(f"- DB commissions: `{summary['db_commissions']}`")
        lines.append(f"- Contacts checked rows: `{summary['contacts_checked_rows']}`")
        lines.append(f"- Failures: `{summary['fail_count']}`")
        lines.append(f"- Warnings: `{summary['warning_count']}`")
        lines.append("")
        if summary["fail_count"] == 0:
            lines.append("## Result")
            lines.append("PASS (no hard failures)")
        else:
            lines.append("## Result")
            lines.append("FAIL")
            lines.append("")
            lines.append("## Failure breakdown")
            for k, v in sorted(summary["fail_by_check"].items(), key=lambda x: (-x[1], x[0])):
                lines.append(f"- {k}: `{v}`")
        if summary["warning_count"] > 0:
            lines.append("")
            lines.append("## Warning breakdown")
            for k, v in sorted(summary["warn_by_check"].items(), key=lambda x: (-x[1], x[0])):
                lines.append(f"- {k}: `{v}`")
        lines.append("")
        lines.append("## Artifacts")
        lines.append(f"- `{out_json}`")
        lines.append(f"- `{out_csv}`")
        lines.append(f"- `{out_md}`")
        out_md.write_text("\n".join(lines))
        return run_dir, out_json, out_csv, out_md


def parse_args():
    p = argparse.ArgumentParser(description="VC2 deterministic audit")
    p.add_argument("--rules", type=Path, default=DEFAULT_RULES, help="Path to rules JSON")
    p.add_argument("--outdir", type=Path, default=DEFAULT_OUTDIR, help="Output directory")
    return p.parse_args()


def main() -> int:
    args = parse_args()
    if not args.rules.exists():
        print(f"Rules file not found: {args.rules}")
        return 2
    rules = json.loads(args.rules.read_text())
    auditor = Auditor(rules, args.outdir)
    summary = auditor.run_checks()
    run_dir, out_json, out_csv, out_md = auditor.write_outputs(summary)
    print(f"RUN_DIR: {run_dir}")
    print(f"REPORT_JSON: {out_json}")
    print(f"REPORT_CSV: {out_csv}")
    print(f"REPORT_MD: {out_md}")
    print(f"FAIL_COUNT: {summary['fail_count']}")
    print(f"WARN_COUNT: {summary['warning_count']}")
    return 0 if summary["fail_count"] == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
