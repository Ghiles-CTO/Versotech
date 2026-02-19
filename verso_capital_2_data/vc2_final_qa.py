import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import load_workbook

ROOT = Path('/Users/ghilesmoussaoui/Desktop/Versotech')
DATA = ROOT / 'verso_capital_2_data'
DASH_XLSX = DATA / 'VERSO DASHBOARD_V1.0.xlsx'
CONTACT_XLSX = DATA / 'VERSO Capital 2 SCSp Emails and Contacts.xlsx'
ENV = ROOT / '.env.local'

OUT_MD = DATA / 'VC2_FINAL_QA_REPORT.md'
OUT_JSON = DATA / 'vc2_final_qa_report.json'
OUT_FAILS = DATA / 'vc2_final_qa_failures.csv'

SCOPE_VC = ['VC201', 'VC202', 'VC203', 'VC206', 'VC207', 'VC209', 'VC210', 'VC211', 'VC215']
SCOPE_WITH_VCL = set(SCOPE_VC + ['VCL001', 'VCL002'])
SCOPE_SHEETS = set(SCOPE_VC)


def load_key():
    for ln in ENV.read_text().splitlines():
        if ln.startswith('SUPABASE_SERVICE_ROLE_KEY='):
            return ln.split('=', 1)[1].strip()
    raise RuntimeError('SUPABASE_SERVICE_ROLE_KEY missing')


def api_get(path: str, params: dict, key: str):
    url = f"https://kagzryotbbnusdcyvqei.supabase.co/rest/v1/{path}?{urlencode(params)}"
    req = Request(url, headers={'apikey': key, 'Authorization': f'Bearer {key}'})
    with urlopen(req, timeout=120) as r:
        return json.loads(r.read().decode('utf-8'))


def to_float(v):
    if v in (None, ''):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(',', '')
    if s == '':
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0


def norm_name(s: str) -> str:
    if s is None:
        return ''
    s = str(s).lower().strip()
    s = s.replace('&', ' and ')
    s = s.replace('–', '-').replace('—', '-')
    s = re.sub(r'\b(mr|mrs|ms|dr)\.?\b', ' ', s)
    s = re.sub(r'[^a-z0-9\- ]+', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def token_sort_name(s: str) -> str:
    n = norm_name(s)
    toks = []
    for t in n.split(' '):
        if not t or t in {'and', 'or'}:
            continue
        if len(t) == 1:
            continue
        if len(t) > 4 and t.endswith('s'):
            t = t[:-1]
        toks.append(t)
    toks.sort()
    return ' '.join(toks)


def parse_date(v):
    if v in (None, ''):
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = str(v).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            pass
    return s


def map_vc(vc):
    # VC203 dashboard sheet contains rows whose Vehicle is VCL001/VCL002, but these are
    # distinct vehicles in the DB (VERSO Capital LLC), not part of VC203 (VERSO Capital 2 SCSp).
    return vc


# ------------------ load dashboard active rows ------------------
wb_dash = load_workbook(DASH_XLSX, data_only=True, read_only=True)
dash_rows = []
zero_rows = []
vehicle_totals_dash = defaultdict(lambda: defaultdict(float))


def find_col(headers, label):
    for idx, h in enumerate(headers, start=1):
        if isinstance(h, str) and h.strip().lower() == label.strip().lower():
            return idx
    return None

for sheet in wb_dash.sheetnames:
    if sheet not in SCOPE_SHEETS:
        continue
    ws = wb_dash[sheet]
    headers = [ws.cell(2, c).value for c in range(1, 80)]
    col = {
        'inv_last': find_col(headers, 'Investor Last Name'),
        'inv_mid': find_col(headers, 'Investor Middle Name'),
        'inv_first': find_col(headers, 'Investor First Name'),
        'inv_entity': find_col(headers, 'Investor Entity'),
        'vehicle': find_col(headers, 'Vehicle'),
        'amount': find_col(headers, 'Amount invested'),
        'shares': find_col(headers, 'Number of shares invested'),
        'ownership': find_col(headers, 'OWNERSHIP POSITION'),
        'date': find_col(headers, 'Contract Date'),
        'spread_fee': find_col(headers, 'Spread PPS Fees'),
        'sub_fee': find_col(headers, 'Subscription fees'),
        'bd_fee': find_col(headers, 'BD fees'),
        'finra_fee': find_col(headers, 'FINRA fees'),
    }
    # Required columns for numeric checks.
    if not all(col[k] for k in ['vehicle', 'amount', 'shares', 'ownership', 'date', 'spread_fee', 'sub_fee', 'bd_fee', 'finra_fee']):
        continue
    for r, row in enumerate(ws.iter_rows(min_row=3, max_col=65, values_only=True), start=3):
        vehicle = row[col['vehicle'] - 1] if len(row) >= col['vehicle'] else None
        if vehicle is None:
            continue
        vehicle = str(vehicle).strip()
        if vehicle not in SCOPE_WITH_VCL:
            continue

        amount = to_float(row[col['amount'] - 1])
        shares = to_float(row[col['shares'] - 1])
        ownership = to_float(row[col['ownership'] - 1])
        date = parse_date(row[col['date'] - 1])
        spread_fee = to_float(row[col['spread_fee'] - 1])
        sub_fee = to_float(row[col['sub_fee'] - 1])
        bd_fee = to_float(row[col['bd_fee'] - 1])
        finra_fee = to_float(row[col['finra_fee'] - 1])

        entity = row[col['inv_entity'] - 1] if col['inv_entity'] else None
        lname = row[col['inv_last'] - 1] if col['inv_last'] else None
        mname = row[col['inv_mid'] - 1] if col['inv_mid'] else None
        fname = row[col['inv_first'] - 1] if col['inv_first'] else None
        if entity and str(entity).strip():
            investor_name = str(entity).strip()
        else:
            investor_name = ' '.join([str(x).strip() for x in (fname, mname, lname) if x not in (None, '')]).strip()
        if not investor_name:
            continue

        rec = {
            'sheet': sheet,
            'row': r,
            'vehicle_raw': vehicle,
            'vehicle': map_vc(vehicle),
            'investor_name': investor_name,
            'investor_norm': norm_name(investor_name),
            'commitment': amount,
            'shares': shares,
            'ownership': ownership,
            'contract_date': date,
            'spread_fee': spread_fee,
            'sub_fee': sub_fee,
            'bd_fee': bd_fee,
            'finra_fee': finra_fee,
        }

        if ownership <= 0:
            zero_rows.append(rec)
            continue

        dash_rows.append(rec)
        v = rec['vehicle']
        vehicle_totals_dash[v]['count'] += 1
        vehicle_totals_dash[v]['commitment'] += amount
        vehicle_totals_dash[v]['shares'] += shares
        vehicle_totals_dash[v]['ownership'] += ownership
        vehicle_totals_dash[v]['spread_fee'] += spread_fee
        vehicle_totals_dash[v]['sub_fee'] += sub_fee
        vehicle_totals_dash[v]['bd_fee'] += bd_fee
        vehicle_totals_dash[v]['finra_fee'] += finra_fee

# multiset key for deterministic matching

def key4(rec):
    return (
        rec['vehicle'],
        round(rec['commitment'], 2),
        round(rec['shares'], 6),
        rec['contract_date'] or ''
    )


dash_key_counts = Counter(key4(r) for r in dash_rows)

# ------------------ load DB ------------------
key = load_key()
# Include VCL001/VCL002 vehicles as distinct vehicles (they appear inside the VC203 dashboard sheet).
vehicles = api_get(
    'vehicles',
    {
        'select': 'id,entity_code',
        'entity_code': 'in.(' + ','.join([f'"{x}"' for x in sorted(SCOPE_WITH_VCL)]) + ')',
    },
    key,
)
vid_to_code = {v['id']: v['entity_code'] for v in vehicles}
vid_in = '(' + ','.join([f'"{x}"' for x in vid_to_code.keys()]) + ')'

investors = api_get('investors', {'select': 'id,legal_name'}, key)
iid_to_name = {i['id']: i.get('legal_name') or '' for i in investors}

subs = api_get('subscriptions', {
    'select': 'id,investor_id,vehicle_id,commitment,num_shares,units,contract_date,spread_fee_amount,subscription_fee_amount,bd_fee_amount,finra_fee_amount,status,funded_amount,currency',
    'vehicle_id': 'in.' + vid_in,
}, key)

positions = api_get('positions', {
    'select': 'id,investor_id,vehicle_id,units,cost_basis,as_of_date',
    'vehicle_id': 'in.' + vid_in,
}, key)

introducers = api_get('introducers', {'select': 'id,legal_name,display_name,email,status'}, key)
intro_name_to_id = {norm_name(i.get('legal_name') or ''): i['id'] for i in introducers}

brokers = api_get('brokers', {'select': 'id,legal_name,display_name,email,status'}, key)
broker_name_to_id = {norm_name(b.get('legal_name') or ''): b['id'] for b in brokers}

# deal mapping for commissions/introductions
deals = api_get('deals', {'select': 'id,vehicle_id', 'vehicle_id': 'in.' + vid_in}, key)
did_to_vcode = {d['id']: vid_to_code.get(d['vehicle_id']) for d in deals}
if did_to_vcode:
    did_in = '(' + ','.join([f'"{x}"' for x in did_to_vcode.keys()]) + ')'
    intros = api_get('introductions', {'select': 'id,introducer_id,prospect_investor_id,deal_id,status'}, key)
    intros = [x for x in intros if x.get('deal_id') in did_to_vcode]
    comms = api_get('introducer_commissions', {'select': 'id,introducer_id,deal_id,investor_id,introduction_id,basis_type,rate_bps,base_amount,accrual_amount,status,tier_number,currency'}, key)
    comms = [c for c in comms if c.get('deal_id') in did_to_vcode]
else:
    intros, comms = [], []

# aggregate DB
vehicle_totals_db = defaultdict(lambda: defaultdict(float))
db_key_counts = Counter()
db_rows = []
for s in subs:
    vc = vid_to_code.get(s['vehicle_id'])
    if not vc:
        continue
    nm = iid_to_name.get(s['investor_id'], '')
    rec = {
        'id': s['id'],
        'vehicle': vc,
        'investor_name': nm,
        'investor_norm': norm_name(nm),
        'commitment': to_float(s.get('commitment')),
        'shares': to_float(s.get('num_shares')),
        'ownership': to_float(s.get('units')),
        'contract_date': parse_date(s.get('contract_date')),
        'spread_fee': to_float(s.get('spread_fee_amount')),
        'sub_fee': to_float(s.get('subscription_fee_amount')),
        'bd_fee': to_float(s.get('bd_fee_amount')),
        'finra_fee': to_float(s.get('finra_fee_amount')),
        'status': s.get('status'),
        'funded_amount': to_float(s.get('funded_amount')),
        'currency': s.get('currency'),
    }
    db_rows.append(rec)
    db_key_counts[key4(rec)] += 1

    v = vc
    vehicle_totals_db[v]['count'] += 1
    vehicle_totals_db[v]['commitment'] += rec['commitment']
    vehicle_totals_db[v]['shares'] += rec['shares']
    vehicle_totals_db[v]['ownership'] += rec['ownership']
    vehicle_totals_db[v]['spread_fee'] += rec['spread_fee']
    vehicle_totals_db[v]['sub_fee'] += rec['sub_fee']
    vehicle_totals_db[v]['bd_fee'] += rec['bd_fee']
    vehicle_totals_db[v]['finra_fee'] += rec['finra_fee']

# position totals by vehicle
vehicle_pos_units = defaultdict(float)
for p in positions:
    vc = vid_to_code.get(p['vehicle_id'])
    if vc:
        vehicle_pos_units[vc] += to_float(p.get('units'))

# ------------------ checks ------------------
fails = []

def add_fail(check, severity, details, vehicle='', row_ref=''):
    fails.append({
        'check': check,
        'severity': severity,
        'vehicle': vehicle,
        'row_ref': row_ref,
        'details': details,
    })

# 1) vehicle totals parity
metrics = ['count', 'commitment', 'shares', 'ownership', 'spread_fee', 'sub_fee', 'bd_fee', 'finra_fee']
for vc in SCOPE_VC:
    for m in metrics:
        dv = vehicle_totals_dash[vc].get(m, 0.0)
        bv = vehicle_totals_db[vc].get(m, 0.0)
        tol = 0.01 if m != 'count' else 0
        if abs(bv - dv) > tol:
            add_fail('vehicle_totals', 'high', f'{m} mismatch dash={dv} db={bv} delta={bv-dv}', vc)
    # ownership vs positions
    if abs(vehicle_pos_units.get(vc, 0.0) - vehicle_totals_dash[vc].get('ownership', 0.0)) > 0.01:
        add_fail('position_vs_dashboard_ownership', 'high',
                 f"position_units={vehicle_pos_units.get(vc,0.0)} dashboard_ownership={vehicle_totals_dash[vc].get('ownership',0.0)}",
                 vc)

# 2) row-key multiset parity
all_keys = set(dash_key_counts) | set(db_key_counts)
for k in sorted(all_keys):
    d = dash_key_counts.get(k, 0)
    b = db_key_counts.get(k, 0)
    if d != b:
        add_fail('row_key_multiset', 'high', f'key={k} dash_count={d} db_count={b}', k[0])

# 3) zero ownership rows must not exist in DB subscriptions/positions by key
zero_key_counts = Counter((map_vc(z['vehicle_raw']), round(z['commitment'],2), round(z['shares'],6), z['contract_date'] or '') for z in zero_rows)
for k, cnt in zero_key_counts.items():
    db_cnt = db_key_counts.get(k, 0)
    if db_cnt > 0:
        add_fail('zero_ownership_loaded', 'high', f'key={k} zero_rows={cnt} db_subs={db_cnt}', k[0])

# 4) funded/status rule
for r in db_rows:
    if (r['status'] or '').lower() != 'funded':
        add_fail('status_not_funded', 'medium', f"sub_id={r['id']} status={r['status']}", r['vehicle'])
    # committed=funded rule
    if abs(r['funded_amount'] - r['commitment']) > 0.01:
        add_fail('funded_amount_mismatch', 'medium', f"sub_id={r['id']} commitment={r['commitment']} funded_amount={r['funded_amount']}", r['vehicle'])

# 5) zero units in DB
for r in db_rows:
    if abs(r['ownership']) < 1e-9:
        add_fail('subscription_zero_units', 'high', f"sub_id={r['id']} has units=0", r['vehicle'])
for p in positions:
    if abs(to_float(p.get('units'))) < 1e-9:
        vc = vid_to_code.get(p['vehicle_id'], '')
        add_fail('position_zero_units', 'high', f"position_id={p['id']} has units=0", vc)

# 6) positions duplicate investor+vehicle check
pos_dup = Counter((p['investor_id'], p['vehicle_id']) for p in positions)
for k, c in pos_dup.items():
    if c > 1:
        vc = vid_to_code.get(k[1], '')
        add_fail('position_duplicate', 'high', f'investor_id={k[0]} vehicle_id={k[1]} count={c}', vc)

# 7) commission duplicate exact logical key
comm_dup = Counter((
    c['introducer_id'],
    c['deal_id'],
    c['investor_id'],
    c.get('introduction_id'),
    c.get('basis_type'),
    c.get('tier_number'),
    c.get('rate_bps'),
    str(c.get('base_amount')),
    str(c.get('accrual_amount'))
) for c in comms)
for k, c in comm_dup.items():
    if c > 1:
        vc = did_to_vcode.get(k[1], '')
        add_fail('commission_duplicate_exact', 'medium', f'key={k} count={c}', vc)

# 8) introduction + commission linkage check (for VC2)
intro_ids = {x['id'] for x in intros}
for c in comms:
    if c.get('introduction_id') and c['introduction_id'] not in intro_ids:
        vc = did_to_vcode.get(c['deal_id'], '')
        add_fail('commission_broken_introduction_fk', 'high', f"commission_id={c['id']} introduction_id={c.get('introduction_id')} missing", vc)

# 9) contacts file checks (red exclusion + names present)
wb_contacts = load_workbook(CONTACT_XLSX, data_only=True)
ws = wb_contacts[wb_contacts.sheetnames[0]]
red_rows = []
active_contact_rows = []
for r in range(2, ws.max_row + 1):
    series = ws.cell(r, 1).value
    legal = ws.cell(r, 2).value
    role = ws.cell(r, 7).value
    intro = ws.cell(r, 8).value
    if series in (None, '') and legal in (None, ''):
        continue
    # detect full row red by fill on first two important cells
    c1 = ws.cell(r, 1).fill.fgColor.rgb if ws.cell(r, 1).fill else None
    c2 = ws.cell(r, 2).fill.fgColor.rgb if ws.cell(r, 2).fill else None
    is_red = False
    for c in (c1, c2):
        if isinstance(c, str) and c.upper().endswith('FF0000'):
            is_red = True
    if is_red:
        red_rows.append(r)
        continue
    s = '' if series is None else str(series).strip()
    if s not in SCOPE_WITH_VCL:
        continue
    active_contact_rows.append((r, s, str(legal or '').strip(), str(role or '').strip(), str(intro or '').strip()))

# check all non-red contact investor names exist in dashboard names (normalized)
dash_names_by_vc = defaultdict(set)
dash_token_names_by_vc = defaultdict(set)
dash_token_sets_by_vc = defaultdict(list)
for r in dash_rows:
    dash_names_by_vc[r['vehicle']].add(r['investor_norm'])
    ts = token_sort_name(r['investor_name'])
    dash_token_names_by_vc[r['vehicle']].add(ts)
    dash_token_sets_by_vc[r['vehicle']].append(set(ts.split()))
for r in zero_rows:
    dash_names_by_vc[r['vehicle']].add(r['investor_norm'])
    ts = token_sort_name(r['investor_name'])
    dash_token_names_by_vc[r['vehicle']].add(ts)
    dash_token_sets_by_vc[r['vehicle']].append(set(ts.split()))

for r, s, legal, role, intro in active_contact_rows:
    vc = map_vc(s)
    n = norm_name(legal)
    ts = token_sort_name(legal)
    ok = n in dash_names_by_vc[vc] or ts in dash_token_names_by_vc[vc]
    if not ok and ts:
        tset = set(ts.split())
        for dset in dash_token_sets_by_vc[vc]:
            inter = len(tset & dset)
            # Allow middle-name expansions when core first/last tokens align.
            if inter >= 2 and (tset.issubset(dset) or dset.issubset(tset)):
                ok = True
                break
    if not ok:
        add_fail('contact_investor_not_in_dashboard', 'medium', f'row={r} series={s} legal_name={legal}', vc, f'contacts:{r}')

# introducer names existence check (contacts -> introducers or brokers)
for r, s, legal, role, intro in active_contact_rows:
    if not intro or intro == '-':
        continue
    ni = norm_name(intro)
    if ni not in intro_name_to_id and ni not in broker_name_to_id:
        add_fail('contact_introducer_missing_in_db_master', 'medium', f'row={r} series={s} intro={intro}', map_vc(s), f'contacts:{r}')

# ------------------ output ------------------
summary = {
    'scope_vehicles': SCOPE_VC,
    'dashboard_active_rows': len(dash_rows),
    'dashboard_zero_rows': len(zero_rows),
    'db_subscriptions': len(db_rows),
    'db_positions': len(positions),
    'db_introductions': len(intros),
    'db_commissions': len(comms),
    'contact_non_red_scope_rows': len(active_contact_rows),
    'contact_red_rows': red_rows,
    'fail_count': len(fails),
    'fail_by_check': dict(Counter(f['check'] for f in fails)),
}

OUT_JSON.write_text(json.dumps({'summary': summary, 'fails': fails}, indent=2))

with OUT_FAILS.open('w', newline='') as f:
    w = csv.DictWriter(f, fieldnames=['check', 'severity', 'vehicle', 'row_ref', 'details'])
    w.writeheader()
    w.writerows(fails)

lines = []
lines.append('# VC2 Final QA Report')
lines.append('')
lines.append(f"- Dashboard active rows: `{summary['dashboard_active_rows']}`")
lines.append(f"- Dashboard zero-ownership rows: `{summary['dashboard_zero_rows']}`")
lines.append(f"- DB subscriptions: `{summary['db_subscriptions']}`")
lines.append(f"- DB positions: `{summary['db_positions']}`")
lines.append(f"- DB introductions: `{summary['db_introductions']}`")
lines.append(f"- DB commissions: `{summary['db_commissions']}`")
lines.append(f"- Contact non-red VC2/VCL rows: `{summary['contact_non_red_scope_rows']}`")
lines.append(f"- Contact red rows excluded: `{','.join(map(str,summary['contact_red_rows'])) if summary['contact_red_rows'] else 'none'}`")
lines.append('')
if not fails:
    lines.append('## Result: PASS')
    lines.append('- No critical or medium anomalies detected under the defined VC2 rules.')
else:
    lines.append('## Result: FAIL')
    lines.append(f"- Total anomalies: `{len(fails)}`")
    by = Counter(f['check'] for f in fails)
    for k, v in sorted(by.items(), key=lambda x: (-x[1], x[0])):
        lines.append(f"- {k}: `{v}`")
    lines.append('')
    lines.append('## Top anomalies')
    for f in fails[:30]:
        lines.append(f"- [{f['severity']}] {f['check']} | {f['vehicle']} | {f['details']}")

lines.append('')
lines.append('## Artifacts')
lines.append(f"- `{OUT_JSON.name}`")
lines.append(f"- `{OUT_FAILS.name}`")
OUT_MD.write_text('\n'.join(lines))

print('WROTE', OUT_MD)
print('WROTE', OUT_JSON)
print('WROTE', OUT_FAILS)
print('FAIL_COUNT', len(fails))
