import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = Path("/Users/ghilesmoussaoui/Desktop/Versotech")
DATA = ROOT / "verso_capital_2_data"
ENV = ROOT / ".env.local"

DASHBOARD_XLSX = DATA / "VERSO DASHBOARD_V1.0.xlsx"
OUT_XLSX = DATA / "VC2_Subscriptions_Introducers.xlsx"

SCOPE_CODES = [
    "VC201",
    "VC202",
    "VC203",
    "VC206",
    "VC207",
    "VC209",
    "VC210",
    "VC211",
    "VC215",
    "VCL001",
    "VCL002",
]
SCOPE_DASHBOARD_SHEETS = ["VC201", "VC202", "VC203", "VC206", "VC207", "VC209", "VC210", "VC211", "VC215"]


def to_float(v):
    if v in (None, ""):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if not s:
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0


def parse_date(v):
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            pass
    return s


def find_col(headers, label):
    label_norm = label.strip().lower()
    for i, h in enumerate(headers, start=1):
        if isinstance(h, str) and h.strip().lower() == label_norm:
            return i
    return None


def get_service_key():
    for line in ENV.read_text().splitlines():
        if line.startswith("SUPABASE_SERVICE_ROLE_KEY="):
            return line.split("=", 1)[1].strip()
    raise RuntimeError("SUPABASE_SERVICE_ROLE_KEY missing in .env.local")


def api_get(path, params, key):
    url = f"https://kagzryotbbnusdcyvqei.supabase.co/rest/v1/{path}?{urlencode(params)}"
    req = Request(url, headers={"apikey": key, "Authorization": f"Bearer {key}"})
    with urlopen(req, timeout=120) as r:
        return json.loads(r.read().decode("utf-8"))


def fetch_db():
    key = get_service_key()

    vehicles = api_get(
        "vehicles",
        {
            "select": "id,entity_code,name",
            "entity_code": "in.(" + ",".join([f'"{x}"' for x in SCOPE_CODES]) + ")",
            "limit": 2000,
        },
        key,
    )
    vid_to_code = {v["id"]: v["entity_code"] for v in vehicles}
    vid_to_name = {v["id"]: v.get("name") for v in vehicles}
    if not vid_to_code:
        raise RuntimeError("No vehicles found for VC2 scope")

    vehicle_ids = "(" + ",".join([f'"{x}"' for x in vid_to_code]) + ")"
    deals = api_get(
        "deals",
        {"select": "id,vehicle_id,name", "vehicle_id": "in." + vehicle_ids, "limit": 2000},
        key,
    )
    deal_ids = [d["id"] for d in deals]
    if not deal_ids:
        raise RuntimeError("No deals found for VC2 scope")
    did_to_vehicle = {d["id"]: d["vehicle_id"] for d in deals}
    did_to_name = {d["id"]: d.get("name") for d in deals}

    deal_in = "(" + ",".join([f'"{x}"' for x in deal_ids]) + ")"
    subs = api_get(
        "subscriptions",
        {
            "select": ",".join(
                [
                    "id",
                    "investor_id",
                    "vehicle_id",
                    "deal_id",
                    "currency",
                    "status",
                    "contract_date",
                    "commitment",
                    "funded_amount",
                    "num_shares",
                    "units",
                    "price_per_share",
                    "cost_per_share",
                    "subscription_fee_percent",
                    "subscription_fee_amount",
                    "spread_per_share",
                    "spread_fee_amount",
                    "management_fee_percent",
                    "bd_fee_percent",
                    "bd_fee_amount",
                    "finra_shares",
                    "finra_fee_amount",
                    "performance_fee_tier1_percent",
                    "performance_fee_tier2_percent",
                    "opportunity_name",
                    "sourcing_contract_ref",
                ]
            ),
            "deal_id": "in." + deal_in,
            "limit": 10000,
        },
        key,
    )

    investor_ids = sorted({s["investor_id"] for s in subs if s.get("investor_id")})
    inv_in = "(" + ",".join([f'"{x}"' for x in investor_ids]) + ")" if investor_ids else "()"
    investors = (
        api_get(
            "investors",
            {
                "select": "id,legal_name,display_name,type,representative_name,email,first_name,middle_name,last_name",
                "id": "in." + inv_in,
                "limit": 10000,
            },
            key,
        )
        if investor_ids
        else []
    )
    iid_to_inv = {i["id"]: i for i in investors}

    positions = api_get(
        "positions",
        {"select": "id,investor_id,vehicle_id,units,cost_basis", "vehicle_id": "in." + vehicle_ids, "limit": 10000},
        key,
    )
    pos_units = {(p["vehicle_id"], p["investor_id"]): to_float(p.get("units")) for p in positions}

    introductions = api_get(
        "introductions",
        {
            "select": "id,deal_id,introducer_id,prospect_investor_id,status,introduced_at",
            "deal_id": "in." + deal_in,
            "limit": 20000,
        },
        key,
    )
    intro_ids = [i["id"] for i in introductions]
    intro_by_key = defaultdict(list)
    for i in introductions:
        intro_by_key[(i["deal_id"], i["prospect_investor_id"])].append(i)

    introducer_ids = sorted({i["introducer_id"] for i in introductions if i.get("introducer_id")})
    intro_in = "(" + ",".join([f'"{x}"' for x in introducer_ids]) + ")" if introducer_ids else "()"
    introducers = (
        api_get(
            "introducers",
            {"select": "id,legal_name,email,contact_name", "id": "in." + intro_in, "limit": 10000},
            key,
        )
        if introducer_ids
        else []
    )
    intr_by_id = {i["id"]: i for i in introducers}

    comm_by_intro = defaultdict(lambda: defaultdict(float))
    comm_all = []
    if intro_ids:
        intro_ids_in = "(" + ",".join([f'"{x}"' for x in intro_ids]) + ")"
        comms = api_get(
            "introducer_commissions",
            {
                "select": "introduction_id,basis_type,accrual_amount,deal_id,investor_id",
                "introduction_id": "in." + intro_ids_in,
                "limit": 50000,
            },
            key,
        )
        comm_all = comms
        for c in comms:
            iid = c.get("introduction_id")
            bt = c.get("basis_type")
            if iid and bt:
                comm_by_intro[iid][bt] += to_float(c.get("accrual_amount"))

    brokers = api_get("brokers", {"select": "legal_name", "limit": 10000}, key)
    broker_names = {str(b.get("legal_name") or "").strip() for b in brokers}

    # commitments used for intro commission allocation across multiple subscriptions per investor+deal
    deal_inv_commitment = defaultdict(float)
    for s in subs:
        deal_inv_commitment[(s["deal_id"], s["investor_id"])] += to_float(s.get("commitment"))

    return {
        "subs": subs,
        "vid_to_code": vid_to_code,
        "vid_to_name": vid_to_name,
        "did_to_name": did_to_name,
        "iid_to_inv": iid_to_inv,
        "pos_units": pos_units,
        "intro_by_key": intro_by_key,
        "intr_by_id": intr_by_id,
        "comm_by_intro": comm_by_intro,
        "comm_all": comm_all,
        "broker_names": broker_names,
        "deal_inv_commitment": deal_inv_commitment,
    }


def collect_dashboard_totals():
    wb = load_workbook(DASHBOARD_XLSX, data_only=True, read_only=True)
    totals = defaultdict(lambda: defaultdict(float))
    scope_set = set(SCOPE_CODES)

    for sheet in SCOPE_DASHBOARD_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        headers = [ws.cell(2, c).value for c in range(1, 90)]
        col = {
            "index": find_col(headers, "Index"),
            "vehicle": find_col(headers, "Vehicle"),
            "currency": find_col(headers, "Currency"),
            "amount": find_col(headers, "Amount invested"),
            "shares": find_col(headers, "Number of shares invested"),
            "ownership": find_col(headers, "OWNERSHIP POSITION"),
            "spread_fee": find_col(headers, "Spread PPS Fees"),
            "sub_fee": find_col(headers, "Subscription fees"),
            "bd_fee": find_col(headers, "BD fees"),
            "finra_fee": find_col(headers, "FINRA fees"),
        }
        req = ["vehicle", "amount", "shares", "ownership", "spread_fee", "sub_fee", "bd_fee", "finra_fee"]
        if not all(col[k] for k in req):
            continue

        seen_index = set()
        for row in ws.iter_rows(min_row=3, max_col=90, values_only=True):
            # Some sheets contain repeated row blocks with duplicated Index values.
            # Keep first occurrence per sheet/index to avoid double counting repeated blocks.
            if col["index"]:
                idx_val = row[col["index"] - 1]
                if idx_val not in (None, ""):
                    idx_key = str(idx_val).strip()
                    if idx_key in seen_index:
                        continue
                    seen_index.add(idx_key)

            raw_vehicle = row[col["vehicle"] - 1]
            if raw_vehicle in (None, ""):
                continue
            vc = str(raw_vehicle).strip()
            if vc not in scope_set:
                continue

            ownership = to_float(row[col["ownership"] - 1])
            if ownership <= 0:
                continue

            cur = "USD"
            if col["currency"]:
                c = row[col["currency"] - 1]
                if c not in (None, ""):
                    cur = str(c).strip()

            k = (vc, cur)
            totals[k]["count"] += 1
            totals[k]["commitment"] += to_float(row[col["amount"] - 1])
            totals[k]["shares"] += to_float(row[col["shares"] - 1])
            totals[k]["ownership"] += ownership
            totals[k]["spread_fee"] += to_float(row[col["spread_fee"] - 1])
            totals[k]["subscription_fee"] += to_float(row[col["sub_fee"] - 1])
            totals[k]["bd_fee"] += to_float(row[col["bd_fee"] - 1])
            totals[k]["finra_fee"] += to_float(row[col["finra_fee"] - 1])
    return totals


def build():
    db = fetch_db()
    subs = db["subs"]
    subs.sort(
        key=lambda s: (
            db["vid_to_code"].get(s["vehicle_id"], "ZZZ"),
            (db["iid_to_inv"].get(s["investor_id"], {}).get("legal_name") or ""),
            parse_date(s.get("contract_date")) or "",
            to_float(s.get("commitment")),
            str(s.get("id")),
        )
    )

    rows = []
    max_intro = 0
    file_totals = defaultdict(lambda: defaultdict(float))
    db_totals = defaultdict(lambda: defaultdict(float))
    sub_keys_present = set()

    for s in subs:
        vc = db["vid_to_code"].get(s["vehicle_id"])
        if vc not in SCOPE_CODES:
            continue

        inv = db["iid_to_inv"].get(s["investor_id"], {})
        key = (s["deal_id"], s["investor_id"])
        sub_keys_present.add(key)
        intros = db["intro_by_key"].get(key, [])

        intro_blocks = []
        total_commit = db["deal_inv_commitment"].get(key, 0.0)
        sub_commit = to_float(s.get("commitment"))
        weight = (sub_commit / total_commit) if total_commit > 0 else 0.0

        for i in intros:
            intr = db["intr_by_id"].get(i["introducer_id"], {})
            name = str(intr.get("legal_name") or "").strip()
            role = "Broker" if name in db["broker_names"] else "Introducer"
            ctot = db["comm_by_intro"].get(i["id"], {})
            intro_blocks.append(
                {
                    "role": role,
                    "name": name,
                    "email": intr.get("email"),
                    "contact": intr.get("contact_name"),
                    "invested_amount": round(to_float(ctot.get("invested_amount")) * weight, 2),
                    "spread": round(to_float(ctot.get("spread")) * weight, 2),
                    "performance_fee": round(to_float(ctot.get("performance_fee")) * weight, 2),
                }
            )

        intro_blocks.sort(key=lambda x: (0 if x["role"] == "Introducer" else 1, x["name"]))
        max_intro = max(max_intro, len(intro_blocks))

        rec = {
            "Subscription ID": s["id"],
            "Row Type": "subscription",
            "Vehicle": vc,
            "Vehicle Name": db["vid_to_name"].get(s["vehicle_id"]) or db["did_to_name"].get(s["deal_id"]),
            "Investor": inv.get("legal_name"),
            "Investor Display Name": inv.get("display_name") or inv.get("legal_name"),
            "Investor Type": inv.get("type"),
            "First Name": inv.get("first_name") if (inv.get("type") or "").lower() == "individual" else None,
            "Middle Name": inv.get("middle_name") if (inv.get("type") or "").lower() == "individual" else None,
            "Last Name": inv.get("last_name") if (inv.get("type") or "").lower() == "individual" else None,
            "Investor Contact": inv.get("representative_name"),
            "Investor Email": inv.get("email"),
            "Contract Date": parse_date(s.get("contract_date")),
            "Currency": s.get("currency") or "USD",
            "Subscription Status": s.get("status"),
            "Opportunity Name": s.get("opportunity_name"),
            "Sourcing Contract Ref": s.get("sourcing_contract_ref"),
            "Committed Amount": to_float(s.get("commitment")),
            "Funded Amount": to_float(s.get("funded_amount")),
            "Shares": to_float(s.get("num_shares")),
            "Ownership Units": to_float(s.get("units")),
            "Current Position": to_float(db["pos_units"].get((s["vehicle_id"], s["investor_id"]))),
            "Price Per Share": to_float(s.get("price_per_share")),
            "Cost Per Share": to_float(s.get("cost_per_share")),
            "Subscription Fee (%)": to_float(s.get("subscription_fee_percent")),
            "Subscription Fee Amount": to_float(s.get("subscription_fee_amount")),
            "Spread Per Share": to_float(s.get("spread_per_share")),
            "Spread Fee Amount": to_float(s.get("spread_fee_amount")),
            "Management Fee (%)": to_float(s.get("management_fee_percent")),
            "BD Fee (%)": to_float(s.get("bd_fee_percent")),
            "BD Fee Amount": to_float(s.get("bd_fee_amount")),
            "FINRA Shares": to_float(s.get("finra_shares")),
            "FINRA Fee Amount": to_float(s.get("finra_fee_amount")),
            "Performance Fee Tier1 (%)": to_float(s.get("performance_fee_tier1_percent")),
            "Performance Fee Tier2 (%)": to_float(s.get("performance_fee_tier2_percent")),
            "_intros": intro_blocks,
        }
        rows.append(rec)

        k = (rec["Vehicle"], rec["Currency"])
        file_totals[k]["count"] += 1
        file_totals[k]["commitment"] += rec["Committed Amount"]
        file_totals[k]["shares"] += rec["Shares"]
        file_totals[k]["ownership"] += rec["Ownership Units"]
        file_totals[k]["spread_fee"] += rec["Spread Fee Amount"]
        file_totals[k]["subscription_fee"] += rec["Subscription Fee Amount"]
        file_totals[k]["bd_fee"] += rec["BD Fee Amount"]
        file_totals[k]["finra_fee"] += rec["FINRA Fee Amount"]

        db_totals[k]["count"] += 1
        db_totals[k]["commitment"] += rec["Committed Amount"]
        db_totals[k]["shares"] += rec["Shares"]
        db_totals[k]["ownership"] += rec["Ownership Units"]
        db_totals[k]["spread_fee"] += rec["Spread Fee Amount"]
        db_totals[k]["subscription_fee"] += rec["Subscription Fee Amount"]
        db_totals[k]["bd_fee"] += rec["BD Fee Amount"]
        db_totals[k]["finra_fee"] += rec["FINRA Fee Amount"]

    # Add commission-only rows (investor+deal has commissions/introductions but no subscription row in scope).
    for key, intros in db["intro_by_key"].items():
        if key in sub_keys_present:
            continue
        deal_id, investor_id = key
        # Keep only VC2 scope deals.
        vc = None
        for s in subs:
            if s["deal_id"] == deal_id:
                vc = db["vid_to_code"].get(s["vehicle_id"])
                break
        if vc is None:
            # Fallback: detect from any introduction-linked commission in scope.
            vc_candidates = []
            for i in intros:
                intro_id = i.get("id")
                for c in db["comm_all"]:
                    if c.get("introduction_id") == intro_id:
                        # deal_id always in scope due fetch filter, so this lookup is safe.
                        for s in subs:
                            if s["deal_id"] == c.get("deal_id"):
                                vc_candidates.append(db["vid_to_code"].get(s["vehicle_id"]))
                        break
            vc = vc_candidates[0] if vc_candidates else None
        if vc not in SCOPE_CODES:
            continue

        inv = db["iid_to_inv"].get(investor_id, {})
        intro_blocks = []
        for i in intros:
            intr = db["intr_by_id"].get(i["introducer_id"], {})
            name = str(intr.get("legal_name") or "").strip()
            role = "Broker" if name in db["broker_names"] else "Introducer"
            ctot = db["comm_by_intro"].get(i.get("id"), {})
            intro_blocks.append(
                {
                    "role": role,
                    "name": name,
                    "email": intr.get("email"),
                    "contact": intr.get("contact_name"),
                    "invested_amount": round(to_float(ctot.get("invested_amount")), 2),
                    "spread": round(to_float(ctot.get("spread")), 2),
                    "performance_fee": round(to_float(ctot.get("performance_fee")), 2),
                }
            )
        if not intro_blocks:
            continue
        intro_blocks.sort(key=lambda x: (0 if x["role"] == "Introducer" else 1, x["name"]))
        max_intro = max(max_intro, len(intro_blocks))

        rec = {
            "Subscription ID": None,
            "Row Type": "commission_only",
            "Vehicle": vc,
            "Vehicle Name": None,
            "Investor": inv.get("legal_name"),
            "Investor Display Name": inv.get("display_name") or inv.get("legal_name"),
            "Investor Type": inv.get("type"),
            "First Name": inv.get("first_name") if (inv.get("type") or "").lower() == "individual" else None,
            "Middle Name": inv.get("middle_name") if (inv.get("type") or "").lower() == "individual" else None,
            "Last Name": inv.get("last_name") if (inv.get("type") or "").lower() == "individual" else None,
            "Investor Contact": inv.get("representative_name"),
            "Investor Email": inv.get("email"),
            "Contract Date": None,
            "Currency": "USD",
            "Subscription Status": None,
            "Opportunity Name": None,
            "Sourcing Contract Ref": None,
            "Committed Amount": 0.0,
            "Funded Amount": 0.0,
            "Shares": 0.0,
            "Ownership Units": 0.0,
            "Current Position": 0.0,
            "Price Per Share": 0.0,
            "Cost Per Share": 0.0,
            "Subscription Fee (%)": 0.0,
            "Subscription Fee Amount": 0.0,
            "Spread Per Share": 0.0,
            "Spread Fee Amount": 0.0,
            "Management Fee (%)": 0.0,
            "BD Fee (%)": 0.0,
            "BD Fee Amount": 0.0,
            "FINRA Shares": 0.0,
            "FINRA Fee Amount": 0.0,
            "Performance Fee Tier1 (%)": 0.0,
            "Performance Fee Tier2 (%)": 0.0,
            "_intros": intro_blocks,
        }
        rows.append(rec)

    wb = Workbook()
    ws = wb.active
    ws.title = "Subscriptions_Review"

    base_headers = [
        "Subscription ID",
        "Row Type",
        "Vehicle",
        "Vehicle Name",
        "Investor",
        "Investor Display Name",
        "Investor Type",
        "First Name",
        "Middle Name",
        "Last Name",
        "Investor Contact",
        "Investor Email",
        "Contract Date",
        "Currency",
        "Subscription Status",
        "Opportunity Name",
        "Sourcing Contract Ref",
        "Committed Amount",
        "Funded Amount",
        "Shares",
        "Ownership Units",
        "Current Position",
        "Price Per Share",
        "Cost Per Share",
        "Subscription Fee (%)",
        "Subscription Fee Amount",
        "Spread Per Share",
        "Spread Fee Amount",
        "Management Fee (%)",
        "BD Fee (%)",
        "BD Fee Amount",
        "FINRA Shares",
        "FINRA Fee Amount",
        "Performance Fee Tier1 (%)",
        "Performance Fee Tier2 (%)",
    ]
    intro_headers = []
    for n in range(1, max_intro + 1):
        intro_headers.extend(
            [
                f"Introducer {n} Role",
                f"Introducer {n} Name",
                f"Introducer {n} Email",
                f"Introducer {n} Contact",
                f"Introducer {n} Commission - Invested Amount",
                f"Introducer {n} Commission - Spread",
                f"Introducer {n} Commission - Performance Fee",
            ]
        )
    headers = base_headers + intro_headers

    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).font = Font(bold=True)
    ws.freeze_panes = "A2"

    for rec in rows:
        row = [rec[h] for h in base_headers]
        intros = rec["_intros"]
        for n in range(max_intro):
            if n < len(intros):
                b = intros[n]
                row.extend(
                    [
                        b["role"],
                        b["name"],
                        b["email"],
                        b["contact"],
                        b["invested_amount"],
                        b["spread"],
                        b["performance_fee"],
                    ]
                )
            else:
                row.extend([None, None, None, None, None, None, None])
        ws.append(row)

    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 34
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["L"].width = 30
    ws.column_dimensions["P"].width = 26
    ws.column_dimensions["Q"].width = 28

    # 3-way totals sheet
    ws2 = wb.create_sheet("Totals_3Way_Check")
    h2 = [
        "vehicle",
        "currency",
        "dashboard_count",
        "db_count",
        "file_count",
        "dashboard_commitment",
        "db_commitment",
        "file_commitment",
        "dashboard_shares",
        "db_shares",
        "file_shares",
        "dashboard_ownership",
        "db_ownership",
        "file_ownership",
        "dashboard_spread_fee",
        "db_spread_fee",
        "file_spread_fee",
        "dashboard_subscription_fee",
        "db_subscription_fee",
        "file_subscription_fee",
        "dashboard_bd_fee",
        "db_bd_fee",
        "file_bd_fee",
        "dashboard_finra_fee",
        "db_finra_fee",
        "file_finra_fee",
        "delta_db_minus_dashboard_commitment",
        "delta_file_minus_db_commitment",
        "delta_db_minus_dashboard_shares",
        "delta_file_minus_db_shares",
        "delta_db_minus_dashboard_ownership",
        "delta_file_minus_db_ownership",
        "delta_db_minus_dashboard_spread_fee",
        "delta_file_minus_db_spread_fee",
    ]
    ws2.append(h2)
    for c in range(1, len(h2) + 1):
        ws2.cell(1, c).font = Font(bold=True)
    ws2.freeze_panes = "A2"

    dash = collect_dashboard_totals()
    all_keys = sorted(set(dash.keys()) | set(db_totals.keys()) | set(file_totals.keys()), key=lambda x: (x[0], x[1]))
    mismatch_rows = 0
    for k in all_keys:
        d = dash[k]
        b = db_totals[k]
        f = file_totals[k]
        ws2.append(
            [
                k[0],
                k[1],
                int(d["count"]),
                int(b["count"]),
                int(f["count"]),
                d["commitment"],
                b["commitment"],
                f["commitment"],
                d["shares"],
                b["shares"],
                f["shares"],
                d["ownership"],
                b["ownership"],
                f["ownership"],
                d["spread_fee"],
                b["spread_fee"],
                f["spread_fee"],
                d["subscription_fee"],
                b["subscription_fee"],
                f["subscription_fee"],
                d["bd_fee"],
                b["bd_fee"],
                f["bd_fee"],
                d["finra_fee"],
                b["finra_fee"],
                f["finra_fee"],
                b["commitment"] - d["commitment"],
                f["commitment"] - b["commitment"],
                b["shares"] - d["shares"],
                f["shares"] - b["shares"],
                b["ownership"] - d["ownership"],
                f["ownership"] - b["ownership"],
                b["spread_fee"] - d["spread_fee"],
                f["spread_fee"] - b["spread_fee"],
            ]
        )
        checks = [
            abs(b["count"] - d["count"]),
            abs(f["count"] - b["count"]),
            abs(b["commitment"] - d["commitment"]),
            abs(f["commitment"] - b["commitment"]),
            abs(b["shares"] - d["shares"]),
            abs(f["shares"] - b["shares"]),
            abs(b["ownership"] - d["ownership"]),
            abs(f["ownership"] - b["ownership"]),
            abs(b["spread_fee"] - d["spread_fee"]),
            abs(f["spread_fee"] - b["spread_fee"]),
            abs(b["subscription_fee"] - d["subscription_fee"]),
            abs(f["subscription_fee"] - b["subscription_fee"]),
            abs(b["bd_fee"] - d["bd_fee"]),
            abs(f["bd_fee"] - b["bd_fee"]),
            abs(b["finra_fee"] - d["finra_fee"]),
            abs(f["finra_fee"] - b["finra_fee"]),
        ]
        if any(x > 1e-5 for x in checks):
            mismatch_rows += 1

    ws3 = wb.create_sheet("Summary")
    ws3["A1"] = "VC2 Subscriptions + Introducers"
    ws3["A1"].font = Font(bold=True)
    ws3["A3"] = "Rows"
    ws3["B3"] = len(rows)
    ws3["A4"] = "Max introducers per subscription"
    ws3["B4"] = max_intro
    ws3["A5"] = "Totals mismatch rows"
    ws3["B5"] = mismatch_rows
    ws3["A6"] = "Generated at"
    ws3["B6"] = datetime.now().isoformat(timespec="seconds")
    ws3["A7"] = "Scope"
    ws3["B7"] = ", ".join(SCOPE_CODES)

    wb.save(OUT_XLSX)
    print("out_file", OUT_XLSX)
    print("rows", len(rows))
    print("max_intro", max_intro)
    print("totals_mismatch_rows", mismatch_rows)


if __name__ == "__main__":
    build()
