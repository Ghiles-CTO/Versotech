from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "verso_capital_2_data"


def norm_name(value: str | None) -> str:
    if not value:
        return ""
    return re.sub(r"[^a-z0-9]", "", value.lower())


def clean(value: str | None) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


ENTITY_WORDS = {
    "llc",
    "ltd",
    "limited",
    "inc",
    "lp",
    "l.p",
    "sarl",
    "company",
    "capital",
    "fund",
    "holdings",
    "holding",
    "trust",
    "group",
    "spv",
    "series",
    "pte",
    "pty",
    "dwc",
    "llp",
    "sa",
    "plc",
    "establishment",
    "advisory",
}


def looks_like_person(name: str) -> bool:
    n = clean(name)
    if not n:
        return False
    low = n.lower()
    if any(word in low for word in ENTITY_WORDS):
        return False
    tokens = re.findall(r"[A-Za-z][A-Za-z'.-]*", n)
    if len(tokens) < 2 or len(tokens) > 6:
        return False
    return True


def looks_like_email(value: str) -> bool:
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", clean(value)))


def clean_phone(value: str) -> str:
    v = clean(value)
    if not v:
        return ""
    if len(re.sub(r"[^0-9]", "", v)) < 6:
        return ""
    return v


@dataclass
class ContactRow:
    source: str
    priority: int
    legal_name: str
    norm: str
    email: str
    contact_person: str
    phone: str


def read_vc2_contacts(path: Path) -> list[ContactRow]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Sheet1"]
    out: list[ContactRow] = []
    for r in range(2, ws.max_row + 1):
        legal_name = clean(ws.cell(r, 2).value)
        row_type = clean(ws.cell(r, 3).value).lower()
        contact = clean(ws.cell(r, 4).value)
        email = clean(ws.cell(r, 5).value)
        if row_type not in {"entity", "individual"}:
            continue
        if not legal_name:
            continue
        out.append(
            ContactRow(
                source="VC2 Contacts",
                priority=1,
                legal_name=legal_name,
                norm=norm_name(legal_name),
                email=email if looks_like_email(email) else "",
                contact_person=contact if looks_like_person(contact) else "",
                phone="",
            )
        )
    return out


def read_fred_missing_emails(path: Path) -> list[ContactRow]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Missing Emails"]
    out: list[ContactRow] = []
    for r in range(2, ws.max_row + 1):
        rec_type = clean(ws.cell(r, 1).value).lower()
        if rec_type != "investor":
            continue
        legal_name = clean(ws.cell(r, 2).value)
        email = clean(ws.cell(r, 3).value)
        if not legal_name:
            continue
        out.append(
            ContactRow(
                source="Fred Missing Emails",
                priority=2,
                legal_name=legal_name,
                norm=norm_name(legal_name),
                email=email if looks_like_email(email) else "",
                contact_person="",
                phone="",
            )
        )
    return out


def read_julien_v2(path: Path) -> list[ContactRow]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Sheet1"]
    out: list[ContactRow] = []
    for r in range(2, ws.max_row + 1):
        legal_name = clean(ws.cell(r, 1).value)
        row_type = clean(ws.cell(r, 2).value).lower()
        email = clean(ws.cell(r, 4).value)
        name_col_10 = clean(ws.cell(r, 10).value)
        name_col_12 = clean(ws.cell(r, 12).value)
        contact = name_col_12 or name_col_10
        if row_type not in {"entity", "individual"}:
            continue
        if not legal_name:
            continue
        out.append(
            ContactRow(
                source="Julien V2",
                priority=3,
                legal_name=legal_name,
                norm=norm_name(legal_name),
                email=email if looks_like_email(email) else "",
                contact_person=contact if looks_like_person(contact) else "",
                phone="",
            )
        )
    return out


def read_vc_firms(path: Path) -> list[ContactRow]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Sheet1"]
    out: list[ContactRow] = []
    for r in range(2, ws.max_row + 1):
        email = clean(ws.cell(r, 1).value)
        corporate_name = clean(ws.cell(r, 2).value)
        first_name = clean(ws.cell(r, 3).value)
        last_name = clean(ws.cell(r, 4).value)
        contact = clean(f"{first_name} {last_name}")
        if not corporate_name:
            continue
        out.append(
            ContactRow(
                source="VC Firms DB",
                priority=4,
                legal_name=corporate_name,
                norm=norm_name(corporate_name),
                email=email if looks_like_email(email) else "",
                contact_person=contact if looks_like_person(contact) else "",
                phone="",
            )
        )
    return out


def read_monaco_fo(path: Path) -> list[ContactRow]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Sheet1"]
    out: list[ContactRow] = []
    for r in range(2, ws.max_row + 1):
        company = clean(ws.cell(r, 1).value)
        phone = clean_phone(clean(ws.cell(r, 2).value))
        name = clean(ws.cell(r, 3).value)
        email = clean(ws.cell(r, 4).value)
        if not company:
            continue
        out.append(
            ContactRow(
                source="Monaco FO",
                priority=5,
                legal_name=company,
                norm=norm_name(company),
                email=email if looks_like_email(email) else "",
                contact_person=name if looks_like_person(name) else "",
                phone=phone,
            )
        )
    return out


def first_non_empty(rows: Iterable[ContactRow], attr: str) -> str:
    for row in rows:
        val = getattr(row, attr)
        if val:
            return val
    return ""


def sql_escape(value: str) -> str:
    return value.replace("'", "''")


def main() -> None:
    rows: list[ContactRow] = []
    rows.extend(read_vc2_contacts(ROOT / "verso_capital_2_data" / "VERSO Capital 2 SCSp Emails and Contacts.xlsx"))
    rows.extend(read_fred_missing_emails(ROOT / "dashboardreconciliations" / "Missing_Emails_Report_Fred_Updates.xlsx"))
    rows.extend(read_julien_v2(ROOT / "VERSO" / "datafixing" / "Investors emails missing for Julien_V2.xlsx"))
    rows.extend(read_vc_firms(ROOT / "dashboardreconciliations" / "VC Firms and Email Database.xlsx"))
    rows.extend(read_monaco_fo(ROOT / "dashboardreconciliations" / "DB - MONACO FO.xlsx"))

    grouped: dict[str, list[ContactRow]] = {}
    for row in rows:
        if not row.norm:
            continue
        grouped.setdefault(row.norm, []).append(row)

    for k in grouped:
        grouped[k].sort(key=lambda x: x.priority)

    merged = []
    for norm, candidates in grouped.items():
        best = candidates[0]
        merged.append(
            {
                "norm": norm,
                "legal_name": best.legal_name,
                "email": first_non_empty(candidates, "email"),
                "contact_person": first_non_empty(candidates, "contact_person"),
                "phone": first_non_empty(candidates, "phone"),
                "source_chain": " | ".join(sorted(set(c.source for c in candidates))),
            }
        )

    merged.sort(key=lambda x: x["legal_name"].lower())

    csv_path = OUT_DIR / "investor_profile_backfill_sources_merged.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "norm",
                "legal_name",
                "email",
                "contact_person",
                "phone",
                "source_chain",
            ],
        )
        w.writeheader()
        w.writerows(merged)

    values_lines = []
    for r in merged:
        values_lines.append(
            "('{}','{}','{}','{}')".format(
                sql_escape(r["norm"]),
                sql_escape(r["email"]),
                sql_escape(r["contact_person"]),
                sql_escape(r["phone"]),
            )
        )

    sql_path = OUT_DIR / "investor_profile_backfill_updates.sql"
    values_block = ",\n    ".join(values_lines) if values_lines else "('','','','')"
    sql = f"""-- Generated by build_investor_profile_backfill.py
with src(norm_name, src_email, src_contact_person, src_phone) as (
  values
    {values_block}
),
src_clean as (
  select norm_name,
         nullif(trim(src_email), '') as src_email,
         nullif(trim(src_contact_person), '') as src_contact_person,
         nullif(trim(src_phone), '') as src_phone
  from src
),
investor_keys as (
  select i.*,
         regexp_replace(lower(coalesce(i.legal_name,'')), '[^a-z0-9]', '', 'g') as norm_name
  from investors i
),
unique_norms as (
  select norm_name
  from investor_keys
  where norm_name <> ''
  group by norm_name
  having count(*) = 1
),
matched as (
  select i.id,
         i.type,
         i.legal_name,
         i.display_name,
         i.email,
         i.phone,
         i.representative_name,
         s.src_email,
         s.src_contact_person,
         s.src_phone
  from investor_keys i
  join unique_norms u on u.norm_name = i.norm_name
  join src_clean s
    on i.norm_name = s.norm_name
)
update investors i
set
  display_name = case
    when coalesce(trim(i.display_name),'') = '' then i.legal_name
    else i.display_name
  end,
  email = case
    when coalesce(trim(i.email),'') = '' and m.src_email is not null then m.src_email
    else i.email
  end,
  phone = case
    when coalesce(trim(i.phone),'') = '' and m.src_phone is not null then m.src_phone
    else i.phone
  end,
  representative_name = case
    when i.type = 'entity'
      and coalesce(trim(i.representative_name),'') = ''
      and m.src_contact_person is not null
      then m.src_contact_person
    else i.representative_name
  end
from matched m
where i.id = m.id;

-- Fill display_name everywhere it is still blank.
update investors
set display_name = legal_name
where coalesce(trim(display_name), '') = ''
  and coalesce(trim(legal_name), '') <> '';

-- Safe individual name fill (single known parseable record).
update investors
set first_name = 'Michael',
    last_name = 'RYAN'
where id = '494ada77-4fb7-4bbd-a7f1-784ecc263338'
  and type = 'individual'
  and coalesce(trim(first_name),'') = ''
  and coalesce(trim(last_name),'') = '';
"""
    sql_path.write_text(sql, encoding="utf-8")

    print(f"rows_extracted={len(rows)}")
    print(f"rows_merged={len(merged)}")
    print(f"csv={csv_path}")
    print(f"sql={sql_path}")


if __name__ == "__main__":
    main()
