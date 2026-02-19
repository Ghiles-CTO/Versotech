import json
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

ROOT = Path("/Users/ghilesmoussaoui/Desktop/Versotech")
DATA = ROOT / "verso_capital_2_data"
ENV = ROOT / ".env.local"

SOURCE_EXPORT = DATA / "VC2_Client_Data_Export.xlsx"
DASH_XLSX = DATA / "VERSO DASHBOARD_V1.0.xlsx"
OUT_XLSX = DATA / "VC2_Client_Review.xlsx"

SCOPE_VC = ["VC201", "VC202", "VC203", "VC206", "VC207", "VC209", "VC210", "VC211", "VC215"]
SCOPE_WITH_VCL = set(SCOPE_VC + ["VCL001", "VCL002"])


def to_float(v):
    if v in (None, ""):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if not s:
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0


def parse_date(v):
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            pass
    return s


def map_vc(vc):
    # VC203 dashboard sheet contains rows whose Vehicle is VCL001/VCL002, but these are
    # distinct vehicles in the DB (VERSO Capital LLC), not part of VC203 (VERSO Capital 2 SCSp).
    return vc


def find_col(headers, label):
    for idx, h in enumerate(headers, start=1):
        if isinstance(h, str) and h.strip().lower() == label.strip().lower():
            return idx
    return None


def load_key():
    for ln in ENV.read_text().splitlines():
        if ln.startswith("SUPABASE_SERVICE_ROLE_KEY="):
            return ln.split("=", 1)[1].strip()
    raise RuntimeError("SUPABASE_SERVICE_ROLE_KEY missing")


def api_get(path: str, params: dict, key: str):
    url = f"https://kagzryotbbnusdcyvqei.supabase.co/rest/v1/{path}?{urlencode(params)}"
    req = Request(url, headers={"apikey": key, "Authorization": f"Bearer {key}"})
    with urlopen(req, timeout=120) as r:
        return json.loads(r.read().decode("utf-8"))


def collect_dashboard_totals():
    wb_dash = load_workbook(DASH_XLSX, data_only=True, read_only=True)
    totals = defaultdict(lambda: defaultdict(float))
    for sheet in wb_dash.sheetnames:
        if sheet not in set(SCOPE_VC):
            continue
        ws = wb_dash[sheet]
        headers = [ws.cell(2, c).value for c in range(1, 90)]
        col = {
            "vehicle": find_col(headers, "Vehicle"),
            "currency": find_col(headers, "Currency"),
            "amount": find_col(headers, "Amount invested"),
            "shares": find_col(headers, "Number of shares invested"),
            "ownership": find_col(headers, "OWNERSHIP POSITION"),
            "spread_fee": find_col(headers, "Spread PPS Fees"),
            "sub_fee": find_col(headers, "Subscription fees"),
            "bd_fee": find_col(headers, "BD fees"),
            "finra_fee": find_col(headers, "FINRA fees"),
        }
        required = ["vehicle", "amount", "shares", "ownership", "spread_fee", "sub_fee", "bd_fee", "finra_fee"]
        if not all(col[k] for k in required):
            continue

        for row in ws.iter_rows(min_row=3, max_col=90, values_only=True):
            vehicle_raw = row[col["vehicle"] - 1] if len(row) >= col["vehicle"] else None
            if vehicle_raw is None:
                continue
            vehicle_raw = str(vehicle_raw).strip()
            if vehicle_raw not in SCOPE_WITH_VCL:
                continue
            ownership = to_float(row[col["ownership"] - 1])
            if ownership <= 0:
                continue

            vc = map_vc(vehicle_raw)
            cur = "USD"
            if col["currency"] and len(row) >= col["currency"]:
                raw_cur = row[col["currency"] - 1]
                if raw_cur not in (None, ""):
                    cur = str(raw_cur).strip()

            key = (vc, cur)
            totals[key]["count"] += 1
            totals[key]["commitment"] += to_float(row[col["amount"] - 1])
            totals[key]["shares"] += to_float(row[col["shares"] - 1])
            totals[key]["ownership"] += ownership
            totals[key]["spread_fee"] += to_float(row[col["spread_fee"] - 1])
            totals[key]["subscription_fee"] += to_float(row[col["sub_fee"] - 1])
            totals[key]["bd_fee"] += to_float(row[col["bd_fee"] - 1])
            totals[key]["finra_fee"] += to_float(row[col["finra_fee"] - 1])
    return totals


def collect_db_totals():
    key = load_key()
    vehicles = api_get(
        "vehicles",
        {"select": "id,entity_code", "entity_code": "in.(" + ",".join([f'"{x}"' for x in SCOPE_VC]) + ")"},
        key,
    )
    vid_to_code = {v["id"]: v["entity_code"] for v in vehicles}
    vid_in = "(" + ",".join([f'"{x}"' for x in vid_to_code.keys()]) + ")"
    subs = api_get(
        "subscriptions",
        {
            "select": "vehicle_id,currency,commitment,num_shares,units,spread_fee_amount,subscription_fee_amount,bd_fee_amount,finra_fee_amount",
            "vehicle_id": "in." + vid_in,
        },
        key,
    )
    totals = defaultdict(lambda: defaultdict(float))
    for s in subs:
        vc = vid_to_code.get(s["vehicle_id"])
        if not vc:
            continue
        cur = (s.get("currency") or "USD").strip() if isinstance(s.get("currency"), str) else (s.get("currency") or "USD")
        key2 = (vc, cur)
        totals[key2]["count"] += 1
        totals[key2]["commitment"] += to_float(s.get("commitment"))
        totals[key2]["shares"] += to_float(s.get("num_shares"))
        totals[key2]["ownership"] += to_float(s.get("units"))
        totals[key2]["spread_fee"] += to_float(s.get("spread_fee_amount"))
        totals[key2]["subscription_fee"] += to_float(s.get("subscription_fee_amount"))
        totals[key2]["bd_fee"] += to_float(s.get("bd_fee_amount"))
        totals[key2]["finra_fee"] += to_float(s.get("finra_fee_amount"))
    return totals


def collect_investor_email_map():
    key = load_key()
    vehicles = api_get(
        "vehicles",
        {"select": "id,entity_code", "entity_code": "in.(" + ",".join([f'"{x}"' for x in SCOPE_VC]) + ")"},
        key,
    )
    vid_in = "(" + ",".join([f'"{x["id"]}"' for x in vehicles]) + ")"
    subs = api_get("subscriptions", {"select": "investor_id", "vehicle_id": "in." + vid_in}, key)
    investor_ids = sorted({s["investor_id"] for s in subs if s.get("investor_id")})
    if not investor_ids:
        return {}
    iid_in = "(" + ",".join([f'"{x}"' for x in investor_ids]) + ")"
    inv = api_get(
        "investors",
        {"select": "id,email,display_name,first_name,middle_name,last_name,type,legal_name", "id": "in." + iid_in},
        key,
    )
    out = {}
    for x in inv:
        out[x["id"]] = {
            "email": x.get("email"),
            "display_name": x.get("display_name"),
            "first_name": x.get("first_name"),
            "middle_name": x.get("middle_name"),
            "last_name": x.get("last_name"),
            "type": x.get("type"),
            "legal_name": x.get("legal_name"),
        }
    return out


def build_review():
    # Source rows from existing full export (already VC2-scoped with intros).
    wb_src = load_workbook(SOURCE_EXPORT, data_only=True)
    ws_src = wb_src["Subscriptions"]
    headers = [ws_src.cell(1, c).value for c in range(1, ws_src.max_column + 1)]
    hidx = {h: i + 1 for i, h in enumerate(headers) if isinstance(h, str)}

    # Dynamic introducer columns present in source.
    intro_nums = sorted(
        {
            int(m.group(1))
            for h in headers
            if isinstance(h, str)
            for m in [re.match(r"intro_(\d+)_role$", h)]
            if m
        }
    )

    rows = []
    used_intro_nums = set()
    investor_meta_by_id = collect_investor_email_map()

    base_cols = [
        "id",
        "vehicle_entity_code",
        "vehicle_name",
        "investor_legal_name",
        "investor_type",
        "investor_representative_name",
        "investor_email",
        "contract_date",
        "currency",
        "status",
        "opportunity_name",
        "sourcing_contract_ref",
        "commitment",
        "funded_amount",
        "num_shares",
        "units",
        "position_units",
        "price_per_share",
        "cost_per_share",
        "subscription_fee_percent",
        "subscription_fee_amount",
        "spread_per_share",
        "spread_fee_amount",
        "management_fee_percent",
        "bd_fee_percent",
        "bd_fee_amount",
        "finra_shares",
        "finra_fee_amount",
        "performance_fee_tier1_percent",
        "performance_fee_tier2_percent",
    ]

    for r in range(2, ws_src.max_row + 1):
        vc = ws_src.cell(r, hidx["vehicle_entity_code"]).value
        if vc not in SCOPE_VC:
            continue

        # base payload
        rec = {}
        for c in base_cols:
            rec[c] = ws_src.cell(r, hidx[c]).value if c in hidx else None

        # always use current DB investor metadata when available
        investor_id = ws_src.cell(r, hidx["investor_id"]).value if "investor_id" in hidx else None
        rec["investor_display_name"] = None
        rec["investor_first_name"] = None
        rec["investor_middle_name"] = None
        rec["investor_last_name"] = None
        if investor_id in investor_meta_by_id:
            meta = investor_meta_by_id[investor_id]
            if meta.get("email"):
                rec["investor_email"] = meta.get("email")
            rec["investor_display_name"] = meta.get("display_name") or rec["investor_legal_name"]
            inv_type = (meta.get("type") or rec.get("investor_type") or "").lower()
            if inv_type == "individual":
                rec["investor_first_name"] = meta.get("first_name")
                rec["investor_middle_name"] = meta.get("middle_name")
                rec["investor_last_name"] = meta.get("last_name")

        # dynamic intros
        for n in intro_nums:
            keys = {
                "role": f"intro_{n}_role",
                "name": f"intro_{n}_name",
                "email": f"intro_{n}_email",
                "contact": f"intro_{n}_contact",
                "commission_invested_amount": f"intro_{n}_commission_invested_amount",
                "commission_spread": f"intro_{n}_commission_spread",
                "commission_performance_fee": f"intro_{n}_commission_performance_fee",
            }
            block = {}
            has_data = False
            for k, src_col in keys.items():
                v = ws_src.cell(r, hidx[src_col]).value if src_col in hidx else None
                block[k] = v
                if v not in (None, "", 0):
                    has_data = True
            if has_data:
                used_intro_nums.add(n)
            rec[f"intro_{n}"] = block

        rows.append(rec)

    used_intro_nums = sorted(used_intro_nums)

    # Create output workbook.
    wb = Workbook()
    ws = wb.active
    ws.title = "Subscriptions_Review"

    out_headers = [
        "Subscription ID",
        "Vehicle",
        "Vehicle Name",
        "Investor",
        "Investor Display Name",
        "Investor Type",
        "First Name",
        "Middle Name",
        "Last Name",
        "Investor Contact",
        "Investor Email",
        "Contract Date",
        "Currency",
        "Subscription Status",
        "Opportunity Name",
        "Sourcing Contract Ref",
        "Committed Amount",
        "Funded Amount",
        "Shares",
        "Ownership Units",
        "Current Position",
        "Price Per Share",
        "Cost Per Share",
        "Subscription Fee (%)",
        "Subscription Fee Amount",
        "Spread Per Share",
        "Spread Fee Amount",
        "Management Fee (%)",
        "BD Fee (%)",
        "BD Fee Amount",
        "FINRA Shares",
        "FINRA Fee Amount",
        "Performance Fee Tier1 (%)",
        "Performance Fee Tier2 (%)",
    ]

    for n in used_intro_nums:
        out_headers.extend(
            [
                f"Introducer {n} Role",
                f"Introducer {n} Name",
                f"Introducer {n} Email",
                f"Introducer {n} Contact",
                f"Introducer {n} Commission - Invested Amount",
                f"Introducer {n} Commission - Spread",
                f"Introducer {n} Commission - Performance Fee",
            ]
        )

    ws.append(out_headers)
    for c in range(1, len(out_headers) + 1):
        ws.cell(1, c).font = Font(bold=True)
    ws.freeze_panes = "A2"

    for rec in rows:
        row = [
            rec["id"],
            rec["vehicle_entity_code"],
            rec["vehicle_name"],
            rec["investor_legal_name"],
            rec["investor_display_name"],
            rec["investor_type"],
            rec["investor_first_name"],
            rec["investor_middle_name"],
            rec["investor_last_name"],
            rec["investor_representative_name"],
            rec["investor_email"],
            rec["contract_date"],
            rec["currency"],
            rec["status"],
            rec["opportunity_name"],
            rec["sourcing_contract_ref"],
            rec["commitment"],
            rec["funded_amount"],
            rec["num_shares"],
            rec["units"],
            rec["position_units"],
            rec["price_per_share"],
            rec["cost_per_share"],
            rec["subscription_fee_percent"],
            rec["subscription_fee_amount"],
            rec["spread_per_share"],
            rec["spread_fee_amount"],
            rec["management_fee_percent"],
            rec["bd_fee_percent"],
            rec["bd_fee_amount"],
            rec["finra_shares"],
            rec["finra_fee_amount"],
            rec["performance_fee_tier1_percent"],
            rec["performance_fee_tier2_percent"],
        ]
        for n in used_intro_nums:
            b = rec.get(f"intro_{n}", {})
            row.extend(
                [
                    b.get("role"),
                    b.get("name"),
                    b.get("email"),
                    b.get("contact"),
                    b.get("commission_invested_amount"),
                    b.get("commission_spread"),
                    b.get("commission_performance_fee"),
                ]
            )
        ws.append(row)

    # Format widths
    widths = {
        1: 10,
        2: 24,
        3: 44,
        4: 12,
        5: 22,
        6: 34,
        7: 14,
        8: 10,
        9: 14,
    }
    for i in range(10, 22):
        widths[i] = 16
    start = 22
    for _ in used_intro_nums:
        widths[start] = 14
        widths[start + 1] = 34
        widths[start + 2] = 30
        widths[start + 3] = 24
        widths[start + 4] = 18
        widths[start + 5] = 18
        widths[start + 6] = 18
        start += 7
    for i, w in widths.items():
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = w
    # Better width handling for >26 columns
    from openpyxl.utils import get_column_letter

    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = w

    # 3-way totals sheet
    ws2 = wb.create_sheet("Totals_3Way_Check")
    hdr2 = [
        "vehicle",
        "currency",
        "dashboard_count",
        "db_count",
        "file_count",
        "dashboard_commitment",
        "db_commitment",
        "file_commitment",
        "dashboard_shares",
        "db_shares",
        "file_shares",
        "dashboard_ownership",
        "db_ownership",
        "file_ownership",
        "dashboard_spread_fee",
        "db_spread_fee",
        "file_spread_fee",
        "dashboard_subscription_fee",
        "db_subscription_fee",
        "file_subscription_fee",
        "dashboard_bd_fee",
        "db_bd_fee",
        "file_bd_fee",
        "dashboard_finra_fee",
        "db_finra_fee",
        "file_finra_fee",
        "delta_db_minus_dashboard_commitment",
        "delta_file_minus_db_commitment",
        "delta_db_minus_dashboard_shares",
        "delta_file_minus_db_shares",
        "delta_db_minus_dashboard_ownership",
        "delta_file_minus_db_ownership",
        "delta_db_minus_dashboard_spread_fee",
        "delta_file_minus_db_spread_fee",
    ]
    ws2.append(hdr2)
    for c in range(1, len(hdr2) + 1):
        ws2.cell(1, c).font = Font(bold=True)
    ws2.freeze_panes = "A2"

    dash = collect_dashboard_totals()
    db = collect_db_totals()

    # file totals from review rows
    file_totals = defaultdict(lambda: defaultdict(float))
    for rec in rows:
        vc = rec["vehicle_entity_code"]
        cur = rec["currency"] or "USD"
        key = (vc, cur)
        file_totals[key]["count"] += 1
        file_totals[key]["commitment"] += to_float(rec["commitment"])
        file_totals[key]["shares"] += to_float(rec["num_shares"])
        # Ownership parity must use subscription units (row-level ownership), not position_units
        # because positions are investor-level aggregates and can repeat across multiple subscriptions.
        file_totals[key]["ownership"] += to_float(rec["units"])
        file_totals[key]["spread_fee"] += to_float(rec["spread_fee_amount"])
        file_totals[key]["subscription_fee"] += to_float(rec["subscription_fee_amount"])
        file_totals[key]["bd_fee"] += to_float(rec["bd_fee_amount"])
        file_totals[key]["finra_fee"] += to_float(rec["finra_fee_amount"])

    all_keys = sorted(set(dash.keys()) | set(db.keys()) | set(file_totals.keys()), key=lambda x: (x[0], x[1]))
    for vc, cur in all_keys:
        d = dash[(vc, cur)]
        b = db[(vc, cur)]
        f = file_totals[(vc, cur)]
        ws2.append(
            [
                vc,
                cur,
                int(d["count"]),
                int(b["count"]),
                int(f["count"]),
                d["commitment"],
                b["commitment"],
                f["commitment"],
                d["shares"],
                b["shares"],
                f["shares"],
                d["ownership"],
                b["ownership"],
                f["ownership"],
                d["spread_fee"],
                b["spread_fee"],
                f["spread_fee"],
                d["subscription_fee"],
                b["subscription_fee"],
                f["subscription_fee"],
                d["bd_fee"],
                b["bd_fee"],
                f["bd_fee"],
                d["finra_fee"],
                b["finra_fee"],
                f["finra_fee"],
                b["commitment"] - d["commitment"],
                f["commitment"] - b["commitment"],
                b["shares"] - d["shares"],
                f["shares"] - b["shares"],
                b["ownership"] - d["ownership"],
                f["ownership"] - b["ownership"],
                b["spread_fee"] - d["spread_fee"],
                f["spread_fee"] - b["spread_fee"],
            ]
        )

    # Summary row for quick visual pass/fail
    ws3 = wb.create_sheet("Summary")
    ws3["A1"] = "VC2 Client Review Export"
    ws3["A1"].font = Font(bold=True)
    ws3["A3"] = "Subscriptions rows"
    ws3["B3"] = len(rows)
    ws3["A4"] = "Dynamic introducer columns used"
    ws3["B4"] = len(used_intro_nums)
    ws3["A5"] = "Max introducers per subscription"
    ws3["B5"] = max(used_intro_nums) if used_intro_nums else 0
    ws3["A7"] = "Built from"
    ws3["B7"] = "DB + dashboard + VC2 full export"
    ws3["A8"] = "Generated at"
    ws3["B8"] = datetime.now().isoformat(timespec="seconds")
    ws3["A9"] = "Email sync applied"
    ws3["B9"] = "Infinitas SPV VII email set from SPV V"

    wb.save(OUT_XLSX)

    # Console verification
    mismatch_rows = 0
    for vc, cur in all_keys:
        d = dash[(vc, cur)]
        b = db[(vc, cur)]
        f = file_totals[(vc, cur)]
        checks = [
            abs(b["count"] - d["count"]),
            abs(f["count"] - b["count"]),
            abs(b["commitment"] - d["commitment"]),
            abs(f["commitment"] - b["commitment"]),
            abs(b["shares"] - d["shares"]),
            abs(f["shares"] - b["shares"]),
            abs(b["ownership"] - d["ownership"]),
            abs(f["ownership"] - b["ownership"]),
            abs(b["spread_fee"] - d["spread_fee"]),
            abs(f["spread_fee"] - b["spread_fee"]),
            abs(b["subscription_fee"] - d["subscription_fee"]),
            abs(f["subscription_fee"] - b["subscription_fee"]),
            abs(b["bd_fee"] - d["bd_fee"]),
            abs(f["bd_fee"] - b["bd_fee"]),
            abs(b["finra_fee"] - d["finra_fee"]),
            abs(f["finra_fee"] - b["finra_fee"]),
        ]
        if any(x > 1e-5 for x in checks):
            mismatch_rows += 1
            print("MISMATCH", vc, cur, d, b, f)
    print("out_file", OUT_XLSX)
    print("subscriptions_rows", len(rows))
    print("introducer_columns_used", len(used_intro_nums))
    print("max_introducers", (max(used_intro_nums) if used_intro_nums else 0))
    print("totals_mismatch_rows", mismatch_rows)


if __name__ == "__main__":
    build_review()
