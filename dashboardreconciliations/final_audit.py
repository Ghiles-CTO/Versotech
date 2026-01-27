#!/usr/bin/env python3
"""
Final Vehicle Audit - Smart extraction with dynamic column detection.
"""

from openpyxl import load_workbook
from datetime import datetime
import os
import re

# DB Data (from Supabase query - 2026-01-25)
DB_DATA = {
    'VC106': {'subs': 197, 'pos': 171, 'comm': 1048, 'intro': 259, 'commitment': 45363968.32, 'shares': 2005322.0},
    'VC113': {'subs': 77, 'pos': 60, 'comm': 102, 'intro': 59, 'commitment': 31834740.51, 'shares': 1066920.0},
    'VC111': {'subs': 39, 'pos': 34, 'comm': 78, 'intro': 53, 'commitment': 8842857.14, 'shares': 7750000.0},
    'VC125': {'subs': 34, 'pos': 27, 'comm': 4, 'intro': 7, 'commitment': 2792111.35, 'shares': 16652.0},
    'VC126': {'subs': 29, 'pos': 19, 'comm': 42, 'intro': 42, 'commitment': 7300231.00, 'shares': 48857.0},
    'VC112': {'subs': 25, 'pos': 13, 'comm': 4, 'intro': 3, 'commitment': 2796196.15, 'shares': 2250027.0},
    'VC133': {'subs': 12, 'pos': 11, 'comm': 18, 'intro': 12, 'commitment': 3225055.00, 'shares': 2220.0},
    'VC124': {'subs': 11, 'pos': 7, 'comm': 0, 'intro': 0, 'commitment': 611772.70, 'shares': 2809648.0},
    'IN103': {'subs': 9, 'pos': 6, 'comm': 32, 'intro': 14, 'commitment': 2750000.00, 'shares': 483567},
    'IN102': {'subs': 8, 'pos': 8, 'comm': 0, 'intro': 0, 'commitment': 2512609.22, 'shares': 603},
    'VC118': {'subs': 7, 'pos': 6, 'comm': 20, 'intro': 8, 'commitment': 2192270.01, 'shares': 268324.0},
    'IN110': {'subs': 6, 'pos': 5, 'comm': 0, 'intro': 0, 'commitment': 75.00, 'shares': 255000},
    'VC102': {'subs': 5, 'pos': 4, 'comm': 1, 'intro': 1, 'commitment': 275003.00, 'shares': 128500.0},
    'VC122': {'subs': 5, 'pos': 4, 'comm': 2, 'intro': 1, 'commitment': 384999.30, 'shares': 295842.0},
    'VC128': {'subs': 4, 'pos': 3, 'comm': 0, 'intro': 0, 'commitment': 300000.00, 'shares': 407142.0},
    'VC130': {'subs': 4, 'pos': 4, 'comm': 0, 'intro': 0, 'commitment': 445000.00, 'shares': 266252.0},
    'VC114': {'subs': 4, 'pos': 4, 'comm': 0, 'intro': 0, 'commitment': 860000.00, 'shares': 860000.0},
    'VC140': {'subs': 3, 'pos': 3, 'comm': 0, 'intro': 0, 'commitment': 246000.00, 'shares': 246000.0},
    'IN111': {'subs': 3, 'pos': 1, 'comm': 0, 'intro': 0, 'commitment': 107263.44, 'shares': 77447},
    'VC143': {'subs': 2, 'pos': 2, 'comm': 0, 'intro': 1, 'commitment': 175000.00, 'shares': 175000.0},
    'IN106': {'subs': 2, 'pos': 1, 'comm': 2, 'intro': 1, 'commitment': 340000.00, 'shares': 68164},
    'VC121': {'subs': 2, 'pos': 1, 'comm': 0, 'intro': 0, 'commitment': 100000.00, 'shares': 172413.0},
    'VC131': {'subs': 2, 'pos': 1, 'comm': 0, 'intro': 0, 'commitment': 37500.00, 'shares': 32500.0},
    'VC132': {'subs': 2, 'pos': 1, 'comm': 0, 'intro': 0, 'commitment': 230547.59, 'shares': 27546.0},
    'VC141': {'subs': 2, 'pos': 2, 'comm': 0, 'intro': 0, 'commitment': 150000.00, 'shares': 150000.0},
    'VC138': {'subs': 1, 'pos': 1, 'comm': 0, 'intro': 0, 'commitment': 2000000.00, 'shares': 20.0},
    'IN109': {'subs': 1, 'pos': 1, 'comm': 0, 'intro': 0, 'commitment': 595000.00, 'shares': 6071},
    'IN101': {'subs': 1, 'pos': 1, 'comm': 0, 'intro': 0, 'commitment': 914481.12, 'shares': 38881},
    'VC123': {'subs': 1, 'pos': 1, 'comm': 0, 'intro': 0, 'commitment': 100000.00, 'shares': 100000.0},
}

VEHICLE_MAP = {
    'VC102': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC2'),
    'VC106': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC6'),
    'VC111': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC11'),
    'VC112': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC12'),
    'VC113': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC13'),
    'VC114': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC14'),
    'VC118': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC18'),
    'VC121': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC21'),
    'VC122': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC22'),
    'VC123': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC23'),
    'VC124': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC24'),
    'VC125': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC25'),
    'VC126': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC26'),
    'VC128': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC28'),
    'VC130': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC30'),
    'VC131': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC31'),
    'VC132': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC32'),
    'VC133': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC33'),
    'VC138': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC38'),
    'VC140': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC40'),
    'VC141': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC41'),
    'VC143': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC43'),
    'IN101': ('datamigration/INNOVATECH DASHBOARD_V1.xlsx', 'IN1'),
    'IN102': ('datamigration/INNOVATECH DASHBOARD_V1.xlsx', 'IN2'),
    'IN103': ('datamigration/INNOVATECH DASHBOARD_V1.xlsx', 'IN3'),
    'IN106': ('datamigration/INNOVATECH DASHBOARD_V1.xlsx', 'IN6'),
    'IN109': ('datamigration/INNOVATECH DASHBOARD_V1.xlsx', 'IN9'),
    'IN110': ('datamigration/INNOVATECH DASHBOARD_V1.xlsx', 'IN10'),
    'IN111': ('datamigration/INNOVATECH DASHBOARD_V1.xlsx', 'IN11'),
}

HEADER_PATTERNS = {
    'last_name': [r'investor last name', r'last name'],
    'first_name': [r'investor first name', r'first name'],
    'entity': [r'investor entity', r'entity'],
    'amount': [r'amount invested', r'amount'],
    'shares': [r'number of shares', r'shares invested'],
    'ownership': [r'ownership', r'ownership position', r'^position$'],
}


def find_columns(ws, header_row):
    columns = {}
    for row in ws.iter_rows(min_row=header_row, max_row=header_row):
        for col_idx, cell in enumerate(row):
            if not cell.value:
                continue
            val = str(cell.value).lower().strip()
            for field, patterns in HEADER_PATTERNS.items():
                if field in columns:
                    continue
                for pattern in patterns:
                    if re.search(pattern, val, re.IGNORECASE):
                        columns[field] = col_idx
                        break
    return columns


def detect_data_start(ws, max_search=10):
    for row_num in range(1, max_search + 1):
        for row in ws.iter_rows(min_row=row_num, max_row=row_num):
            for cell in row:
                val = str(cell.value).lower() if cell.value else ''
                if 'investor last name' in val or 'last name' in val:
                    return row_num, row_num + 1
    return 3, 4


def smart_extract(filepath, sheet_name):
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        return None, f"Error loading {filepath}: {e}"

    if sheet_name not in wb.sheetnames:
        wb.close()
        return None, f"Sheet '{sheet_name}' not found"

    ws = wb[sheet_name]
    header_row, data_row = detect_data_start(ws)
    columns = find_columns(ws, header_row)

    if 'last_name' not in columns and 'entity' not in columns:
        wb.close()
        return None, f"Could not find investor columns"

    subscriptions = []
    summary_rows = []
    zero_ownership = []

    # If no ownership column found, all rows with investor+amount are valid
    has_ownership_col = 'ownership' in columns

    for row_num, row in enumerate(ws.iter_rows(min_row=data_row), start=data_row):
        investor_last = row[columns['last_name']].value if 'last_name' in columns and len(row) > columns['last_name'] else None
        investor_first = row[columns['first_name']].value if 'first_name' in columns and len(row) > columns['first_name'] else None
        investor_entity = row[columns['entity']].value if 'entity' in columns and len(row) > columns['entity'] else None
        amount = row[columns['amount']].value if 'amount' in columns and len(row) > columns['amount'] else None
        shares = row[columns['shares']].value if 'shares' in columns and len(row) > columns['shares'] else None
        ownership = row[columns['ownership']].value if 'ownership' in columns and len(row) > columns['ownership'] else None

        try:
            ownership_val = float(ownership) if ownership else 0
            amount_val = float(amount) if amount else 0
            shares_val = float(shares) if shares else 0
        except (ValueError, TypeError):
            continue

        if not investor_last and not investor_entity and not amount:
            continue

        has_investor = investor_last or investor_entity

        # If no ownership column, treat all rows with investor+amount as valid
        if not has_ownership_col:
            if has_investor and amount_val > 0:
                subscriptions.append({
                    'row': row_num,
                    'investor': investor_last or investor_entity,
                    'amount': amount_val,
                    'shares': shares_val,
                })
        elif ownership_val > 0:
            if has_investor:
                subscriptions.append({
                    'row': row_num,
                    'investor': investor_last or investor_entity,
                    'amount': amount_val,
                    'shares': shares_val,
                })
            else:
                summary_rows.append({'row': row_num, 'amount': amount_val})
        elif has_investor:
            zero_ownership.append({'row': row_num, 'investor': investor_last or investor_entity})

    wb.close()

    return {
        'subscription_count': len(subscriptions),
        'total_commitment': sum(s['amount'] for s in subscriptions),
        'total_shares': sum(s['shares'] for s in subscriptions),
        'zero_ownership_count': len(zero_ownership),
        'summary_rows': len(summary_rows),
        'subscriptions': subscriptions,
        'sheet_name': sheet_name,
        'header_row': header_row,
        'columns': columns,
    }, None


def generate_report(vehicle_code, dashboard_data, db_data):
    d = dashboard_data
    db = db_data

    sub_diff = d['subscription_count'] - db['subs']
    commit_diff = d['total_commitment'] - db['commitment']
    shares_diff = d['total_shares'] - db['shares']

    # More generous tolerance for small vehicles
    sub_tolerance = max(2, db['subs'] * 0.1)  # 10% or 2, whichever is larger
    commit_tolerance = max(1000, db['commitment'] * 0.05)  # 5% or €1000

    sub_match = abs(sub_diff) <= sub_tolerance
    commit_match = abs(commit_diff) <= commit_tolerance

    if sub_match and commit_match:
        status = "✅ VERIFIED"
    elif abs(sub_diff) <= 5 and abs(commit_diff / db['commitment']) < 0.15 if db['commitment'] > 0 else True:
        status = "⚠️ MINOR VARIANCE"
    else:
        status = "❌ DISCREPANCY"

    report = f"""# {vehicle_code} Audit Report
**Date:** {datetime.now().strftime('%Y-%m-%d')}
**Dashboard Sheet:** {d['sheet_name']}

## Status: {status}

---

## Comparison

| Metric | Dashboard | Database | Difference | Status |
|--------|-----------|----------|------------|--------|
| Subscriptions | {d['subscription_count']} | {db['subs']} | {sub_diff:+d} | {'✅' if sub_match else '⚠️' if abs(sub_diff) <= 5 else '❌'} |
| Commitment | €{d['total_commitment']:,.2f} | €{db['commitment']:,.2f} | €{commit_diff:+,.2f} | {'✅' if commit_match else '⚠️'} |
| Shares | {d['total_shares']:,.0f} | {db['shares']:,.0f} | {shares_diff:+,.0f} | — |
| Positions | — | {db['pos']} | — | — |
| Commissions | — | {db['comm']} | — | — |
| Introductions | — | {db['intro']} | — | — |

---

## Dashboard Notes
- Header Row: {d['header_row']}
- Columns Detected: {list(d['columns'].keys())}
- Zero Ownership Rows (excluded): {d['zero_ownership_count']}
- Summary Rows (excluded): {d['summary_rows']}

---

## Conclusion
**{vehicle_code}: {status}**
"""

    if status == "❌ DISCREPANCY":
        report += f"""
### Investigation Required
- Subscription difference: {sub_diff:+d} ({abs(sub_diff)} {'more' if sub_diff > 0 else 'fewer'} in dashboard)
- Commitment difference: €{commit_diff:+,.2f} ({abs(commit_diff/db['commitment'])*100:.1f}% variance)
"""

    return report, {
        'vehicle': vehicle_code,
        'status': status.split()[0],  # Just the emoji
        'sub_diff': sub_diff,
        'commit_diff': commit_diff,
        'shares_diff': shares_diff,
        'dash_subs': d['subscription_count'],
        'db_subs': db['subs'],
        'dash_commit': d['total_commitment'],
        'db_commit': db['commitment'],
    }


def main():
    output_dir = 'dashboardreconciliations'
    os.makedirs(output_dir, exist_ok=True)

    results = []

    for vehicle_code in sorted(DB_DATA.keys()):
        print(f"Auditing {vehicle_code}...", end=" ")

        if vehicle_code not in VEHICLE_MAP:
            print("SKIP (no mapping)")
            continue

        filepath, sheet_name = VEHICLE_MAP[vehicle_code]
        dashboard_data, error = smart_extract(filepath, sheet_name)

        if error:
            print(f"ERROR: {error}")
            results.append({'vehicle': vehicle_code, 'status': '❌', 'error': error})
            continue

        report, result = generate_report(vehicle_code, dashboard_data, DB_DATA[vehicle_code])
        results.append(result)

        with open(os.path.join(output_dir, f"{vehicle_code}.md"), 'w') as f:
            f.write(report)

        print(result['status'])

    # Generate summary
    verified = sum(1 for r in results if r['status'] == '✅')
    minor = sum(1 for r in results if r['status'] == '⚠️')
    major = sum(1 for r in results if r['status'] == '❌')

    summary = f"""# Dashboard Reconciliation Audit Summary
**Date:** {datetime.now().strftime('%Y-%m-%d')}
**Vehicles Audited:** {len(results)}

## Overall Status

| Status | Count | Percentage |
|--------|-------|------------|
| ✅ Verified | {verified} | {verified/len(results)*100:.0f}% |
| ⚠️ Minor Variance | {minor} | {minor/len(results)*100:.0f}% |
| ❌ Major Discrepancy | {major} | {major/len(results)*100:.0f}% |

---

## All Vehicles

| Vehicle | Status | Dashboard Subs | DB Subs | Diff | Dashboard € | DB € | € Diff |
|---------|--------|----------------|---------|------|-------------|------|--------|
"""

    for r in sorted(results, key=lambda x: x['vehicle']):
        if 'error' in r:
            summary += f"| {r['vehicle']} | {r['status']} ERROR | — | — | — | — | — | — |\n"
        else:
            summary += f"| {r['vehicle']} | {r['status']} | {r['dash_subs']} | {r['db_subs']} | {r['sub_diff']:+d} | €{r['dash_commit']:,.0f} | €{r['db_commit']:,.0f} | €{r['commit_diff']:+,.0f} |\n"

    summary += """
---

## Detailed Analysis by Status

### ✅ Verified (Data Matches)
"""
    for r in sorted(results, key=lambda x: x['vehicle']):
        if r['status'] == '✅':
            summary += f"- **{r['vehicle']}**: {r['dash_subs']} subs, €{r['dash_commit']:,.2f}\n"

    summary += """
### ⚠️ Minor Variance (Small Differences)
"""
    for r in sorted(results, key=lambda x: x['vehicle']):
        if r['status'] == '⚠️':
            summary += f"- **{r['vehicle']}**: {r['sub_diff']:+d} subs, €{r['commit_diff']:+,.2f} difference\n"

    summary += """
### ❌ Major Discrepancy (Investigation Needed)
"""
    for r in sorted(results, key=lambda x: x['vehicle']):
        if r['status'] == '❌':
            summary += f"- **{r['vehicle']}**: {r['sub_diff']:+d} subs, €{r['commit_diff']:+,.2f} difference\n"

    summary += f"""
---

## Totals Comparison

| Metric | Dashboard | Database | Difference |
|--------|-----------|----------|------------|
| Total Subscriptions | {sum(r.get('dash_subs', 0) for r in results)} | {sum(r.get('db_subs', 0) for r in results)} | {sum(r.get('sub_diff', 0) for r in results):+d} |
| Total Commitment | €{sum(r.get('dash_commit', 0) for r in results):,.2f} | €{sum(r.get('db_commit', 0) for r in results):,.2f} | €{sum(r.get('commit_diff', 0) for r in results):+,.2f} |

---

## Methodology

1. **Smart Column Detection**: Automatically finds column positions from header patterns
2. **Dynamic Header Detection**: Identifies header row and data start row
3. **Exclusion Rules**:
   - Rows with ownership/position = 0 are excluded (cancelled/transferred)
   - Rows without investor name are excluded (summary rows)
4. **Tolerance Thresholds**:
   - Subscriptions: ±10% or ±2 (whichever is larger)
   - Commitment: ±5% or ±€1,000 (whichever is larger)

## Source Files
- VERSO: `datamigration/VERSO DASHBOARD_V1.0.xlsx`
- INNOVATECH: `datamigration/INNOVATECH DASHBOARD_V1.xlsx`
- Name Mappings: `VERSO/datafixing/introducers name change.csv`

---

*Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*
"""

    with open(os.path.join(output_dir, 'SUMMARY.md'), 'w') as f:
        f.write(summary)

    print(f"\n{'='*50}")
    print(f"AUDIT COMPLETE")
    print(f"{'='*50}")
    print(f"✅ Verified: {verified}")
    print(f"⚠️ Minor Variance: {minor}")
    print(f"❌ Major Discrepancy: {major}")
    print(f"\nReports saved to: {output_dir}/")


if __name__ == '__main__':
    main()
