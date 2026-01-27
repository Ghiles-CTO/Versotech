#!/usr/bin/env python3
"""
Fresh Independent Data Audit - Dashboard vs Production DB
Extracts data directly from Excel dashboards and compares with Supabase.
"""

from openpyxl import load_workbook
from datetime import datetime
import csv
import os

# Name normalization mappings
NAME_MAPPINGS = {
    'Anand': 'Setcap',
    'Anand Sethia': 'Setcap',
    'Dan': 'Daniel Baumslag',
    'Daniel': 'Daniel Baumslag',
    'Rick': 'Altras Capital Financing Broker',
    'Elevation+Rick': 'Altras Capital Financing Broker',
    'Denis': 'Denis Matthey',
    'Robin': 'Robin Doble',
    'Sandro': 'GEMERA Consulting Pte Ltd',
    'Sandro-Gemera': 'GEMERA Consulting Pte Ltd',
    'Gemera': 'GEMERA Consulting Pte Ltd',
    'Simone': 'Manna Capital',
    'Terra': 'Terra Financial & Management Services SA',
    'Elevation': 'Elevation Securities',
    'Gary': 'Game Venture Management LLC',
    'John': 'Moore & Moore Investments Ltd',
    'Omar': 'Omar ADI',
    'Stableton': 'Stableton Financial AG',
    'Andrew': 'Andrew Stewart',
    'Aboud': 'Aboud Khaddam',
    'AUX': 'AUX Business Support Ltd',
    'Enguerrand': 'Enguerrand Elbaz',
    'Gio': 'Giovanni SALADINO',
}

def normalize_introducer_name(name):
    """Normalize introducer name using mappings."""
    if not name:
        return None
    name = str(name).strip()
    return NAME_MAPPINGS.get(name, name)

def get_sheet_name(vehicle_code):
    """Convert vehicle code to sheet name."""
    # VC106 -> VC6, VC111 -> VC11, IN103 -> IN3, etc.
    if vehicle_code.startswith('VC1'):
        # VC106 -> VC6, VC111 -> VC11, etc.
        num = vehicle_code[2:]  # Remove 'VC'
        if len(num) == 3 and num.startswith('1'):
            return 'VC' + num[1:]  # VC106 -> VC6, VC111 -> VC11
        elif len(num) == 3 and num.startswith('2'):
            return 'VC' + num  # VC201 stays VC201
    elif vehicle_code.startswith('VC2'):
        return vehicle_code  # VC201, VC202 stay as is
    elif vehicle_code.startswith('IN1'):
        # IN101 -> IN1, IN103 -> IN3, etc.
        num = vehicle_code[2:]
        if len(num) == 3 and num.startswith('1'):
            return 'IN' + num[1:]
    elif vehicle_code.startswith('IN'):
        return vehicle_code
    return vehicle_code

def extract_dashboard_data(dashboard_path, sheet_name):
    """Extract investor and introducer data directly from Excel."""
    wb = load_workbook(dashboard_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        wb.close()
        return None, f"Sheet '{sheet_name}' not found in workbook"

    ws = wb[sheet_name]

    subscriptions = []
    introductions = []

    for row_num, row in enumerate(ws.iter_rows(min_row=4), start=4):
        # Extract investor data (fixed columns)
        investor_last = row[6].value if len(row) > 6 else None  # Col 7
        investor_first = row[8].value if len(row) > 8 else None  # Col 9
        investor_entity = row[9].value if len(row) > 9 else None  # Col 10
        amount = row[11].value if len(row) > 11 else None  # Col 12
        price_per_share = row[12].value if len(row) > 12 else None  # Col 13
        shares = row[13].value if len(row) > 13 else None  # Col 14
        ownership = row[14].value if len(row) > 14 else None  # Col 15
        contract_date = row[15].value if len(row) > 15 else None  # Col 16

        # Extract introducer data (dynamic columns)
        partner_name = row[44].value if len(row) > 44 else None  # Col 45
        bi_name = row[54].value if len(row) > 54 else None  # Col 55

        # Subscription fees from each section
        partner_sub_pct = row[45].value if len(row) > 45 else None  # Col 46
        partner_sub_amt = row[46].value if len(row) > 46 else None  # Col 47
        bi_sub_pct = row[55].value if len(row) > 55 else None  # Col 56
        bi_sub_amt = row[56].value if len(row) > 56 else None  # Col 57

        # Spread fees
        partner_spread_pct = row[51].value if len(row) > 51 else None  # Col 52
        partner_spread_amt = row[52].value if len(row) > 52 else None  # Col 53
        bi_spread_pct = None  # Need to find correct column
        bi_spread_amt = None

        # Skip empty rows
        if not investor_last and not amount:
            continue

        # Handle ownership as number
        try:
            ownership_val = float(ownership) if ownership else 0
        except (ValueError, TypeError):
            ownership_val = 0

        # Create subscription record
        sub_record = {
            'row': row_num,
            'investor_last': investor_last,
            'investor_first': investor_first,
            'investor_entity': investor_entity,
            'amount': amount,
            'price_per_share': price_per_share,
            'shares': shares,
            'ownership': ownership_val,
            'contract_date': contract_date,
            'has_subscription': ownership_val > 0
        }
        subscriptions.append(sub_record)

        # Create introduction records for Partner and BI
        if partner_name and partner_name not in ['None', '', 'N/A']:
            intro_record = {
                'row': row_num,
                'investor_last': investor_last,
                'investor_first': investor_first,
                'amount': amount,
                'introducer_name': normalize_introducer_name(partner_name),
                'raw_name': partner_name,
                'type': 'partner',
                'sub_fee_pct': partner_sub_pct,
                'sub_fee_amt': partner_sub_amt,
                'spread_pct': partner_spread_pct,
                'spread_amt': partner_spread_amt,
            }
            introductions.append(intro_record)

        if bi_name and bi_name not in ['None', '', 'N/A', 'VERSO BI']:
            intro_record = {
                'row': row_num,
                'investor_last': investor_last,
                'investor_first': investor_first,
                'amount': amount,
                'introducer_name': normalize_introducer_name(bi_name),
                'raw_name': bi_name,
                'type': 'introducer',
                'sub_fee_pct': bi_sub_pct,
                'sub_fee_amt': bi_sub_amt,
                'spread_pct': bi_spread_pct,
                'spread_amt': bi_spread_amt,
            }
            introductions.append(intro_record)

    wb.close()

    return {
        'subscriptions': subscriptions,
        'introductions': introductions,
        'active_subscriptions': [s for s in subscriptions if s['has_subscription']],
        'zero_ownership': [s for s in subscriptions if not s['has_subscription']],
    }, None


if __name__ == '__main__':
    # Test extraction
    result, error = extract_dashboard_data(
        'datamigration/VERSO DASHBOARD_V1.0.xlsx',
        'VC6'
    )

    if error:
        print(f"Error: {error}")
    else:
        print(f"Total rows: {len(result['subscriptions'])}")
        print(f"Active subscriptions (ownership > 0): {len(result['active_subscriptions'])}")
        print(f"Zero ownership rows: {len(result['zero_ownership'])}")
        print(f"Introductions: {len(result['introductions'])}")

        # Print first 5 subscriptions
        print("\nFirst 5 active subscriptions:")
        for sub in result['active_subscriptions'][:5]:
            print(f"  Row {sub['row']}: {sub['investor_last']} - €{sub['amount']} - {sub['shares']} shares")

        # Print unique introducers
        unique_introducers = set(i['introducer_name'] for i in result['introductions'])
        print(f"\nUnique introducers: {unique_introducers}")
