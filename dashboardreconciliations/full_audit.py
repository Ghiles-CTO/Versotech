#!/usr/bin/env python3
"""
Full Vehicle Audit - Extract from dashboards and compare with DB data.
Handles different sheet structures across vehicles.
"""

from openpyxl import load_workbook
from datetime import datetime
import os

# DB Data (from Supabase query)
DB_DATA = {
    'VC106': {'subs': 197, 'pos': 171, 'comm': 1048, 'intro': 259, 'commitment': 45363968.32, 'shares': 2005322.0},
    'VC113': {'subs': 77, 'pos': 60, 'comm': 102, 'intro': 59, 'commitment': 31834740.51, 'shares': 1066920.0},
    'VC111': {'subs': 39, 'pos': 34, 'comm': 78, 'intro': 53, 'commitment': 8842857.14, 'shares': 7750000.0},
    'VC125': {'subs': 34, 'pos': 27, 'comm': 4, 'intro': 7, 'commitment': 2792111.35, 'shares': 16652.0},
    'VC126': {'subs': 29, 'pos': 19, 'comm': 42, 'intro': 42, 'commitment': 7300231.00, 'shares': 48857.0},
    'VC112': {'subs': 25, 'pos': 13, 'comm': 4, 'intro': 3, 'commitment': 2796196.15, 'shares': 2250027.0},
    'VC133': {'subs': 12, 'pos': 11, 'comm': 18, 'intro': 12, 'commitment': 3225055.00, 'shares': 2220.0},
    'VC124': {'subs': 11, 'pos': 7, 'comm': 0, 'intro': 0, 'commitment': 611772.70, 'shares': 2809648.0},
    'IN103': {'subs': 9, 'pos': 6, 'comm': 32, 'intro': 14, 'commitment': 2750000.00, 'shares': 483567},
    'IN102': {'subs': 8, 'pos': 8, 'comm': 0, 'intro': 0, 'commitment': 2512609.22, 'shares': 603},
    'VC118': {'subs': 7, 'pos': 6, 'comm': 20, 'intro': 8, 'commitment': 2192270.01, 'shares': 268324.0},
    'IN110': {'subs': 6, 'pos': 5, 'comm': 0, 'intro': 0, 'commitment': 75.00, 'shares': 255000},
    'VC102': {'subs': 5, 'pos': 4, 'comm': 1, 'intro': 1, 'commitment': 275003.00, 'shares': 128500.0},
    'VC122': {'subs': 5, 'pos': 4, 'comm': 2, 'intro': 1, 'commitment': 384999.30, 'shares': 295842.0},
    'VC128': {'subs': 4, 'pos': 3, 'comm': 0, 'intro': 0, 'commitment': 300000.00, 'shares': 407142.0},
    'VC130': {'subs': 4, 'pos': 4, 'comm': 0, 'intro': 0, 'commitment': 445000.00, 'shares': 266252.0},
    'VC114': {'subs': 4, 'pos': 4, 'comm': 0, 'intro': 0, 'commitment': 860000.00, 'shares': 860000.0},
    'VC140': {'subs': 3, 'pos': 3, 'comm': 0, 'intro': 0, 'commitment': 246000.00, 'shares': 246000.0},
    'IN111': {'subs': 3, 'pos': 1, 'comm': 0, 'intro': 0, 'commitment': 107263.44, 'shares': 77447},
    'VC143': {'subs': 2, 'pos': 2, 'comm': 0, 'intro': 1, 'commitment': 175000.00, 'shares': 175000.0},
    'IN106': {'subs': 2, 'pos': 1, 'comm': 2, 'intro': 1, 'commitment': 340000.00, 'shares': 68164},
    'VC121': {'subs': 2, 'pos': 1, 'comm': 0, 'intro': 0, 'commitment': 100000.00, 'shares': 172413.0},
    'VC131': {'subs': 2, 'pos': 1, 'comm': 0, 'intro': 0, 'commitment': 37500.00, 'shares': 32500.0},
    'VC132': {'subs': 2, 'pos': 1, 'comm': 0, 'intro': 0, 'commitment': 230547.59, 'shares': 27546.0},
    'VC141': {'subs': 2, 'pos': 2, 'comm': 0, 'intro': 0, 'commitment': 150000.00, 'shares': 150000.0},
    'VC138': {'subs': 1, 'pos': 1, 'comm': 0, 'intro': 0, 'commitment': 2000000.00, 'shares': 20.0},
    'IN109': {'subs': 1, 'pos': 1, 'comm': 0, 'intro': 0, 'commitment': 595000.00, 'shares': 6071},
    'IN101': {'subs': 1, 'pos': 1, 'comm': 0, 'intro': 0, 'commitment': 914481.12, 'shares': 38881},
    'VC123': {'subs': 1, 'pos': 1, 'comm': 0, 'intro': 0, 'commitment': 100000.00, 'shares': 100000.0},
}

# Sheet configurations - (filepath, sheet_name, header_row, data_start_row, col_offsets)
# col_offsets: (last_name, first_name, entity, amount, shares, ownership)
SHEET_CONFIGS = {
    # Standard VERSO layout (most vehicles)
    'standard_verso': {
        'file': 'datamigration/VERSO DASHBOARD_V1.0.xlsx',
        'header_row': 3, 'data_row': 4,
        'cols': {'last': 6, 'first': 8, 'entity': 9, 'amount': 11, 'shares': 13, 'ownership': 14}
    },
    # VC11 has different layout
    'vc11_layout': {
        'file': 'datamigration/VERSO DASHBOARD_V1.0.xlsx',
        'header_row': 2, 'data_row': 3,
        'cols': {'last': 5, 'first': 7, 'entity': 8, 'amount': 10, 'shares': 12, 'ownership': 13}
    },
    # INNOVATECH layout
    'innovatech': {
        'file': 'datamigration/INNOVATECH DASHBOARD_V1.xlsx',
        'header_row': 3, 'data_row': 4,
        'cols': {'last': 6, 'first': 8, 'entity': 9, 'amount': 11, 'shares': 13, 'ownership': 14}
    },
}

# Vehicle to sheet mapping
VEHICLE_SHEETS = {
    'VC102': ('VC2', 'standard_verso'), 'VC106': ('VC6', 'standard_verso'),
    'VC111': ('VC11', 'vc11_layout'), 'VC112': ('VC12', 'standard_verso'),
    'VC113': ('VC13', 'standard_verso'), 'VC114': ('VC14', 'standard_verso'),
    'VC118': ('VC18', 'standard_verso'), 'VC121': ('VC21', 'standard_verso'),
    'VC122': ('VC22', 'standard_verso'), 'VC123': ('VC23', 'standard_verso'),
    'VC124': ('VC24', 'standard_verso'), 'VC125': ('VC25', 'standard_verso'),
    'VC126': ('VC26', 'standard_verso'), 'VC128': ('VC28', 'standard_verso'),
    'VC130': ('VC30', 'standard_verso'), 'VC131': ('VC31', 'standard_verso'),
    'VC132': ('VC32', 'standard_verso'), 'VC133': ('VC33', 'standard_verso'),
    'VC138': ('VC38', 'standard_verso'), 'VC140': ('VC40', 'standard_verso'),
    'VC141': ('VC41', 'standard_verso'), 'VC143': ('VC43', 'standard_verso'),
    'IN101': ('IN1', 'innovatech'), 'IN102': ('IN2', 'innovatech'),
    'IN103': ('IN3', 'innovatech'), 'IN106': ('IN6', 'innovatech'),
    'IN109': ('IN9', 'innovatech'), 'IN110': ('IN10', 'innovatech'),
    'IN111': ('IN11', 'innovatech'),
}


def extract_vehicle_data(vehicle_code):
    """Extract subscription data from dashboard for a vehicle."""
    if vehicle_code not in VEHICLE_SHEETS:
        return None, f"Unknown vehicle: {vehicle_code}"

    sheet_name, config_key = VEHICLE_SHEETS[vehicle_code]
    config = SHEET_CONFIGS[config_key]

    try:
        wb = load_workbook(config['file'], read_only=True, data_only=True)
    except Exception as e:
        return None, f"Error loading {config['file']}: {e}"

    if sheet_name not in wb.sheetnames:
        wb.close()
        return None, f"Sheet '{sheet_name}' not found"

    ws = wb[sheet_name]
    cols = config['cols']

    subscriptions = []
    summary_rows = []
    zero_ownership = []

    for row_num, row in enumerate(ws.iter_rows(min_row=config['data_row']), start=config['data_row']):
        # Extract using 0-indexed columns
        investor_last = row[cols['last']].value if len(row) > cols['last'] else None
        investor_first = row[cols['first']].value if len(row) > cols['first'] else None
        investor_entity = row[cols['entity']].value if len(row) > cols['entity'] else None
        amount = row[cols['amount']].value if len(row) > cols['amount'] else None
        shares = row[cols['shares']].value if len(row) > cols['shares'] else None
        ownership = row[cols['ownership']].value if len(row) > cols['ownership'] else None

        try:
            ownership_val = float(ownership) if ownership else 0
            amount_val = float(amount) if amount else 0
            shares_val = float(shares) if shares else 0
        except (ValueError, TypeError):
            continue

        if not investor_last and not investor_entity and not amount:
            continue

        has_investor = investor_last or investor_entity

        if ownership_val > 0:
            if has_investor:
                subscriptions.append({
                    'row': row_num,
                    'investor': investor_last or investor_entity,
                    'amount': amount_val,
                    'shares': shares_val,
                    'ownership': ownership_val
                })
            else:
                summary_rows.append({'row': row_num, 'amount': amount_val})
        elif has_investor:
            zero_ownership.append({'row': row_num, 'investor': investor_last or investor_entity})

    wb.close()

    return {
        'subscription_count': len(subscriptions),
        'total_commitment': sum(s['amount'] for s in subscriptions),
        'total_shares': sum(s['shares'] for s in subscriptions),
        'zero_ownership_count': len(zero_ownership),
        'summary_rows': len(summary_rows),
        'subscriptions': subscriptions,
        'sheet_name': sheet_name,
        'config': config_key,
    }, None


def generate_report(vehicle_code, dashboard_data, db_data):
    """Generate audit report for a vehicle."""
    d = dashboard_data
    db = db_data

    # Calculate differences
    sub_diff = d['subscription_count'] - db['subs']
    commit_diff = d['total_commitment'] - db['commitment']
    shares_diff = d['total_shares'] - db['shares']

    # Determine status
    sub_match = abs(sub_diff) <= 2  # Allow 2 tolerance
    commit_match = abs(commit_diff) < 1000  # Allow €1000 tolerance
    shares_match = abs(shares_diff) < 100  # Allow 100 shares tolerance

    overall_status = "✅ VERIFIED" if (sub_match and commit_match) else "⚠️ DISCREPANCIES FOUND"

    report = f"""# {vehicle_code} Audit Report
**Date:** {datetime.now().strftime('%Y-%m-%d')}
**Dashboard Sheet:** {d['sheet_name']} ({d['config']})

## Status: {overall_status}

---

## Counts Comparison

| Type | Dashboard | DB | Difference | Match |
|------|-----------|-----|------------|-------|
| Subscriptions | {d['subscription_count']} | {db['subs']} | {sub_diff:+d} | {'✅' if sub_match else '❌'} |
| Positions | — | {db['pos']} | — | — |
| Commissions | — | {db['comm']} | — | — |
| Introductions | — | {db['intro']} | — | — |

---

## Financial Verification

| Metric | Dashboard | Database | Difference |
|--------|-----------|----------|------------|
| Total Commitment | €{d['total_commitment']:,.2f} | €{db['commitment']:,.2f} | €{commit_diff:+,.2f} |
| Total Shares | {d['total_shares']:,.0f} | {db['shares']:,.0f} | {shares_diff:+,.0f} |

---

## Dashboard Notes
- Zero Ownership Rows (excluded): {d['zero_ownership_count']}
- Summary Rows (excluded): {d['summary_rows']}

---

## Verification Checklist

- [{'x' if sub_match else ' '}] Subscription count within tolerance (±2)
- [{'x' if commit_match else ' '}] Commitment amount within tolerance (±€1,000)
- [{'x' if shares_match else ' '}] Share count within tolerance (±100)
- [x] Zero-ownership rows correctly excluded
- [x] Summary rows correctly excluded

---

## Conclusion
**{vehicle_code} data integrity: {overall_status}**
"""

    if not sub_match or not commit_match:
        report += f"""
### Discrepancies to Investigate
"""
        if not sub_match:
            report += f"- Subscription count differs by {abs(sub_diff)} records\n"
        if not commit_match:
            report += f"- Commitment amount differs by €{abs(commit_diff):,.2f}\n"
        if not shares_match:
            report += f"- Share count differs by {abs(shares_diff):,.0f}\n"

    return report, {
        'vehicle': vehicle_code,
        'status': 'VERIFIED' if (sub_match and commit_match) else 'DISCREPANCY',
        'sub_diff': sub_diff,
        'commit_diff': commit_diff,
        'shares_diff': shares_diff,
    }


def main():
    output_dir = 'dashboardreconciliations'
    os.makedirs(output_dir, exist_ok=True)

    summary_results = []

    for vehicle_code in sorted(DB_DATA.keys()):
        print(f"Auditing {vehicle_code}...", end=" ")

        dashboard_data, error = extract_vehicle_data(vehicle_code)

        if error:
            print(f"ERROR: {error}")
            summary_results.append({
                'vehicle': vehicle_code,
                'status': 'ERROR',
                'error': error
            })
            continue

        report, result = generate_report(vehicle_code, dashboard_data, DB_DATA[vehicle_code])
        summary_results.append(result)

        # Write report
        report_path = os.path.join(output_dir, f"{vehicle_code}.md")
        with open(report_path, 'w') as f:
            f.write(report)

        print(result['status'])

    # Generate summary
    verified = sum(1 for r in summary_results if r.get('status') == 'VERIFIED')
    discrepancies = sum(1 for r in summary_results if r.get('status') == 'DISCREPANCY')
    errors = sum(1 for r in summary_results if r.get('status') == 'ERROR')

    summary = f"""# Dashboard Reconciliation Summary
**Date:** {datetime.now().strftime('%Y-%m-%d')}
**Vehicles Audited:** {len(summary_results)}

## Overall Status

| Status | Count |
|--------|-------|
| ✅ Verified | {verified} |
| ⚠️ Discrepancies | {discrepancies} |
| ❌ Errors | {errors} |

---

## Vehicle Status

| Vehicle | Status | Sub Diff | Commitment Diff | Notes |
|---------|--------|----------|-----------------|-------|
"""

    for r in sorted(summary_results, key=lambda x: x['vehicle']):
        if r.get('status') == 'ERROR':
            summary += f"| {r['vehicle']} | ❌ ERROR | — | — | {r.get('error', '')} |\n"
        else:
            status_icon = '✅' if r['status'] == 'VERIFIED' else '⚠️'
            notes = ""
            if abs(r['sub_diff']) > 2:
                notes += f"{r['sub_diff']:+d} subs; "
            if abs(r['commit_diff']) > 1000:
                notes += f"€{r['commit_diff']:+,.0f}; "
            summary += f"| {r['vehicle']} | {status_icon} {r['status']} | {r['sub_diff']:+d} | €{r['commit_diff']:+,.2f} | {notes.rstrip('; ')} |\n"

    summary += """
---

## Key Findings

### Perfect Matches (Subscription count and amounts match exactly or within tolerance)
"""
    for r in sorted(summary_results, key=lambda x: x['vehicle']):
        if r.get('status') == 'VERIFIED':
            summary += f"- **{r['vehicle']}**: ✅ Verified\n"

    summary += """
### Vehicles with Discrepancies (Require investigation)
"""
    for r in sorted(summary_results, key=lambda x: x['vehicle']):
        if r.get('status') == 'DISCREPANCY':
            summary += f"- **{r['vehicle']}**: {r['sub_diff']:+d} subscriptions, €{r['commit_diff']:+,.2f} commitment difference\n"

    summary += """
---

## Methodology

1. **Data Extraction**: Directly from Excel dashboards using openpyxl (read_only=True, data_only=True)
2. **Filtering**: Only rows with ownership > 0 AND investor/entity name counted
3. **Exclusions**: Summary rows and zero-ownership rows properly excluded
4. **Tolerance**: ±2 subscriptions, ±€1,000 commitment considered acceptable variance
5. **Comparison**: Against Supabase production database

## Files Generated
- Individual vehicle reports: `dashboardreconciliations/<VEHICLE>.md`
- This summary: `dashboardreconciliations/SUMMARY.md`
"""

    summary_path = os.path.join(output_dir, 'SUMMARY.md')
    with open(summary_path, 'w') as f:
        f.write(summary)

    print(f"\n=== Audit Complete ===")
    print(f"Verified: {verified}")
    print(f"Discrepancies: {discrepancies}")
    print(f"Errors: {errors}")
    print(f"\nReports saved to {output_dir}/")


if __name__ == '__main__':
    main()
