#!/usr/bin/env python3
"""
Vehicle Audit Script - Extracts data from Excel and compares with DB results.
Usage: python run_audit.py <vehicle_code>
"""

import sys
from openpyxl import load_workbook
from datetime import datetime

# Vehicle to Sheet mapping
VERSO_VEHICLES = {
    'VC102': 'VC2', 'VC106': 'VC6', 'VC111': 'VC11', 'VC112': 'VC12', 'VC113': 'VC13',
    'VC114': 'VC14', 'VC118': 'VC18', 'VC121': 'VC21', 'VC122': 'VC22', 'VC123': 'VC23',
    'VC124': 'VC24', 'VC125': 'VC25', 'VC126': 'VC26', 'VC128': 'VC28', 'VC130': 'VC30',
    'VC131': 'VC31', 'VC132': 'VC32', 'VC133': 'VC33', 'VC138': 'VC38', 'VC140': 'VC40',
    'VC141': 'VC41', 'VC143': 'VC43',
}

INNOVATECH_VEHICLES = {
    'IN101': 'IN1', 'IN102': 'IN2', 'IN103': 'IN3', 'IN106': 'IN6',
    'IN109': 'IN9', 'IN110': 'IN10', 'IN111': 'IN11',
}

# Name normalization
NAME_MAPPINGS = {
    'Anand': 'Setcap', 'Anand Sethia': 'Setcap',
    'Dan': 'Daniel Baumslag', 'Daniel': 'Daniel Baumslag',
    'Rick': 'Altras Capital Financing Broker', 'Elevation+Rick': 'Altras Capital Financing Broker',
    'Denis': 'Denis Matthey', 'Robin': 'Robin Doble',
    'Sandro': 'GEMERA Consulting Pte Ltd', 'Sandro-Gemera': 'GEMERA Consulting Pte Ltd',
    'Gemera': 'GEMERA Consulting Pte Ltd', 'Simone': 'Manna Capital',
    'Terra': 'Terra Financial & Management Services SA',
    'Elevation': 'Elevation Securities', 'Gary': 'Game Venture Management LLC',
    'John': 'Moore & Moore Investments Ltd', 'Omar': 'Omar ADI',
    'Stableton': 'Stableton Financial AG', 'Andrew': 'Andrew Stewart',
    'Aboud': 'Aboud Khaddam', 'AUX': 'AUX Business Support Ltd',
    'Enguerrand': 'Enguerrand Elbaz', 'Gio': 'Giovanni SALADINO',
}

def get_dashboard_path(vehicle_code):
    """Get the Excel file path for a vehicle."""
    if vehicle_code in VERSO_VEHICLES:
        return 'datamigration/VERSO DASHBOARD_V1.0.xlsx', VERSO_VEHICLES[vehicle_code]
    elif vehicle_code in INNOVATECH_VEHICLES:
        return 'datamigration/INNOVATECH DASHBOARD_V1.xlsx', INNOVATECH_VEHICLES[vehicle_code]
    return None, None

def extract_dashboard_data(vehicle_code):
    """Extract data from dashboard for a vehicle."""
    filepath, sheet_name = get_dashboard_path(vehicle_code)
    if not filepath:
        return None, f"Unknown vehicle: {vehicle_code}"

    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        return None, f"Error loading workbook: {e}"

    if sheet_name not in wb.sheetnames:
        wb.close()
        return None, f"Sheet '{sheet_name}' not found"

    ws = wb[sheet_name]

    subscriptions = []
    introductions = []
    summary_rows = []
    zero_ownership = []

    for row_num, row in enumerate(ws.iter_rows(min_row=4), start=4):
        # Extract data
        investor_last = row[6].value if len(row) > 6 else None  # Col 7
        investor_first = row[8].value if len(row) > 8 else None  # Col 9
        investor_entity = row[9].value if len(row) > 9 else None  # Col 10
        amount = row[11].value if len(row) > 11 else None  # Col 12
        shares = row[13].value if len(row) > 13 else None  # Col 14
        ownership = row[14].value if len(row) > 14 else None  # Col 15
        contract_date = row[15].value if len(row) > 15 else None  # Col 16

        # Partner/Introducer columns
        partner_name = row[44].value if len(row) > 44 else None  # Col 45
        bi_name = row[54].value if len(row) > 54 else None  # Col 55

        # Parse values
        try:
            ownership_val = float(ownership) if ownership else 0
            amount_val = float(amount) if amount else 0
            shares_val = float(shares) if shares else 0
        except (ValueError, TypeError):
            continue

        # Skip completely empty rows
        if not investor_last and not investor_entity and not amount:
            continue

        has_investor = investor_last or investor_entity

        if ownership_val > 0:
            if has_investor:
                sub_record = {
                    'row': row_num,
                    'investor_last': investor_last,
                    'investor_first': investor_first,
                    'investor_entity': investor_entity,
                    'amount': amount_val,
                    'shares': shares_val,
                    'ownership': ownership_val,
                    'contract_date': str(contract_date) if contract_date else None,
                }
                subscriptions.append(sub_record)

                # Track introductions
                if partner_name and partner_name not in ['None', '', 'N/A', 'VERSO PARTNER']:
                    introductions.append({
                        'row': row_num,
                        'introducer': NAME_MAPPINGS.get(partner_name, partner_name),
                        'type': 'partner'
                    })
                if bi_name and bi_name not in ['None', '', 'N/A', 'VERSO BI']:
                    introductions.append({
                        'row': row_num,
                        'introducer': NAME_MAPPINGS.get(bi_name, bi_name),
                        'type': 'introducer'
                    })
            else:
                summary_rows.append({
                    'row': row_num,
                    'amount': amount_val,
                    'ownership': ownership_val
                })
        elif investor_last or investor_entity:
            zero_ownership.append({
                'row': row_num,
                'investor': investor_last or investor_entity,
                'amount': amount_val
            })

    wb.close()

    # Calculate totals
    total_commitment = sum(s['amount'] for s in subscriptions)
    total_shares = sum(s['shares'] for s in subscriptions)
    unique_introducers = list(set(i['introducer'] for i in introductions))

    return {
        'subscriptions': subscriptions,
        'subscription_count': len(subscriptions),
        'total_commitment': total_commitment,
        'total_shares': total_shares,
        'introductions': introductions,
        'introduction_count': len(introductions),
        'unique_introducers': unique_introducers,
        'zero_ownership_count': len(zero_ownership),
        'summary_rows': summary_rows,
        'sheet_name': sheet_name,
        'filepath': filepath,
    }, None


def main():
    if len(sys.argv) < 2:
        print("Usage: python run_audit.py <vehicle_code>")
        print("Example: python run_audit.py VC113")
        sys.exit(1)

    vehicle_code = sys.argv[1].upper()

    print(f"=== Extracting Dashboard Data for {vehicle_code} ===")
    result, error = extract_dashboard_data(vehicle_code)

    if error:
        print(f"Error: {error}")
        sys.exit(1)

    print(f"Sheet: {result['sheet_name']} in {result['filepath']}")
    print(f"")
    print(f"Subscriptions: {result['subscription_count']}")
    print(f"Total Commitment: €{result['total_commitment']:,.2f}")
    print(f"Total Shares: {result['total_shares']:,.0f}")
    print(f"Zero Ownership Rows: {result['zero_ownership_count']}")
    print(f"Summary Rows Excluded: {len(result['summary_rows'])}")
    print(f"")
    print(f"Introductions: {result['introduction_count']}")
    print(f"Unique Introducers: {result['unique_introducers']}")

    if result['summary_rows']:
        print(f"")
        print("Summary rows excluded:")
        for sr in result['summary_rows']:
            print(f"  Row {sr['row']}: €{sr['amount']:,.2f}")


if __name__ == '__main__':
    main()
