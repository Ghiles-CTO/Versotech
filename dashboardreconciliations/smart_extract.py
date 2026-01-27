#!/usr/bin/env python3
"""
Smart Extraction - Dynamically detects column positions from headers.
"""

from openpyxl import load_workbook
import re

# Column header patterns (case-insensitive)
HEADER_PATTERNS = {
    'last_name': [r'investor last name', r'last name'],
    'first_name': [r'investor first name', r'first name'],
    'entity': [r'investor entity', r'entity'],
    'amount': [r'amount invested', r'amount'],
    'shares': [r'number of shares', r'shares invested'],
    'ownership': [r'ownership', r'ownership position', r'^position$'],  # INNOVATECH uses "Position"
}


def find_columns(ws, header_row):
    """Find column indices for key fields based on header patterns."""
    columns = {}

    for row in ws.iter_rows(min_row=header_row, max_row=header_row):
        for col_idx, cell in enumerate(row):
            if not cell.value:
                continue
            val = str(cell.value).lower().strip()

            for field, patterns in HEADER_PATTERNS.items():
                if field in columns:
                    continue
                for pattern in patterns:
                    if re.search(pattern, val, re.IGNORECASE):
                        columns[field] = col_idx
                        break

    return columns


def detect_data_start(ws, max_search=10):
    """Detect which row contains headers and which starts data."""
    for row_num in range(1, max_search + 1):
        for row in ws.iter_rows(min_row=row_num, max_row=row_num):
            for cell in row:
                val = str(cell.value).lower() if cell.value else ''
                if 'investor last name' in val or 'last name' in val:
                    # Found header row, data starts next row
                    return row_num, row_num + 1
    return 3, 4  # Default


def smart_extract(filepath, sheet_name):
    """Extract data with automatic column detection."""
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        return None, f"Error loading {filepath}: {e}"

    if sheet_name not in wb.sheetnames:
        wb.close()
        return None, f"Sheet '{sheet_name}' not found"

    ws = wb[sheet_name]

    # Detect header and data rows
    header_row, data_row = detect_data_start(ws)

    # Find column positions
    columns = find_columns(ws, header_row)

    # Validate we found essential columns
    if 'last_name' not in columns and 'entity' not in columns:
        wb.close()
        return None, f"Could not find investor columns in headers (row {header_row})"

    subscriptions = []
    summary_rows = []
    zero_ownership = []

    for row_num, row in enumerate(ws.iter_rows(min_row=data_row), start=data_row):
        # Extract data using detected columns
        investor_last = row[columns['last_name']].value if 'last_name' in columns and len(row) > columns['last_name'] else None
        investor_first = row[columns['first_name']].value if 'first_name' in columns and len(row) > columns['first_name'] else None
        investor_entity = row[columns['entity']].value if 'entity' in columns and len(row) > columns['entity'] else None
        amount = row[columns['amount']].value if 'amount' in columns and len(row) > columns['amount'] else None
        shares = row[columns['shares']].value if 'shares' in columns and len(row) > columns['shares'] else None
        ownership = row[columns['ownership']].value if 'ownership' in columns and len(row) > columns['ownership'] else None

        try:
            ownership_val = float(ownership) if ownership else 0
            amount_val = float(amount) if amount else 0
            shares_val = float(shares) if shares else 0
        except (ValueError, TypeError):
            continue

        if not investor_last and not investor_entity and not amount:
            continue

        has_investor = investor_last or investor_entity

        if ownership_val > 0:
            if has_investor:
                subscriptions.append({
                    'row': row_num,
                    'investor': investor_last or investor_entity,
                    'amount': amount_val,
                    'shares': shares_val,
                    'ownership': ownership_val
                })
            else:
                summary_rows.append({'row': row_num, 'amount': amount_val})
        elif has_investor:
            zero_ownership.append({'row': row_num, 'investor': investor_last or investor_entity})

    wb.close()

    return {
        'subscription_count': len(subscriptions),
        'total_commitment': sum(s['amount'] for s in subscriptions),
        'total_shares': sum(s['shares'] for s in subscriptions),
        'zero_ownership_count': len(zero_ownership),
        'summary_rows': len(summary_rows),
        'subscriptions': subscriptions,
        'sheet_name': sheet_name,
        'header_row': header_row,
        'data_row': data_row,
        'columns_found': columns,
    }, None


# Vehicle to file/sheet mapping
VEHICLE_MAP = {
    'VC102': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC2'),
    'VC106': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC6'),
    'VC111': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC11'),
    'VC112': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC12'),
    'VC113': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC13'),
    'VC114': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC14'),
    'VC118': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC18'),
    'VC121': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC21'),
    'VC122': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC22'),
    'VC123': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC23'),
    'VC124': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC24'),
    'VC125': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC25'),
    'VC126': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC26'),
    'VC128': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC28'),
    'VC130': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC30'),
    'VC131': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC31'),
    'VC132': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC32'),
    'VC133': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC33'),
    'VC138': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC38'),
    'VC140': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC40'),
    'VC141': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC41'),
    'VC143': ('datamigration/VERSO DASHBOARD_V1.0.xlsx', 'VC43'),
    'IN101': ('datamigration/INNOVATECH DASHBOARD_V1.xlsx', 'IN1'),
    'IN102': ('datamigration/INNOVATECH DASHBOARD_V1.xlsx', 'IN2'),
    'IN103': ('datamigration/INNOVATECH DASHBOARD_V1.xlsx', 'IN3'),
    'IN106': ('datamigration/INNOVATECH DASHBOARD_V1.xlsx', 'IN6'),
    'IN109': ('datamigration/INNOVATECH DASHBOARD_V1.xlsx', 'IN9'),
    'IN110': ('datamigration/INNOVATECH DASHBOARD_V1.xlsx', 'IN10'),
    'IN111': ('datamigration/INNOVATECH DASHBOARD_V1.xlsx', 'IN11'),
}


if __name__ == '__main__':
    import sys

    if len(sys.argv) < 2:
        print("Usage: python smart_extract.py <vehicle_code>")
        sys.exit(1)

    vehicle = sys.argv[1].upper()

    if vehicle not in VEHICLE_MAP:
        print(f"Unknown vehicle: {vehicle}")
        sys.exit(1)

    filepath, sheet_name = VEHICLE_MAP[vehicle]
    result, error = smart_extract(filepath, sheet_name)

    if error:
        print(f"Error: {error}")
        sys.exit(1)

    print(f"=== {vehicle} Extraction Results ===")
    print(f"Sheet: {result['sheet_name']}")
    print(f"Header Row: {result['header_row']}, Data Row: {result['data_row']}")
    print(f"Columns Found: {result['columns_found']}")
    print()
    print(f"Subscriptions: {result['subscription_count']}")
    print(f"Total Commitment: €{result['total_commitment']:,.2f}")
    print(f"Total Shares: {result['total_shares']:,.0f}")
    print(f"Zero Ownership: {result['zero_ownership_count']}")
    print(f"Summary Rows: {result['summary_rows']}")
