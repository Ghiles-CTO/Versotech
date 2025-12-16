import argparse
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl


def _clean_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        text = (
            value.replace("\u00a0", " ")
            .replace("\r\n", "\n")
            .replace("\r", "\n")
        )
        text = "\n".join(line.rstrip() for line in text.split("\n"))
        text = text.strip()
        return text if text else None
    return str(value).strip() or None


def _format_ref(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, (int,)):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        # Avoid scientific notation while keeping significant digits.
        text = f"{value}"
        return text.rstrip("0").rstrip(".")
    return _clean_text(value)


def _markdown_bullet(text: str, indent: str = "  ") -> str:
    lines = text.split("\n")
    if not lines:
        return "- "
    out = f"- {lines[0]}"
    for line in lines[1:]:
        out += f"\n{indent}{line}"
    return out


class SheetConfig:
    def __init__(
        self,
        feature_ref_col: int,
        feature_name_col: int,
        sub_ref_col: int,
        sub_name_col: int,
        story_col: int,
        mvp_col: Optional[int] = None,
    ) -> None:
        self.feature_ref_col = feature_ref_col
        self.feature_name_col = feature_name_col
        self.sub_ref_col = sub_ref_col
        self.sub_name_col = sub_name_col
        self.story_col = story_col
        self.mvp_col = mvp_col


def extract_sheet(ws: Any, config: SheetConfig) -> List[Dict[str, Any]]:
    current_feature_ref: Optional[str] = None
    current_feature_name: Optional[str] = None
    current_sub_ref: Optional[str] = None
    current_sub_name: Optional[str] = None
    current_mvp: Optional[str] = None

    rows: List[Dict[str, Any]] = []

    for row_index in range(2, ws.max_row + 1):
        feature_ref = _format_ref(ws.cell(row=row_index, column=config.feature_ref_col).value)
        feature_name = _clean_text(ws.cell(row=row_index, column=config.feature_name_col).value)

        mvp_value: Optional[str] = None
        if config.mvp_col is not None:
            mvp_value = _clean_text(ws.cell(row=row_index, column=config.mvp_col).value)

        sub_ref = _clean_text(ws.cell(row=row_index, column=config.sub_ref_col).value)
        sub_name = _clean_text(ws.cell(row=row_index, column=config.sub_name_col).value)
        story = _clean_text(ws.cell(row=row_index, column=config.story_col).value)

        # Forward-fill hierarchy (Excel omits repeated values across rows).
        if feature_ref is not None or feature_name is not None:
            if feature_ref is not None:
                current_feature_ref = feature_ref
            if feature_name is not None:
                current_feature_name = feature_name
            if config.mvp_col is not None:
                # MVP is often repeated in the sheet; capture when present.
                if mvp_value is not None:
                    current_mvp = mvp_value

        if sub_ref is not None or sub_name is not None:
            if sub_ref is not None:
                current_sub_ref = sub_ref
            if sub_name is not None:
                current_sub_name = sub_name

        rows.append(
            {
                "sheet_row": row_index,
                "feature_ref": current_feature_ref,
                "feature": current_feature_name,
                "mvp": current_mvp,
                "sub_ref": current_sub_ref,
                "sub_feature": current_sub_name,
                "story": story,
            }
        )

    return rows


def group_rows(rows: List[Dict[str, Any]]) -> Dict[Tuple[str, str, Optional[str]], Dict[Tuple[str, str], List[Dict[str, Any]]]]:
    grouped: Dict[Tuple[str, str, Optional[str]], Dict[Tuple[str, str], List[Dict[str, Any]]]] = defaultdict(
        lambda: defaultdict(list)
    )

    for entry in rows:
        feature_ref = entry.get("feature_ref") or "(missing feature ref)"
        feature = entry.get("feature") or "(missing feature name)"
        mvp = entry.get("mvp")
        sub_ref = entry.get("sub_ref") or "(missing sub ref)"
        sub_feature = entry.get("sub_feature") or "(missing sub-feature name)"
        grouped[(feature_ref, feature, mvp)][(sub_ref, sub_feature)].append(entry)

    return grouped


def render_markdown(
    workbook_path: Path,
    sheet_names: List[str],
    include_blank_story_rows: bool,
) -> str:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)

    lines: List[str] = []
    lines.append("# User Stories Extract (Mobile V6)")
    lines.append("")
    lines.append(f"Source: `{workbook_path.name}`")
    lines.append("")
    lines.append("Notes:")
    lines.append("- This file is generated from the Excel sheets; Excel row numbers are preserved as `Row N`.")
    lines.append("- Some rows in the workbook have empty `User stories` cells (kept only when `--include-blank-story-rows` is used).")
    lines.append("")

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            raise SystemExit(f"Sheet not found: {sheet_name}")

        ws = wb[sheet_name]
        has_mvp = sheet_name.startswith("0.")
        config = (
            SheetConfig(
                feature_ref_col=2,
                feature_name_col=3,
                sub_ref_col=5,
                sub_name_col=6,
                story_col=7,
                mvp_col=4,
            )
            if has_mvp
            else SheetConfig(
                feature_ref_col=2,
                feature_name_col=3,
                sub_ref_col=4,
                sub_name_col=5,
                story_col=6,
            )
        )

        raw_rows = extract_sheet(ws, config)

        if include_blank_story_rows:
            rows = raw_rows
        else:
            rows = [r for r in raw_rows if r.get("story") is not None]

        lines.append(f"## {sheet_name}")
        lines.append("")
        lines.append(f"- Excel `max_row`: {ws.max_row} (header row + data rows)")
        lines.append(f"- Extracted rows: {len(rows)} ({'including' if include_blank_story_rows else 'excluding'} blank `User stories` cells)")
        if not include_blank_story_rows:
            blank_rows = [r["sheet_row"] for r in raw_rows if r.get("story") is None]
            if blank_rows:
                preview = ", ".join(str(n) for n in blank_rows[:20])
                more = "" if len(blank_rows) <= 20 else f", … (+{len(blank_rows) - 20} more)"
                lines.append(f"- Blank `User stories` rows skipped: {len(blank_rows)} (first: {preview}{more})")
        lines.append("")

        grouped = group_rows(rows)

        # Sort by feature_ref if possible (e.g., "4.1" < "4.2"), otherwise lexicographic.
        def sort_key_feature(item: Tuple[str, str, Optional[str]]) -> Tuple:
            ref = item[0]
            parts: List[Any] = []
            for token in ref.split("."):
                try:
                    parts.append(int(token))
                except ValueError:
                    parts.append(token)
            return tuple(parts)

        for (feature_ref, feature_name, mvp) in sorted(grouped.keys(), key=sort_key_feature):
            mvp_suffix = f" (MVP: {mvp})" if has_mvp and mvp is not None else ""
            lines.append(f"### {feature_ref} — {feature_name}{mvp_suffix}")
            lines.append("")

            sub_groups = grouped[(feature_ref, feature_name, mvp)]

            def sort_key_sub(item: Tuple[str, str]) -> Tuple:
                ref = item[0]
                parts: List[Any] = []
                for token in ref.split("."):
                    try:
                        parts.append(int(token))
                    except ValueError:
                        parts.append(token)
                return tuple(parts)

            for (sub_ref, sub_name) in sorted(sub_groups.keys(), key=sort_key_sub):
                lines.append(f"#### {sub_ref} — {sub_name}")
                lines.append("")

                entries = sub_groups[(sub_ref, sub_name)]
                for entry in entries:
                    story = entry.get("story")
                    if story is None:
                        lines.append(f"- (Row {entry['sheet_row']}) (blank user story)")
                    else:
                        row_number = entry["sheet_row"]
                        lines.append(_markdown_bullet(f"(Row {row_number}) {story}"))
                lines.append("")

    return "\n".join(lines).rstrip() + "\n"


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract user stories from the Mobile V6 Excel workbook.")
    parser.add_argument(
        "--workbook",
        default="User_stories_Verso_Capital_mobile_V6 (1).xlsx",
        help="Path to the Excel workbook",
    )
    parser.add_argument(
        "--out",
        default="docs/important/user_stories_mobile_v6_extracted.md",
        help="Output markdown path",
    )
    parser.add_argument(
        "--include-blank-story-rows",
        action="store_true",
        help="Include rows where the 'User stories' cell is blank",
    )
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # Preserve the sheet order specified in docs/important/plan_next_phase.md.
    sheet_names = [
        "0. Enablers",
        "1.CEO",
        "2.Arranger",
        "3.Lawyer",
        "4.Investor",
        "5.Partner",
        "6.Introducer",
        "7.Commercial Partner",
    ]

    markdown = render_markdown(
        workbook_path=workbook_path,
        sheet_names=sheet_names,
        include_blank_story_rows=bool(args.include_blank_story_rows),
    )

    out_path.write_text(markdown, encoding="utf-8")
    print(f"Wrote: {out_path}")


if __name__ == "__main__":
    main()
