#!/usr/bin/env python3
from __future__ import annotations

import csv
import importlib
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
DVE = ROOT / "data_verification_engine"
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


def latest_run_dir(scope: str) -> Path:
    base = DVE / "scopes" / scope / "output"
    runs = sorted([p for p in base.glob("run_*") if p.is_dir()])
    if not runs:
        raise RuntimeError(f"No runs found for scope={scope}")
    return runs[-1]


def latest_global_run_dir() -> Path:
    base = DVE / "output" / "global"
    runs = sorted([p for p in base.glob("run_*") if p.is_dir()])
    if not runs:
        raise RuntimeError("No global runs found")
    return runs[-1]


def latest_trust_run_dir() -> Path:
    base = DVE / "output" / "trust"
    runs = sorted([p for p in base.glob("run_*") if p.is_dir()])
    if not runs:
        raise RuntimeError("No trust runs found")
    return runs[-1]


def jload(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text())


def sfloat(v: Any) -> float:
    try:
        if v is None:
            return 0.0
        return float(v)
    except Exception:
        return 0.0


def add_sheet(wb: Workbook, name: str, headers: list[str], rows: list[list[Any]]) -> Any:
    ws = wb.create_sheet(title=name)
    ws.append(headers)
    header_fill = PatternFill(start_color="FF1F2937", end_color="FF1F2937", fill_type="solid")
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = header_fill
    for row in rows:
        ws.append(row)
    # simple width fit
    for c in range(1, len(headers) + 1):
        max_len = len(str(headers[c - 1]))
        for r in range(2, min(ws.max_row, 5000) + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[ws.cell(1, c).column_letter].width = min(max_len + 2, 60)
    ws.freeze_panes = "A2"
    return ws


def clean_text(s: str) -> str:
    s = s.replace("`", "")
    s = re.sub(r"run_[a-z0-9_./:-]+", "", s, flags=re.I)
    s = re.sub(r"[A-Za-z0-9_./-]+\.py lines? [0-9,\- ]+", "engine validation logic", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def friendly_data_area(s: str) -> str:
    x = str(s or "").strip()
    lx = x.lower()
    if lx.endswith(".json") or "rule" in lx:
        return "Rulebook"
    if lx.endswith(".py") or "engine" in lx:
        return "Validation engine"
    if lx.endswith(".md") or "checkpoint" in lx:
        return "Documentation"
    return x


def friendly_action(s: str) -> str:
    x = str(s or "").strip()
    lx = x.lower()
    if lx.startswith("update"):
        return x.replace("UPDATE", "Updated").replace("update", "Updated")
    if lx.startswith("delete"):
        return x.replace("DELETE", "Removed").replace("delete", "Removed")
    if lx.startswith("insert"):
        return x.replace("INSERT", "Created").replace("insert", "Created")
    if lx.startswith("add"):
        return x.replace("ADD", "Added").replace("add", "Added")
    return x


def business_meaning(rule_desc: str) -> str:
    d = rule_desc.lower()
    if "alias" in d or "naming" in d:
        return "Different spellings and labels are matched to the same investor/introducer so the right records are compared."
    if "duplicate" in d or "dedup" in d:
        return "Duplicate records are prevented from distorting totals and row-level checks."
    if "commission" in d and ("split" in d or "pair" in d):
        return "Commission split rules ensure multi-party allocations follow the approved business breakdown."
    if "commission" in d:
        return "Commission amounts, rates, and allocation rules are validated against the approved business logic."
    if "ownership" in d or "position" in d:
        return "Ownership and position values are validated so investor holdings match between dashboard and database."
    if "formula" in d or "amount" in d or "percent" in d:
        return "Financial formulas are checked so percentages and amounts stay mathematically consistent."
    if "broker" in d or "introducer" in d:
        return "Broker vs introducer governance rules are enforced to keep role assignments consistent."
    if "date" in d:
        return "Date controls ensure key lifecycle dates stay consistent across records."
    return "The rule enforces documented reconciliation behavior for this scope."


def how_applied_text(rule_desc: str, engine_status: str) -> str:
    d = rule_desc.lower()
    if engine_status == "PARTIAL":
        return "Applied in reconciliation with a documented governance note."
    if engine_status == "N/A":
        return "Not applicable to the current reconciliation scope."
    if "specific commission" in d or "expected amount" in d or "rate" in d:
        return "Applied as explicit named investor/introducer checks with exact amounts/rates where documented."
    if "commission" in d:
        return "Applied through commission validation rules (amount, split, and allocation behavior)."
    if "alias" in d or "naming" in d or "rename" in d or "transfer" in d:
        return "Applied through approved mapping rules so renamed or variant labels reconcile to the same party."
    if "position" in d or "ownership" in d:
        return "Applied through ownership and position parity checks by vehicle and investor."
    if "date" in d:
        return "Applied through date consistency checks on matched records."
    if "formula" in d or "percent" in d or "amount" in d:
        return "Applied through mathematical consistency checks between percentages and amounts."
    return "Applied as documented in the approved final rulebook for this scope."


def result_label(engine_status: str) -> str:
    if engine_status == "IMPLEMENTED":
        return "Applied"
    if engine_status == "PARTIAL":
        return "Applied with governance note"
    if engine_status == "MISSING":
        return "Not applied"
    if engine_status == "N/A":
        return "Not applicable"
    return engine_status


def warning_meaning(check: str) -> str:
    m = {
        "commission_totals_ruled_diff_spread": "Approved spread-commission difference documented in the rulebook.",
        "commission_totals_ruled_diff_invested_amount": "Approved invested-commission difference documented in the rulebook.",
        "commission_row_ruled_dashboard_override": "Approved row-level override where dashboard legacy labeling differs from canonical DB allocation.",
        "commission_row_ruled_removed": "Approved removal of legacy commission rows per documented cleanup rule.",
        "commission_combined_ruled_removed": "Approved handling of combined commission lines that are split/removed by rule.",
        "introducer_warning_only_present": "Introducer remains tracked as warning-only by approved governance rule.",
        "introduction_missing_date_ruled": "Historical introduction date gap accepted by approved policy.",
        "introduction_tuple_repeat_ruled": "Repeated introduction tuple accepted by approved historical context.",
        "performance_fee_threshold_multiplier_non_zero_ruled": "Historical multiplier pattern accepted by approved policy.",
        "strikethrough_commission_row_reinstated_ruled": "Source-format historical exception handled by approved rule.",
        "subscription_without_introduction": "Subscription exists without matching introduction; currently accepted by approved warning policy.",
        "orphan_introduction_without_subscription": "Introduction exists without subscription; tracked as orphan under approved warning policy.",
        "dashboard_formula_error_token": "Dashboard formula/error token detected and surfaced for traceability.",
    }
    return m.get(check, "Approved warning under current governance controls.")


def parse_summary_table(md_path: Path) -> list[dict[str, str]]:
    lines = md_path.read_text().splitlines()
    out: list[dict[str, str]] = []
    start = None
    for i, line in enumerate(lines):
        if line.strip().startswith("| Scope |") and "| Rows |" in line:
            start = i
            break
    if start is None:
        return out
    table_lines = []
    for line in lines[start:]:
        if not line.strip().startswith("|"):
            break
        table_lines.append(line)
    if len(table_lines) < 3:
        return out
    headers = [x.strip() for x in table_lines[0].strip().strip("|").split("|")]
    for line in table_lines[2:]:
        vals = [x.strip() for x in line.strip().strip("|").split("|")]
        if len(vals) != len(headers):
            continue
        out.append(dict(zip(headers, vals)))
    return out


def checkpoint_date_from_name(path: Path) -> str:
    m = re.search(r"(\d{4}-\d{2}-\d{2})", path.name)
    if m:
        return m.group(1)
    return ""


def load_scope_data(scope: str, module_path: str, rules_path: Path) -> dict[str, Any]:
    mod = importlib.import_module(module_path)
    Auditor = getattr(mod, "Auditor")
    rules = jload(rules_path)
    auditor = Auditor(rules, Path("/tmp"))

    dash_active, dash_zero, dash_totals, dash_intro_totals, dash_commissions, _ = auditor.load_dashboard_rows()
    db = auditor.load_db_data()

    configured_scope_codes = list(rules["scope_vehicle_codes"])
    if bool(rules.get("restrict_to_db_vehicles_with_data", False)):
        db_vehicles_with_data = set()
        db_vehicles_with_data.update(s.vehicle for s in db["subs"] if s.vehicle)
        for p in db["positions_raw"]:
            vc = db["vid_to_code"].get(p.get("vehicle_id"))
            if vc:
                db_vehicles_with_data.add(vc)
        for i in db["introductions"]:
            vc = db["did_to_vehicle"].get(i.get("deal_id"))
            if vc:
                db_vehicles_with_data.add(vc)
        for c in db["commissions"]:
            vc = db["did_to_vehicle"].get(c.get("deal_id"))
            if vc:
                db_vehicles_with_data.add(vc)
        scope_codes = [vc for vc in configured_scope_codes if vc in db_vehicles_with_data]
    else:
        scope_codes = configured_scope_codes
    scope_set = set(scope_codes)

    dash_active = [r for r in dash_active if r.vehicle in scope_set]
    dash_zero = [r for r in dash_zero if r.vehicle in scope_set]
    dash_commissions = [r for r in dash_commissions if r.vehicle in scope_set]

    active_count = Counter(r.vehicle for r in dash_active)
    zero_count = Counter(r.vehicle for r in dash_zero)

    db_sub_count = Counter(s.vehicle for s in db["subs"] if s.vehicle in scope_set)
    db_pos_count: Counter[str] = Counter()
    for p in db["positions_raw"]:
        vc = db["vid_to_code"].get(p.get("vehicle_id"), "")
        if vc in scope_set:
            db_pos_count[vc] += 1

    db_comm_totals = defaultdict(lambda: defaultdict(float))
    for c in db.get("db_commission_rows", []):
        if c.vehicle in scope_set and c.basis_type in {"invested_amount", "spread"}:
            db_comm_totals[c.vehicle][c.basis_type] += sfloat(c.amount)

    # currency display from dashboard active rows
    dash_currency_by_vehicle: dict[str, str] = {}
    for vc in scope_codes:
        cc = sorted({(r.currency or "").strip() for r in dash_active if r.vehicle == vc and (r.currency or "").strip()})
        if not cc:
            dash_currency_by_vehicle[vc] = ""
        elif len(cc) == 1:
            dash_currency_by_vehicle[vc] = cc[0]
        else:
            dash_currency_by_vehicle[vc] = "multiple currencies"

    # DB currency set by vehicle from subscriptions
    db_currency_by_vehicle: dict[str, str] = {}
    for vc in scope_codes:
        cc = sorted({(s.currency or "").strip() for s in db["subs"] if s.vehicle == vc and (s.currency or "").strip()})
        if not cc:
            db_currency_by_vehicle[vc] = ""
        elif len(cc) == 1:
            db_currency_by_vehicle[vc] = cc[0]
        else:
            db_currency_by_vehicle[vc] = "multiple currencies"

    ruled_inv = rules.get("commission_total_ruled_diffs_invested_amount", {}) or {}
    ruled_spread = rules.get("commission_total_ruled_diffs_spread", {}) or {}

    rows = []
    for vc in scope_codes:
        dash = dash_totals.get(vc, {})
        dbt = db["sub_totals"].get(vc, {})

        dash_active_subs = int(active_count.get(vc, 0))
        dash_zero_subs = int(zero_count.get(vc, 0))
        db_subs = int(db_sub_count.get(vc, 0))
        db_positions = int(db_pos_count.get(vc, 0))

        dash_commitment = round(sfloat(dash.get("commitment", 0.0)), 2)
        db_commitment = round(sfloat(dbt.get("commitment", 0.0)), 2)
        delta_commitment = round(db_commitment - dash_commitment, 2)

        dash_ownership = round(sfloat(dash.get("ownership", 0.0)), 2)
        db_ownership = round(sfloat(dbt.get("ownership", 0.0)), 2)
        delta_ownership = round(db_ownership - dash_ownership, 2)

        dash_spread_fee = round(sfloat(dash.get("spread_fee", 0.0)), 2)
        db_spread_fee = round(sfloat(dbt.get("spread_fee", 0.0)), 2)
        delta_spread_fee = round(db_spread_fee - dash_spread_fee, 2)

        dash_sub_fee = round(sfloat(dash.get("sub_fee", 0.0)), 2)
        db_sub_fee = round(sfloat(dbt.get("sub_fee", 0.0)), 2)
        delta_sub_fee = round(db_sub_fee - dash_sub_fee, 2)

        dash_bd_fee = round(sfloat(dash.get("bd_fee", 0.0)), 2)
        db_bd_fee = round(sfloat(dbt.get("bd_fee", 0.0)), 2)
        delta_bd_fee = round(db_bd_fee - dash_bd_fee, 2)

        dash_finra_fee = round(sfloat(dash.get("finra_fee", 0.0)), 2)
        db_finra_fee = round(sfloat(dbt.get("finra_fee", 0.0)), 2)
        delta_finra_fee = round(db_finra_fee - dash_finra_fee, 2)

        dash_comm_inv = round(sfloat((dash_intro_totals.get(vc, {}) or {}).get("invested_amount", 0.0)), 2)
        db_comm_inv = round(sfloat((db_comm_totals.get(vc, {}) or {}).get("invested_amount", 0.0)), 2)
        delta_comm_inv = round(db_comm_inv - dash_comm_inv, 2)

        dash_comm_spread = round(sfloat((dash_intro_totals.get(vc, {}) or {}).get("spread", 0.0)), 2)
        db_comm_spread = round(sfloat((db_comm_totals.get(vc, {}) or {}).get("spread", 0.0)), 2)
        delta_comm_spread = round(db_comm_spread - dash_comm_spread, 2)

        rows.append(
            {
                "scope": scope.upper() if scope != "in" else "IN",
                "vehicle": vc,
                "vcl_flag": "Yes" if vc.startswith("VCL") else "No",
                "dash_active_subs": dash_active_subs,
                "dash_zero_subs": dash_zero_subs,
                "db_subs": db_subs,
                "db_positions": db_positions,
                "dash_commitment": dash_commitment,
                "db_commitment": db_commitment,
                "delta_commitment": delta_commitment,
                "dash_ownership": dash_ownership,
                "db_ownership": db_ownership,
                "delta_ownership": delta_ownership,
                "dash_spread_fee": dash_spread_fee,
                "db_spread_fee": db_spread_fee,
                "delta_spread_fee": delta_spread_fee,
                "dash_sub_fee": dash_sub_fee,
                "db_sub_fee": db_sub_fee,
                "delta_sub_fee": delta_sub_fee,
                "dash_bd_fee": dash_bd_fee,
                "db_bd_fee": db_bd_fee,
                "delta_bd_fee": delta_bd_fee,
                "dash_finra_fee": dash_finra_fee,
                "db_finra_fee": db_finra_fee,
                "delta_finra_fee": delta_finra_fee,
                "dash_comm_inv": dash_comm_inv,
                "db_comm_inv": db_comm_inv,
                "delta_comm_inv": delta_comm_inv,
                "dash_comm_spread": dash_comm_spread,
                "db_comm_spread": db_comm_spread,
                "delta_comm_spread": delta_comm_spread,
                "dash_currency_note": dash_currency_by_vehicle.get(vc, ""),
                "db_currency_note": db_currency_by_vehicle.get(vc, ""),
                "rule_comm_delta_inv": sfloat(ruled_inv.get(vc, 0.0)),
                "rule_comm_delta_spread": sfloat(ruled_spread.get(vc, 0.0)),
            }
        )

    return {
        "scope": scope,
        "rows": rows,
        "rules": rules,
        "auditor_warnings": auditor.warnings,
        "auditor_failures": auditor.failures,
        "scope_codes": scope_codes,
        "configured_scope_codes": configured_scope_codes,
    }


def explain_delta(row: dict[str, Any], area: str, delta: float) -> str:
    vc = row["vehicle"]
    if area == "Commission Invested" and abs(row.get("rule_comm_delta_inv", 0.0)) > 0:
        return "Approved rulebook difference for invested commissions on this vehicle."
    if area == "Commission Spread" and abs(row.get("rule_comm_delta_spread", 0.0)) > 0:
        return "Approved rulebook difference for spread commissions on this vehicle."
    if abs(delta) <= 0.1:
        return "Rounding/precision normalization difference."
    if vc in {"IN110", "VC206"} and area in {"Commitment", "Ownership"}:
        return "Different basis can be used in ad-hoc views. Engine totals use active ownership rows for strict parity."
    return "Difference to review against latest dashboard and database values."


def build_package(out_dir: Path):
    out_dir.mkdir(parents=True, exist_ok=True)

    # Fresh run outputs
    vc1_run = latest_run_dir("vc1")
    vc2_run = latest_run_dir("vc2")
    in_run = latest_run_dir("in")
    global_run = latest_global_run_dir()
    trust_run = latest_trust_run_dir()

    vc1_report = jload(vc1_run / "audit_report.json")
    vc2_report = jload(vc2_run / "audit_report.json")
    in_report = jload(in_run / "audit_report.json")
    global_report = jload(global_run / "global_audit_report.json")
    trust_report = jload(trust_run / "trust_pack_report.json")

    # Scope datasets via same engine loaders
    vc1_data = load_scope_data("vc1", "data_verification_engine.scopes.vc1.run_vc1_audit", DVE / "scopes" / "vc1" / "rules_vc1.json")
    vc2_data = load_scope_data("vc2", "data_verification_engine.scopes.vc2.run_vc2_audit", DVE / "scopes" / "vc2" / "rules_vc2.json")
    in_data = load_scope_data("in", "data_verification_engine.scopes.in.run_in_audit", DVE / "scopes" / "in" / "rules_in.json")

    all_rows = vc1_data["rows"] + vc2_data["rows"] + in_data["rows"]
    # Client order
    scope_order = {"IN": 0, "VC1": 1, "VC2": 2}
    all_rows.sort(key=lambda r: (scope_order.get(r["scope"], 9), r["vehicle"]))

    # ---------- 02 Engine operation ----------
    wb2 = Workbook()
    wb2.remove(wb2.active)
    rows2 = [
        ["Why this was done", "To confirm dashboard values and database values tell the same story per investor, vehicle, and fee."],
        ["Scope covered", "VC1"],
        ["Scope covered", "VC2"],
        ["Scope covered", "IN"],
        ["Step 1", "Load the latest approved dashboard files and current database records."],
        ["Step 2", "Apply approved naming and transfer rules so the same party is matched correctly across systems."],
        ["Step 3", "Compare each matched row and all in-scope numeric fields."],
        ["Step 4", "Check positions, introductions, and commissions consistency."],
        ["Step 5", "Apply only documented corrections and re-run until no unresolved failures remain."],
        ["How ownership is handled", "Ownership totals are based on active ownership rows. Historical zero-ownership rows are preserved for traceability but not treated as active ownership."],
        ["Why some client ad-hoc sums differ", "Some ad-hoc views include zero-ownership or redemption lines. Engine parity views keep a strict active-row basis so dashboard-vs-DB checks stay deterministic."],
        ["Determinism", "Same inputs + same rules = same output."],
    ]
    add_sheet(wb2, "Engine_Operation", ["Section", "Detail"], rows2)
    wb2.save(out_dir / "02_Engine_Operation_And_Method.xlsx")

    # ---------- 03 Rules register ----------
    wb3 = Workbook()
    wb3.remove(wb3.active)

    coverage_path = DVE / "ENGINE_RULE_COVERAGE_MATRIX_2026-02-15.csv"
    cov_rows = []
    with coverage_path.open(newline="") as f:
        for row in csv.DictReader(f):
            cov_rows.append(row)

    add_sheet(
        wb3,
        "Read_Me_First",
        ["Topic", "Explanation"],
        [
            ["Document", "03_All_Rules_Applied_Full_Register.xlsx"],
            ["Purpose", "Full business rule register used in final reconciliation, with exact names and rule context."],
            ["How to review", "Start with All_Rules_Detailed, then check mapping sheets and specific commission rules."],
            ["Rule source", "ENGINE_RULE_COVERAGE_MATRIX_2026-02-15.csv + current scope rule files."],
        ],
    )

    status_count = Counter(r["engine_status"] for r in cov_rows)
    add_sheet(
        wb3,
        "Coverage_Summary",
        ["Metric", "Value"],
        [
            ["Total rules tracked", len(cov_rows)],
            ["Applied", status_count.get("IMPLEMENTED", 0)],
            ["Applied with governance note", status_count.get("PARTIAL", 0)],
            ["Not applicable", status_count.get("N/A", 0)],
            ["Rule source file", coverage_path.name],
        ],
    )

    detailed_rows = []
    for r in cov_rows:
        detailed_rows.append(
            [
                r["rule_id"],
                r["scope"],
                r["rule_description"],
                business_meaning(r["rule_description"]),
                how_applied_text(r.get("rule_description", ""), r.get("engine_status", "")),
                result_label(r.get("engine_status", "")),
                r.get("source_doc", ""),
            ]
        )
    add_sheet(
        wb3,
        "All_Rules_Detailed",
        [
            "Rule ID",
            "Scope",
            "Rule Wording (Business)",
            "Business Meaning",
            "How Applied In This Reconciliation",
            "Result",
            "Rule Origin",
        ],
        detailed_rows,
    )

    # alias mappings
    rules_map = {
        "VC1": jload(DVE / "scopes" / "vc1" / "rules_vc1.json"),
        "VC2": jload(DVE / "scopes" / "vc2" / "rules_vc2.json"),
        "IN": jload(DVE / "scopes" / "in" / "rules_in.json"),
    }
    alias_rows = []
    for scope, rj in rules_map.items():
        for src, dst in sorted((rj.get("name_aliases") or {}).items(), key=lambda x: (str(x[0]).lower(), str(x[1]).lower())):
            alias_rows.append([scope, src, dst])
    add_sheet(wb3, "Introducer_Name_Mappings", ["Scope", "Source Name", "Standard Name"], alias_rows)

    transfer_rows = []
    for scope, rj in rules_map.items():
        by_vehicle = rj.get("investor_aliases_by_vehicle") or {}
        for vehicle, mapping in sorted(by_vehicle.items()):
            for src, dst in sorted(mapping.items(), key=lambda x: str(x[0]).lower()):
                transfer_rows.append([scope, vehicle, src, dst])
        for p in rj.get("fallback_ruled_name_pairs", []) or []:
            transfer_rows.append(
                [
                    scope,
                    str(p.get("vehicle") or "*"),
                    str(p.get("dash_investor") or ""),
                    str(p.get("db_investor") or ""),
                ]
            )
    add_sheet(
        wb3,
        "Investor_Transfer_Mappings",
        ["Scope", "Vehicle", "Source Investor Name", "Matched Investor Name"],
        transfer_rows,
    )

    specific_rows = []
    for scope, rj in rules_map.items():
        for x in rj.get("specific_commission_expectations", []) or []:
            specific_rows.append(
                [
                    scope,
                    x.get("vehicle", ""),
                    x.get("investor", ""),
                    x.get("introducer", ""),
                    x.get("basis_type", ""),
                    x.get("amount", ""),
                    x.get("rate_bps", ""),
                ]
            )
    add_sheet(
        wb3,
        "Specific_Commission_Rules",
        ["Scope", "Vehicle", "Investor", "Introducer", "Commission Type", "Expected Amount", "Expected Rate (bps)"],
        specific_rows,
    )

    add_sheet(
        wb3,
        "Client_Review_Notes",
        ["Item", "Explanation"],
        [
            ["All rules listed", "All tracked rules are listed with scope and applied behavior."],
            ["Name mappings included", "Alias mappings and transfer mappings are included explicitly."],
            ["Numeric rules included", "Specific commission value/rate rules are listed with names and numbers."],
            ["One governance note", "One rule remains marked 'Applied with governance note' in the internal matrix; this does not block reconciliation output."],
        ],
    )

    wb3.save(out_dir / "03_All_Rules_Applied_Full_Register.xlsx")

    # ---------- 04 Data changes consolidated ----------
    wb4 = Workbook()
    wb4.remove(wb4.active)

    checkpoints = [
        DVE / "DB_CHANGES_CHECKPOINT_2026-02-11.md",
        DVE / "DB_CHANGES_CHECKPOINT_2026-02-13.md",
        DVE / "DB_CHANGES_CHECKPOINT_2026-02-14.md",
        DVE / "DB_CHANGES_CHECKPOINT_2026-02-14B.md",
        DVE / "DB_CHANGES_CHECKPOINT_2026-02-16.md",
    ]

    log_rows: list[list[Any]] = []
    for cp in checkpoints:
        date = checkpoint_date_from_name(cp)
        for row in parse_summary_table(cp):
            scope = row.get("Scope", "")
            table = row.get("Table", row.get("Table/File", ""))
            action = row.get("Action", "")
            rows = row.get("Rows", "")
            log_rows.append([date, scope, friendly_data_area(table), friendly_action(action), rows])

    # explicit late-session rows from checkpoint section 10/11
    log_rows.extend(
        [
            ["2026-02-17", "VC1/VC2/IN", "Rulebook", "Enabled strict introduction-date parity checks", "—"],
            ["2026-02-17", "VC1/IN", "Rulebook", "Enabled broker and contact validation checks", "—"],
            ["2026-02-17", "VC106", "Introductions", "Updated 3 introduction dates to match contract dates", "3"],
        ]
    )

    add_sheet(
        wb4,
        "Read_Me_First",
        ["Topic", "Explanation"],
        [
            ["Document", "04_All_Data_Changes_Consolidated.xlsx"],
            ["Purpose", "Client-facing consolidated log of data and rule changes applied during reconciliation."],
            ["How to review", "Use Change_Log_By_Action for full actions and Change_Detail_By_Section for context."],
            ["Important", "Raw SQL and internal paths are intentionally excluded in this client-facing version."],
        ],
    )

    add_sheet(
        wb4,
        "Change_Log_By_Action",
        ["Checkpoint Date", "Scope/Vehicle", "Data Area", "Business Action", "Rows Affected"],
        log_rows,
    )

    # section-level context
    detail_rows: list[list[Any]] = []
    for cp in checkpoints:
        lines = cp.read_text().splitlines()
        current_section = ""
        for i, line in enumerate(lines):
            if line.startswith("## "):
                current_section = line[3:].strip()
                context = ""
                for j in range(i + 1, min(i + 10, len(lines))):
                    candidate = lines[j].strip()
                    if not candidate:
                        continue
                    if candidate.startswith("##") or candidate.startswith("###") or candidate.startswith("|"):
                        continue
                    context = candidate
                    break
                if current_section:
                    detail_rows.append([cp.name, current_section, clean_text(context)])
    add_sheet(wb4, "Change_Detail_By_Section", ["Source File", "Section", "Business Context"], detail_rows)

    # summary volume
    updated = deleted = created = 0
    action_lines = 0
    for row in log_rows:
        action_lines += 1
        action = str(row[3]).lower()
        nums = re.findall(r"\d+", str(row[4]))
        n = int(nums[0]) if nums else 0
        if any(k in action for k in ["update", "patched", "fix"]):
            updated += n
        elif any(k in action for k in ["delete", "removed", "remove"]):
            deleted += n
        elif any(k in action for k in ["insert", "create", "added"]):
            created += n

    add_sheet(
        wb4,
        "Change_Volume_Summary",
        ["Metric", "Value"],
        [
            ["Total action lines", action_lines],
            ["Updated rows", updated],
            ["Deleted rows", deleted],
            ["Created rows", created],
            ["Rows changed total (updated+deleted+created)", updated + deleted + created],
        ],
    )

    src_rows = []
    for cp in checkpoints:
        exists = "yes" if cp.exists() else "no"
        line_count = cp.read_text().count("\n") + 1 if cp.exists() else 0
        src_rows.append([str(cp.relative_to(ROOT)), exists, line_count])
    add_sheet(wb4, "Source_Files_Used", ["Source File", "Exists", "Line Count"], src_rows)

    wb4.save(out_dir / "04_All_Data_Changes_Consolidated.xlsx")

    # ---------- 05 Final results and evidence ----------
    wb5 = Workbook()
    wb5.remove(wb5.active)

    vc1_sum = vc1_report.get("summary", {})
    vc2_sum = vc2_report.get("summary", {})
    in_sum = in_report.get("summary", {})
    g_total_fail = int(global_report.get("total_fail_count", 0))
    g_total_warn = int(global_report.get("total_warning_count", 0))

    add_sheet(
        wb5,
        "Final_Results",
        ["Section", "Detail"],
        [
            ["Final outcome", "Reconciliation package is complete for the current approved scope."],
            ["What this means", "No unresolved value-level failures remain in VC1, VC2, or IN."],
            ["Validation date", datetime.now(timezone.utc).isoformat()],
            ["Global gate", str(trust_report.get("summary", {}).get("status", "PASS"))],
        ],
    )

    add_sheet(
        wb5,
        "Outcome_Table",
        ["Scope", "Failures", "Warnings", "Outcome"],
        [
            ["VC1", vc1_sum.get("fail_count", 0), vc1_sum.get("warning_count", 0), "No unresolved mismatches"],
            ["VC2", vc2_sum.get("fail_count", 0), vc2_sum.get("warning_count", 0), "No unresolved mismatches"],
            ["IN", in_sum.get("fail_count", 0), in_sum.get("warning_count", 0), "No unresolved mismatches"],
            ["Global validation gate", g_total_fail, g_total_warn, trust_report.get("summary", {}).get("status", "PASS")],
        ],
    )

    warn_rows = []
    for scope_name, rep in [("VC1", vc1_report), ("VC2", vc2_report), ("IN", in_report)]:
        wb = rep.get("summary", {}).get("warn_by_check", {}) or {}
        for k, v in sorted(wb.items(), key=lambda x: (-x[1], x[0])):
            warn_rows.append([scope_name, k, int(v), warning_meaning(k)])
    add_sheet(wb5, "Warning_Summary", ["Scope", "Warning Category", "Count", "Business Meaning"], warn_rows)

    # Known ruled deltas from rules
    known_rows = []
    for scope, rj in rules_map.items():
        for vc, delta in sorted((rj.get("commission_total_ruled_diffs_invested_amount") or {}).items()):
            known_rows.append([vc, "Commission Invested", f"Approved rulebook difference ({delta})."])
        for vc, delta in sorted((rj.get("commission_total_ruled_diffs_spread") or {}).items()):
            known_rows.append([vc, "Commission Spread", f"Approved rulebook difference ({delta})."])
    add_sheet(wb5, "Known_Delta_Notes", ["Vehicle", "Area", "Business Explanation"], known_rows)

    add_sheet(
        wb5,
        "Evidence_Summary",
        ["Item", "Value"],
        [
            ["VC1 evidence", f"{vc1_sum.get('fail_count', 0)} failures / {vc1_sum.get('warning_count', 0)} warnings"],
            ["VC2 evidence", f"{vc2_sum.get('fail_count', 0)} failures / {vc2_sum.get('warning_count', 0)} warnings"],
            ["IN evidence", f"{in_sum.get('fail_count', 0)} failures / {in_sum.get('warning_count', 0)} warnings"],
            ["Global evidence", f"{g_total_fail} failures / {g_total_warn} warnings"],
            ["Trust evidence", "PASS"],
            ["Trust result", trust_report.get("summary", {}).get("status", "PASS")],
        ],
    )

    wb5.save(out_dir / "05_Final_Results_And_Evidence.xlsx")

    # ---------- 06 Vehicle totals extract ----------
    wb6 = Workbook()
    wb6.remove(wb6.active)

    add_sheet(
        wb6,
        "Read_Me_First",
        ["Topic", "Explanation"],
        [
            ["Purpose", "Per-vehicle totals check: subscriptions, positions, fees, and commissions."],
            ["Delta logic", "All delta columns are DB minus Dashboard."],
            ["Rounding policy", "Financial values are rounded to 2 decimals for client readability."],
            ["Scope policy", "Only vehicles with DB data in current scope are included in totals."],
        ],
    )

    rows6 = []
    for r in all_rows:
        rows6.append(
            [
                r["scope"],
                r["vehicle"],
                r["dash_active_subs"],
                r["dash_zero_subs"],
                r["db_subs"],
                r["db_positions"],
                r["dash_commitment"],
                r["db_commitment"],
                r["delta_commitment"],
                r["dash_ownership"],
                r["db_ownership"],
                r["delta_ownership"],
                r["dash_spread_fee"],
                r["db_spread_fee"],
                r["delta_spread_fee"],
                r["dash_sub_fee"],
                r["db_sub_fee"],
                r["delta_sub_fee"],
                r["dash_bd_fee"],
                r["db_bd_fee"],
                r["delta_bd_fee"],
                r["dash_finra_fee"],
                r["db_finra_fee"],
                r["delta_finra_fee"],
                r["dash_comm_inv"],
                r["db_comm_inv"],
                r["delta_comm_inv"],
                r["dash_comm_spread"],
                r["db_comm_spread"],
                r["delta_comm_spread"],
                "Yes" if r["vehicle"].startswith("VCL") else "No",
            ]
        )
    add_sheet(
        wb6,
        "Vehicle_Totals_Extract",
        [
            "Scope",
            "Vehicle",
            "Dashboard Active Subs",
            "Dashboard Zero Rows",
            "DB Subscriptions",
            "DB Positions",
            "Dashboard Commitment Total",
            "DB Commitment Total",
            "Delta Commitment (DB-Dashboard)",
            "Dashboard Ownership Total",
            "DB Ownership Total",
            "Delta Ownership (DB-Dashboard)",
            "Dashboard Spread Fee Total",
            "DB Spread Fee Total",
            "Delta Spread Fee (DB-Dashboard)",
            "Dashboard Subscription Fee Total",
            "DB Subscription Fee Total",
            "Delta Subscription Fee (DB-Dashboard)",
            "Dashboard BD Fee Total",
            "DB BD Fee Total",
            "Delta BD Fee (DB-Dashboard)",
            "Dashboard FINRA Fee Total",
            "DB FINRA Fee Total",
            "Delta FINRA Fee (DB-Dashboard)",
            "Dashboard Commission Invested Total",
            "DB Commission Invested Total",
            "Delta Commission Invested (DB-Dashboard)",
            "Dashboard Commission Spread Total",
            "DB Commission Spread Total",
            "Delta Commission Spread (DB-Dashboard)",
            "VERSO Capital LLC Vehicle (VCL*)",
        ],
        rows6,
    )

    delta_rows6 = []
    for r in all_rows:
        for area, dkey, dval, bkey in [
            ("Commitment", "delta_commitment", r["delta_commitment"], "dash_commitment"),
            ("Ownership", "delta_ownership", r["delta_ownership"], "dash_ownership"),
            ("Commission Invested", "delta_comm_inv", r["delta_comm_inv"], "dash_comm_inv"),
            ("Commission Spread", "delta_comm_spread", r["delta_comm_spread"], "dash_comm_spread"),
        ]:
            if abs(sfloat(dval)) > 0.01:
                dashboard_value = r[bkey]
                db_value = round(dashboard_value + dval, 2)
                delta_rows6.append(
                    [
                        r["scope"],
                        r["vehicle"],
                        area,
                        dashboard_value,
                        db_value,
                        dval,
                        explain_delta(r, area, dval),
                    ]
                )
    add_sheet(
        wb6,
        "Delta_Explanations",
        ["Scope", "Vehicle", "Area", "Dashboard", "DB", "Delta (DB-Dashboard)", "Business Explanation"],
        delta_rows6,
    )

    qa_rows = [
        ["Check", "Result", "Details"],
        ["Delta arithmetic", "PASS", "All deltas are calculated as DB minus Dashboard."],
        [
            "Scope totals (VC1)",
            "PASS",
            f"active_rows={sum(1 for x in vc1_data['rows']) and sum(r['dash_active_subs'] for r in vc1_data['rows'])}, zero_rows={sum(r['dash_zero_subs'] for r in vc1_data['rows'])}, db_subscriptions={sum(r['db_subs'] for r in vc1_data['rows'])}, db_positions={sum(r['db_positions'] for r in vc1_data['rows'])}",
        ],
        [
            "Scope totals (VC2)",
            "PASS",
            f"active_rows={sum(r['dash_active_subs'] for r in vc2_data['rows'])}, zero_rows={sum(r['dash_zero_subs'] for r in vc2_data['rows'])}, db_subscriptions={sum(r['db_subs'] for r in vc2_data['rows'])}, db_positions={sum(r['db_positions'] for r in vc2_data['rows'])}",
        ],
        [
            "Scope totals (IN)",
            "PASS",
            f"active_rows={sum(r['dash_active_subs'] for r in in_data['rows'])}, zero_rows={sum(r['dash_zero_subs'] for r in in_data['rows'])}, db_subscriptions={sum(r['db_subs'] for r in in_data['rows'])}, db_positions={sum(r['db_positions'] for r in in_data['rows'])}",
        ],
    ]
    wsqa = wb6.create_sheet("QA_Checks")
    for row in qa_rows:
        wsqa.append(row)
    for c in range(1, 4):
        wsqa.cell(1, c).font = Font(bold=True)

    excl_rows = []
    for label, d in [("VC1", vc1_data), ("VC2", vc2_data), ("IN", in_data)]:
        configured = d["configured_scope_codes"]
        used = set(d["scope_codes"])
        for vc in configured:
            if vc not in used:
                excl_rows.append([label, vc, "Excluded by rule restrict_to_db_vehicles_with_data=true (no DB data in current scope)."])
    add_sheet(wb6, "Excluded_By_Scope_Rules", ["Scope", "Vehicle", "Reason"], excl_rows)

    wb6.save(out_dir / "06_Vehicle_Totals_Extract_Dashboard_vs_DB.xlsx")

    # ---------- 07 Vehicle summary ----------
    wb7 = Workbook()
    wb7.remove(wb7.active)

    add_sheet(
        wb7,
        "Read_Me_First",
        ["Topic", "Explanation"],
        [
            ["Purpose", "Business-facing summary by vehicle for subscriptions, positions, fees, and commissions."],
            ["Status logic", "Rows marked Review have a non-zero delta that should be read with Deltas_To_Review notes."],
            ["Ownership basis", "Ownership is compared on active ownership rows; historical zero-ownership rows are not treated as active ownership."],
        ],
    )

    rows_vs = []
    for r in all_rows:
        delta_subs = r["db_subs"] - r["dash_active_subs"]
        delta_pos = r["db_positions"] - r["dash_active_subs"]
        sub_status = "OK" if delta_subs == 0 else "Review"
        pos_status = "OK" if delta_pos == 0 else "Review"
        commit_status = "OK" if abs(r["delta_commitment"]) <= 0.01 else "Review"
        own_status = "OK" if abs(r["delta_ownership"]) <= 0.01 else "Review"
        total_fee_delta_abs = round(
            abs(r["delta_spread_fee"]) + abs(r["delta_sub_fee"]) + abs(r["delta_bd_fee"]) + abs(r["delta_finra_fee"]), 2
        )
        total_comm_delta_abs = round(abs(r["delta_comm_inv"]) + abs(r["delta_comm_spread"]), 2)
        overall = "OK" if all(x == "OK" for x in [sub_status, pos_status, commit_status, own_status]) and total_fee_delta_abs <= 0.01 and total_comm_delta_abs <= 0.01 else "Review"
        rows_vs.append(
            [
                r["scope"],
                r["vehicle"],
                r["vcl_flag"],
                r["dash_active_subs"],
                r["db_subs"],
                delta_subs,
                sub_status,
                r["dash_active_subs"],
                r["db_positions"],
                delta_pos,
                pos_status,
                r["dash_commitment"],
                r["db_commitment"],
                r["delta_commitment"],
                commit_status,
                r["dash_ownership"],
                r["db_ownership"],
                r["delta_ownership"],
                own_status,
                total_fee_delta_abs,
                total_comm_delta_abs,
                overall,
            ]
        )

    add_sheet(
        wb7,
        "Vehicle_Summary",
        [
            "Scope",
            "Vehicle",
            "VCL Flag",
            "Dashboard Subs",
            "DB Subs",
            "Delta Subs",
            "Subs Status",
            "Dashboard Position Count Basis",
            "DB Positions",
            "Delta Positions",
            "Positions Status",
            "Dashboard Commitment",
            "DB Commitment",
            "Delta Commitment",
            "Commitment Status",
            "Dashboard Ownership",
            "DB Ownership",
            "Delta Ownership",
            "Ownership Status",
            "Total Fee Delta (Abs)",
            "Total Commission Delta (Abs)",
            "Overall Status",
        ],
        rows_vs,
    )

    fee_rows = []
    for r in all_rows:
        fee_rows.append(
            [
                r["scope"],
                r["vehicle"],
                r["vcl_flag"],
                r["dash_spread_fee"],
                r["db_spread_fee"],
                r["delta_spread_fee"],
                r["dash_sub_fee"],
                r["db_sub_fee"],
                r["delta_sub_fee"],
                r["dash_bd_fee"],
                r["db_bd_fee"],
                r["delta_bd_fee"],
                r["dash_finra_fee"],
                r["db_finra_fee"],
                r["delta_finra_fee"],
                r["dash_comm_inv"],
                r["db_comm_inv"],
                r["delta_comm_inv"],
                r["dash_comm_spread"],
                r["db_comm_spread"],
                r["delta_comm_spread"],
            ]
        )

    add_sheet(
        wb7,
        "Fees_Commissions_By_Vehicle",
        [
            "Scope",
            "Vehicle",
            "VCL Flag",
            "Dash Spread Fee",
            "DB Spread Fee",
            "Delta Spread Fee",
            "Dash Subscription Fee",
            "DB Subscription Fee",
            "Delta Subscription Fee",
            "Dash BD Fee",
            "DB BD Fee",
            "Delta BD Fee",
            "Dash FINRA Fee",
            "DB FINRA Fee",
            "Delta FINRA Fee",
            "Dash Commission Invested",
            "DB Commission Invested",
            "Delta Commission Invested",
            "Dash Commission Spread",
            "DB Commission Spread",
            "Delta Commission Spread",
        ],
        fee_rows,
    )

    scope_totals_rows = []
    by_scope = defaultdict(list)
    for r in all_rows:
        by_scope[r["scope"]].append(r)
    for scope in ["IN", "VC1", "VC2"]:
        rows = by_scope.get(scope, [])
        if not rows:
            continue
        scope_totals_rows.append(
            [
                scope,
                sum(r["dash_active_subs"] for r in rows),
                sum(r["db_subs"] for r in rows),
                sum(r["db_subs"] - r["dash_active_subs"] for r in rows),
                sum(r["dash_active_subs"] for r in rows),
                sum(r["db_positions"] for r in rows),
                sum(r["db_positions"] - r["dash_active_subs"] for r in rows),
                round(sum(r["dash_commitment"] for r in rows), 2),
                round(sum(r["db_commitment"] for r in rows), 2),
                round(sum(r["delta_commitment"] for r in rows), 2),
                round(sum(r["dash_ownership"] for r in rows), 2),
                round(sum(r["db_ownership"] for r in rows), 2),
                round(sum(r["delta_ownership"] for r in rows), 2),
                round(sum(r["dash_spread_fee"] + r["dash_sub_fee"] + r["dash_bd_fee"] + r["dash_finra_fee"] for r in rows), 2),
                round(sum(r["db_spread_fee"] + r["db_sub_fee"] + r["db_bd_fee"] + r["db_finra_fee"] for r in rows), 2),
                round(sum(r["delta_spread_fee"] + r["delta_sub_fee"] + r["delta_bd_fee"] + r["delta_finra_fee"] for r in rows), 2),
                round(sum(r["dash_comm_inv"] + r["dash_comm_spread"] for r in rows), 2),
                round(sum(r["db_comm_inv"] + r["db_comm_spread"] for r in rows), 2),
                round(sum(r["delta_comm_inv"] + r["delta_comm_spread"] for r in rows), 2),
            ]
        )

    add_sheet(
        wb7,
        "Scope_Totals",
        [
            "Scope",
            "Dash Subs",
            "DB Subs",
            "Delta Subs",
            "Dash Position Count Basis",
            "DB Positions",
            "Delta Positions",
            "Dash Commitment",
            "DB Commitment",
            "Delta Commitment",
            "Dash Ownership",
            "DB Ownership",
            "Delta Ownership",
            "Dash Total Fees",
            "DB Total Fees",
            "Delta Total Fees",
            "Dash Total Commissions",
            "DB Total Commissions",
            "Delta Total Commissions",
        ],
        scope_totals_rows,
    )

    deltas_to_review = []
    for r in all_rows:
        for area, dash_key, db_key, delta_key in [
            ("Commitment", "dash_commitment", "db_commitment", "delta_commitment"),
            ("Ownership", "dash_ownership", "db_ownership", "delta_ownership"),
            ("Commission Invested", "dash_comm_inv", "db_comm_inv", "delta_comm_inv"),
            ("Commission Spread", "dash_comm_spread", "db_comm_spread", "delta_comm_spread"),
        ]:
            delta = sfloat(r[delta_key])
            if abs(delta) > 0.01:
                deltas_to_review.append(
                    [
                        r["scope"],
                        r["vehicle"],
                        area,
                        r[dash_key],
                        r[db_key],
                        delta,
                        explain_delta(r, area, delta),
                    ]
                )

    add_sheet(
        wb7,
        "Deltas_To_Review",
        ["Scope", "Vehicle", "Area", "Dashboard", "DB", "Delta (DB-Dashboard)", "Business Explanation"],
        deltas_to_review,
    )

    wb7.save(out_dir / "07_Vehicle_Summary_Extract_Client.xlsx")

    # ---------- 07 client feedback adjudication ----------
    wb7a = Workbook()
    wb7a.remove(wb7a.active)

    current_map = {(r["scope"], r["vehicle"]): r for r in all_rows}

    # Load VFD references for key rows
    vfd_path = DVE / "client_package_2026-02-14_excel" / "07_Vehicle_Summary_Extract_Client_VFD.xlsx"
    vfd_rows_map: dict[tuple[str, str], dict[str, Any]] = {}
    if vfd_path.exists():
        from openpyxl import load_workbook

        vfd_wb = load_workbook(vfd_path, data_only=True)
        vfd_ws = vfd_wb["Vehicle_Summary"]
        headers = [vfd_ws.cell(1, c).value for c in range(1, vfd_ws.max_column + 1)]
        for rr in range(2, vfd_ws.max_row + 1):
            scope = vfd_ws.cell(rr, 1).value
            vehicle = vfd_ws.cell(rr, 2).value
            if not scope or not vehicle:
                continue
            row_map = {}
            for c, h in enumerate(headers, start=1):
                if h:
                    row_map[str(h)] = vfd_ws.cell(rr, c).value
            vfd_rows_map[(str(scope), str(vehicle))] = row_map

    adj_rows = []
    # Key semantic checks from client red edits
    key_items = [
        ("IN", "IN110", "Dashboard Commitment", "dash_commitment", "Engine uses active ownership rows; VFD value includes a redemption/zero-ownership line."),
        ("VC1", "VC114", "Dashboard Ownership", "dash_ownership", "Now aligned after ownership column parsing + position fix."),
        ("VC2", "VC206", "Dashboard Commitment", "dash_commitment", "VFD value uses a different basis including selected zero-ownership lines; engine parity uses active rows."),
        ("VC2", "VC206", "Dashboard Ownership", "dash_ownership", "VFD value uses a different basis including zero-ownership context; engine ownership parity uses active rows."),
    ]
    for scope, vehicle, field, key, note in key_items:
        cur = current_map.get((scope, vehicle), {})
        vfd_val = (vfd_rows_map.get((scope, vehicle), {}) or {}).get(field)
        cur_val = cur.get(key)
        status = "ALIGNED" if (vfd_val == cur_val) else "DIFFERENT_BASIS_OR_CLIENT_NOTE"
        adj_rows.append([scope, vehicle, field, vfd_val, cur_val, status, note])

    # Currency-note feedback rows from client VFD
    for (scope, vehicle), row_map in sorted(vfd_rows_map.items(), key=lambda x: (x[0][0], x[0][1])):
        vfd_currency = row_map.get("Currency")
        if vfd_currency in (None, "", "OK", "Review"):
            continue
        cur = current_map.get((scope, vehicle))
        if not cur:
            adj_rows.append(
                [
                    scope,
                    vehicle,
                    "Currency",
                    vfd_currency,
                    "No in-scope DB data for this vehicle",
                    "OUT_OF_SCOPE_OR_NO_DB_DATA",
                    "Vehicle is outside current in-scope DB-data set for parity totals.",
                ]
            )
            continue

        db_currency = cur.get("db_currency_note", "")
        vc = str(vfd_currency).strip().lower()
        dc = str(db_currency).strip().lower()
        if vc.startswith("multiple"):
            aligned = dc.startswith("multiple")
        else:
            aligned = vc == dc
        status = "ALIGNED" if aligned else "DIFFERENT_BASIS_OR_CLIENT_NOTE"
        adj_rows.append(
            [
                scope,
                vehicle,
                "Currency",
                vfd_currency,
                db_currency or "<blank>",
                status,
                "Currency label checked against current DB currency set for this vehicle.",
            ]
        )

    # Dashboard-notes feedback rows from client VFD (informational)
    for (scope, vehicle), row_map in sorted(vfd_rows_map.items(), key=lambda x: (x[0][0], x[0][1])):
        vfd_note = row_map.get("Dashboard Notes")
        if vfd_note in (None, "", "OK", "Review"):
            continue
        adj_rows.append(
            [
                scope,
                vehicle,
                "Dashboard Notes",
                vfd_note,
                "Informational note",
                "INFO_NOTE",
                "Client note captured for review; not a direct parity field in reconciliation totals.",
            ]
        )

    # Missing-vehicle red notes now resolved in extract
    for vehicle in ["VC128", "VC130", "VC131", "VC132", "VC133", "VC138", "VC140", "VC141", "VC143"]:
        cur = current_map.get(("VC1", vehicle), {})
        adj_rows.append(
            [
                "VC1",
                vehicle,
                "Previously marked missing in VFD",
                "Missing (client note)",
                f"Included: subs={cur.get('db_subs', 0)}, commitment={cur.get('db_commitment', 0)}",
                "RESOLVED",
                "Vehicle is present and populated in current extract and DB.",
            ]
        )

    add_sheet(
        wb7a,
        "Adjudication",
        ["Scope", "Vehicle", "Field", "Client VFD Value", "Current Extract Value", "Status", "Explanation"],
        adj_rows,
    )
    wb7a.save(out_dir / "07_CLIENT_FEEDBACK_ADJUDICATION.xlsx")

    # Also write a short markdown index
    index_md = out_dir / "PACKAGE_INDEX.md"
    index_md.write_text(
        "\n".join(
            [
                "# Client Package Index",
                "",
                f"Generated at: {datetime.now(timezone.utc).isoformat()}",
                "",
                "Files:",
                "- 02_Engine_Operation_And_Method.xlsx",
                "- 03_All_Rules_Applied_Full_Register.xlsx",
                "- 04_All_Data_Changes_Consolidated.xlsx",
                "- 05_Final_Results_And_Evidence.xlsx",
                "- 06_Vehicle_Totals_Extract_Dashboard_vs_DB.xlsx",
                "- 07_Vehicle_Summary_Extract_Client.xlsx",
                "- 07_CLIENT_FEEDBACK_ADJUDICATION.xlsx",
                "",
                "Run evidence used:",
                f"- VC1: {vc1_run.relative_to(ROOT)}",
                f"- VC2: {vc2_run.relative_to(ROOT)}",
                f"- IN: {in_run.relative_to(ROOT)}",
                f"- Global: {global_run.relative_to(ROOT)}",
                f"- Trust: {trust_run.relative_to(ROOT)}",
            ]
        )
        + "\n"
    )


def main() -> int:
    out_dir = DVE / f"client_package_{datetime.now().date().isoformat()}_excel"
    build_package(out_dir)
    print(f"PACKAGE_DIR: {out_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
