#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from itertools import zip_longest
from pathlib import Path
from typing import Any
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[3]
DEFAULT_RULES = ROOT / "data_verification_engine" / "scopes" / "vc1" / "rules_vc1.json"
DEFAULT_OUTDIR = ROOT / "data_verification_engine" / "scopes" / "vc1" / "output"
ENV_PATH = ROOT / ".env.local"
SUPABASE_URL = "https://kagzryotbbnusdcyvqei.supabase.co"


def to_float(v: Any) -> float:
    if v in (None, ""):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if not s:
        return 0.0
    try:
        return float(s)
    except Exception:
        # Handles values like "ETH 15.00", "USD -", "#VALUE!".
        m = re.search(r"[-+]?\d*\.?\d+", s)
        if m:
            try:
                return float(m.group(0))
            except Exception:
                return 0.0
        return 0.0


def is_blank_cell_value(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str):
        return v.strip() == ""
    return False


def parse_date(v: Any) -> str | None:
    if v in (None, ""):
        return None
    if hasattr(v, "date"):
        try:
            return v.date().isoformat()
        except Exception:
            pass
    s = str(v).strip()
    if not s:
        return None
    # Ignore obvious non-date tokens used in dashboard helper cells.
    if normalize_text(s) in {"isin", "check", "todo", "n a", "na"}:
        return None
    # ISO/date-like guard only; avoid carrying arbitrary strings as "dates".
    if not re.search(r"\d", s):
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            pass
    # If it is not parseable as date, treat as missing rather than a literal.
    return None


def normalize_text(s: str) -> str:
    s = s.lower().strip()
    s = s.replace("&", " and ")
    s = s.replace("–", "-").replace("—", "-")
    s = re.sub(r"\b(mr|mrs|ms|dr|sir|madam|mme|m)\.?\b", " ", s)
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalize_currency_token(v: Any) -> str:
    if v in (None, ""):
        return ""
    s = str(v).strip().upper()
    if not s:
        return ""
    symbol_map = {
        "$": "USD",
        "US$": "USD",
        "€": "EUR",
        "£": "GBP",
    }
    if s in symbol_map:
        return symbol_map[s]
    m = re.search(r"\b[A-Z]{3,5}\b", s)
    if m:
        return m.group(0)
    return ""


def canonical_name_key(s: str) -> str:
    n = normalize_text(s)
    parts = []
    for p in n.split(" "):
        if not p or p in {"and", "or"}:
            continue
        if len(p) > 4 and p.endswith("s"):
            p = p[:-1]
        parts.append(p)
    parts.sort()
    return "".join(parts)


def compact_name_key(s: str) -> str:
    n = normalize_text(s)
    toks = [t for t in n.split(" ") if t and t not in {"and", "or"}]
    return "".join(toks)


def alias_key_variants(s: str) -> set[str]:
    out = set()
    c = canonical_name_key(s)
    if c:
        out.add(c)
    k = compact_name_key(s)
    if k:
        out.add(k)
    l = loose_name_key(s)
    if l:
        out.add(l)
    return out


def name_key(s: str, aliases: dict[str, str]) -> str:
    c = canonical_name_key(s)
    if not c:
        return c
    for k in (c, compact_name_key(s), loose_name_key(s)):
        if not k:
            continue
        alias_target = aliases.get(k)
        if alias_target:
            return alias_target
    return c


def loose_name_key(s: str) -> str:
    # Used only for fallback matching when strict normalized keys differ
    # because of middle names/initials in dashboard vs DB.
    n = normalize_text(s)
    toks = [t for t in n.split(" ") if t and t not in {"and", "or"}]
    if not toks:
        return ""
    if len(toks) == 1:
        return toks[0]
    return toks[0] + toks[-1]


def individual_loose_identity_key(s: str) -> str:
    # Duplicate-person heuristic: first+last after normalization.
    return loose_name_key(s)


def values_match(
    metric: str,
    dashboard_value: float,
    db_value: float,
    tol: float = 0.01,
    tolerance_overrides: dict[str, float] | None = None,
    percent_compare_rules: dict[str, Any] | None = None,
) -> bool:
    if tolerance_overrides and metric in tolerance_overrides:
        tol = float(tolerance_overrides[metric])
    if abs(db_value - dashboard_value) <= tol:
        return True
    percent_metrics = {
        "sub_fee_percent",
        "management_fee_percent",
        "perf1_percent",
        "perf2_percent",
        "bd_fee_percent",
    }
    if metric in percent_metrics:
        cfg = (percent_compare_rules or {}).get(metric, {})
        mode = str(cfg.get("mode", "fraction_or_percent_strict")).strip().lower()
        divide_threshold = float(cfg.get("allow_divide_if_abs_gte", 100.0))

        def percent_candidates(v: float) -> set[float]:
            if mode == "exact":
                return {v}
            out = {v}
            # Typical normalization: 0.2 (fraction) <-> 20 (percent)
            if abs(v) <= 1.0:
                out.add(v * 100.0)
            # Optional safeguard for very large loaded values only.
            if abs(v) >= divide_threshold:
                out.add(v / 100.0)
            return out

        d_candidates = percent_candidates(dashboard_value)
        b_candidates = percent_candidates(db_value)
        for dv in d_candidates:
            for bv in b_candidates:
                if abs(bv - dv) <= tol:
                    return True
    return False


def load_env(path: Path) -> dict[str, str]:
    out: dict[str, str] = {}
    if not path.exists():
        return out
    for raw in path.read_text().splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        out[k.strip()] = v.strip().strip('"')
    return out


def api_get_all(table: str, select: str, key: str, filters: dict[str, str] | None = None) -> list[dict[str, Any]]:
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Accept": "application/json",
    }
    out: list[dict[str, Any]] = []
    offset = 0
    limit = 1000
    while True:
        params = {"select": select, "order": "id", "limit": str(limit), "offset": str(offset)}
        if filters:
            params.update(filters)
        url = f"{SUPABASE_URL}/rest/v1/{table}?{urlencode(params)}"
        req = Request(url, headers=headers)
        try:
            with urlopen(req, timeout=120) as r:
                payload = r.read().decode("utf-8")
        except Exception as exc:
            body = ""
            if hasattr(exc, "read"):
                try:
                    body = exc.read().decode("utf-8", errors="ignore")
                except Exception:
                    body = ""
            raise RuntimeError(f"API error for {table} offset={offset}: {body}") from exc
        rows = json.loads(payload)
        if not rows:
            break
        out.extend(rows)
        if len(rows) < limit:
            break
        offset += limit
    return out


def find_col(headers: list[Any], candidates: list[str]) -> int | None:
    # Pass 1: strict raw text match (preserves '%' distinction)
    raw_headers = [str(h).strip().lower() if isinstance(h, str) else "" for h in headers]
    raw_candidates = [str(c).strip().lower() for c in candidates]
    for i, h in enumerate(raw_headers, start=1):
        if h in raw_candidates:
            return i

    # Pass 2: normalized fallback for minor punctuation/spacing drift
    norm_headers = [normalize_text(h) if isinstance(h, str) else "" for h in headers]
    norm_candidates = [normalize_text(c) for c in candidates]
    for i, h in enumerate(norm_headers, start=1):
        if h in norm_candidates:
            return i
    return None


@dataclass
class DashRow:
    vehicle: str
    sheet: str
    row_num: int
    investor_name: str
    investor_key: str
    currency: str
    commitment: float
    cost_per_share: float
    price_per_share: float
    shares: float
    ownership: float
    contract_date: str | None
    spread_per_share: float
    spread_fee: float
    sub_fee_percent: float
    sub_fee: float
    management_fee_percent: float
    perf1_percent: float
    perf1_threshold: float
    perf2_percent: float
    perf2_threshold: float
    bd_fee_percent: float
    bd_fee: float
    finra_shares: float
    finra_fee: float
    raw_numeric: dict[str, Any]


@dataclass
class DbSub:
    id: str
    deal_id: str | None
    vehicle: str
    investor_id: str
    investor_name: str
    investor_key: str
    currency: str
    commitment: float
    cost_per_share: float
    price_per_share: float
    shares: float
    units: float
    contract_date: str | None
    spread_per_share: float
    spread_fee: float
    sub_fee_percent: float
    sub_fee: float
    management_fee_percent: float
    perf1_percent: float
    perf1_threshold: float
    perf2_percent: float
    perf2_threshold: float
    bd_fee_percent: float
    bd_fee: float
    finra_shares: float
    finra_fee: float
    status: str
    funded_amount: float


@dataclass
class DashCommission:
    vehicle: str
    sheet: str
    row_num: int
    investor_name: str
    investor_key: str
    introducer_name: str
    introducer_key: str
    basis_type: str
    amount: float


@dataclass
class DbCommission:
    id: str
    vehicle: str
    investor_name: str
    investor_key: str
    introducer_name: str
    introducer_key: str
    basis_type: str
    amount: float
    rate_bps: int | None


class Auditor:
    def __init__(self, rules: dict[str, Any], outdir: Path):
        self.rules = rules
        self.outdir = outdir
        raw_aliases = rules.get("name_aliases", {}) or {}
        normalized_aliases: dict[str, str] = {}
        for k, v in raw_aliases.items():
            vv = canonical_name_key(str(v))
            if not vv:
                continue
            # Map multiple key variants to canonical target. This avoids false misses
            # when rules are written as compact keys (e.g. "ericpascalleseigneur").
            for kk in alias_key_variants(str(k)):
                normalized_aliases[kk] = vv
            # Also self-normalize target variants for stable canonicalization.
            for kk in alias_key_variants(str(v)):
                normalized_aliases.setdefault(kk, vv)
        self.aliases = normalized_aliases
        # Optional vehicle-scoped investor aliasing for known transfer/rename rows.
        # Shape in rules:
        # "investor_aliases_by_vehicle": { "VC111": { "Zandera (Finco) Limited": "Michael RYAN" } }
        raw_by_vehicle = rules.get("investor_aliases_by_vehicle", {}) or {}
        normalized_by_vehicle: dict[str, dict[str, str]] = {}
        for vc, mapping in raw_by_vehicle.items():
            if not isinstance(mapping, dict):
                continue
            dst: dict[str, str] = {}
            for k, v in mapping.items():
                vv = canonical_name_key(str(v))
                if not vv:
                    continue
                for kk in alias_key_variants(str(k)):
                    dst[kk] = vv
            if dst:
                normalized_by_vehicle[str(vc)] = dst
        self.investor_aliases_by_vehicle = normalized_by_vehicle
        raw_fallback_pairs = rules.get("fallback_ruled_name_pairs", []) or []
        normalized_fallback_pairs: set[tuple[str, str, str]] = set()
        for item in raw_fallback_pairs:
            if not isinstance(item, dict):
                continue
            vc = str(item.get("vehicle") or "*").strip() or "*"
            dname = canonical_name_key(str(item.get("dash_investor") or ""))
            bname = canonical_name_key(str(item.get("db_investor") or ""))
            if dname and bname:
                normalized_fallback_pairs.add((vc, dname, bname))
        self.fallback_ruled_pairs = normalized_fallback_pairs
        fallback_lookup: dict[tuple[str, str], set[str]] = defaultdict(set)
        for vc, dname, bname in normalized_fallback_pairs:
            fallback_lookup[(vc, dname)].add(bname)
        self.fallback_dash_to_db = fallback_lookup
        self.failures: list[dict[str, str]] = []
        self.warnings: list[dict[str, str]] = []

    def add_failure(self, check: str, vehicle: str, details: str, row_ref: str = ""):
        self.failures.append(
            {"severity": "fail", "check": check, "vehicle": vehicle, "row_ref": row_ref, "details": details}
        )

    def add_warning(self, check: str, vehicle: str, details: str, row_ref: str = ""):
        self.warnings.append(
            {"severity": "warn", "check": check, "vehicle": vehicle, "row_ref": row_ref, "details": details}
        )

    def investor_key_for_vehicle(self, vehicle: str, investor_name: str) -> str:
        vmap = self.investor_aliases_by_vehicle.get(vehicle, {})
        for kk in alias_key_variants(investor_name):
            tgt = vmap.get(kk)
            if tgt:
                return tgt
        return name_key(investor_name, self.aliases)

    def apply_ruled_fallback_investor(self, vehicle: str, investor_key: str) -> str:
        candidates = set()
        candidates.update(self.fallback_dash_to_db.get((vehicle, investor_key), set()))
        candidates.update(self.fallback_dash_to_db.get(("*", investor_key), set()))
        if len(candidates) == 1:
            return next(iter(candidates))
        return investor_key

    def load_dashboard_rows(
        self,
    ) -> tuple[
        list[DashRow],
        list[DashRow],
        dict[str, dict[str, float]],
        dict[str, dict[str, float]],
        list[DashCommission],
        set[tuple[str, str, str]],
    ]:
        dash_path = ROOT / self.rules["dashboard_files"]["main"]
        scope_sheets = set(self.rules["scope_dashboard_sheets"])
        scope_codes = set(self.rules["scope_vehicle_codes"])
        alias_map = self.rules.get("dashboard_vehicle_alias", {})
        alias_by_sheet = self.rules.get("dashboard_vehicle_alias_by_sheet", {})
        header_row_by_sheet = self.rules.get("dashboard_header_row_by_sheet", {})
        data_start_row_by_sheet = self.rules.get("dashboard_data_start_row_by_sheet", {})
        ownership_fallback_to_shares = bool(self.rules.get("dashboard_ownership_fallback_to_shares_if_missing", False))
        blank_ownership_active_by_sheet = set(self.rules.get("dashboard_blank_ownership_as_active_by_sheet", []))
        optional_numeric_fields = set(self.rules.get("dashboard_optional_numeric_fields", []))
        combined_intro_policy = str(self.rules.get("combined_introducer_name_policy", "fail")).strip().lower()
        comm_cols = self.rules.get("dashboard_commission_columns", {})
        comm_default = comm_cols.get("default", {})
        comm_by_sheet = comm_cols.get("by_sheet", {})
        default_invested_cols = tuple(comm_default.get("invested_amount", [35, 45, 55]))
        default_spread_cols = tuple(comm_default.get("spread", [41, 51, 61]))
        comm_slots_cfg = self.rules.get("dashboard_commission_slots", {})
        default_slot_cfg = comm_slots_cfg.get("default", [])
        slot_cfg_by_sheet = comm_slots_cfg.get("by_sheet", {})
        dynamic_slots_enabled = bool(self.rules.get("dashboard_dynamic_commission_slots", False))
        dynamic_name_headers = set(
            normalize_text(x)
            for x in self.rules.get(
                "dashboard_commission_name_header_candidates",
                ["name", "names", "partner", "partners", "bi", "verso", "introducer", "introducers"],
            )
        )
        currency_header_candidates = self.rules.get(
            "dashboard_currency_header_candidates",
            ["Currency", "Deal Currency", "CCY", "Curr"],
        )
        comm_name_cols = self.rules.get("dashboard_commission_name_columns", {})
        default_name_cols = tuple(comm_name_cols.get("default", [43, 53]))
        name_by_sheet = comm_name_cols.get("by_sheet", {})

        wb = load_workbook(dash_path, data_only=True, read_only=True)
        active: list[DashRow] = []
        zero: list[DashRow] = []
        dash_commissions: list[DashCommission] = []
        combined_commission_skip_keys: set[tuple[str, str, int, str]] = set()
        totals = defaultdict(lambda: defaultdict(float))
        intro_totals_all_rows = defaultdict(lambda: defaultdict(float))

        for sheet in wb.sheetnames:
            if sheet not in scope_sheets:
                continue
            ws = wb[sheet]
            header_row = int(header_row_by_sheet.get(sheet, 2))
            data_start_row = int(data_start_row_by_sheet.get(sheet, header_row + 1))
            headers = [ws.cell(header_row, c).value for c in range(1, 120)]
            col = {
                "inv_last": find_col(headers, ["Investor Last Name"]),
                "inv_mid": find_col(headers, ["Investor Middle Name"]),
                "inv_first": find_col(headers, ["Investor First Name"]),
                "inv_entity": find_col(headers, ["Investor Entity"]),
                "vehicle": find_col(headers, ["Vehicle"]),
                "currency": find_col(headers, currency_header_candidates),
                "cost_per_share": find_col(headers, ["Cost per Share"]),
                "amount": find_col(headers, ["Amount invested"]),
                "price_per_share": find_col(headers, ["Price per Share", "Price per Note"]),
                "shares": find_col(headers, ["Number of shares invested"]),
                "ownership": find_col(headers, ["OWNERSHIP", "OWNERSHIP POSITION", "Position", "Ownership Position"]),
                "date": find_col(headers, ["Contract Date"]),
                "spread_per_share": find_col(headers, ["Spread PPS"]),
                "spread_fee": find_col(headers, ["Spread PPS Fees"]),
                "sub_fee_percent": find_col(headers, ["Subscription fees %"]),
                "sub_fee": find_col(headers, ["Subscription fees"]),
                "management_fee_percent": find_col(headers, ["Management fees", "Management Fee"]),
                "perf1_percent": find_col(headers, ["Performance fees 1"]),
                "perf1_threshold": find_col(headers, ["Threshold 1"]),
                "perf2_percent": find_col(headers, ["Performance fees 2"]),
                "perf2_threshold": find_col(headers, ["Threshold 2"]),
                "bd_fee_percent": find_col(headers, ["BD Fees %"]),
                "bd_fee": find_col(headers, ["BD fees"]),
                "finra_shares": find_col(headers, ["FINRA fees in share"]),
                "finra_fee": find_col(headers, ["FINRA fees"])
            }
            required = [
                "vehicle",
                "cost_per_share",
                "amount",
                "price_per_share",
                "shares",
                "date",
                "spread_per_share",
                "spread_fee",
                "sub_fee_percent",
                "sub_fee",
                "perf1_percent",
                "perf1_threshold",
                "perf2_percent",
                "perf2_threshold",
                "management_fee_percent",
                "bd_fee_percent",
                "bd_fee",
                "finra_shares",
                "finra_fee",
            ]
            if not ownership_fallback_to_shares:
                required.append("ownership")
            required = [k for k in required if k not in optional_numeric_fields]
            if not all(col.get(k) for k in required):
                self.add_failure("dashboard_columns_missing", sheet, f"missing columns in sheet {sheet}")
                continue

            sheet_comm = comm_by_sheet.get(sheet, {})
            invested_cols = tuple(sheet_comm.get("invested_amount", default_invested_cols))
            spread_cols = tuple(sheet_comm.get("spread", default_spread_cols))
            name_cols = tuple(name_by_sheet.get(sheet, default_name_cols))
            slot_cfg = slot_cfg_by_sheet.get(sheet, default_slot_cfg)

            commission_slots: list[tuple[int, int, int]] = []
            if slot_cfg:
                for s in slot_cfg:
                    try:
                        nidx = int(s.get("name_col"))
                        iidx = int(s.get("invested_amount_col") or 0)
                        sidx = int(s.get("spread_col") or 0)
                        if iidx or sidx:
                            commission_slots.append((nidx, iidx, sidx))
                    except Exception:
                        continue
            elif dynamic_slots_enabled:
                hdr_norm = [normalize_text(str(h or "")) for h in headers]
                name_positions = [i + 1 for i, h in enumerate(hdr_norm) if h in dynamic_name_headers]
                for p_idx, nidx in enumerate(name_positions):
                    next_name = name_positions[p_idx + 1] if p_idx + 1 < len(name_positions) else len(headers) + 1
                    iidx = None
                    sidx = None
                    for cidx in range(nidx + 1, next_name):
                        h_raw = str(headers[cidx - 1] or "")
                        h_norm = normalize_text(h_raw)
                        if iidx is None and "subscription fees" in h_norm and "%" not in h_raw:
                            iidx = cidx
                        if sidx is None and "spread pps fees" in h_norm:
                            sidx = cidx
                    if iidx or sidx:
                        commission_slots.append((nidx, iidx or 0, sidx or 0))
            else:
                # Backward-compatible fallback:
                # pair N name columns with the rightmost N amount columns.
                # Example default: names [43,53], invested [35,45,55] -> pair to [45,55].
                n = min(len(name_cols), len(invested_cols), len(spread_cols))
                if n > 0:
                    inv_pair = invested_cols[-n:]
                    spr_pair = spread_cols[-n:]
                    for idx in range(n):
                        commission_slots.append((name_cols[idx], inv_pair[idx], spr_pair[idx]))

            # Dashboard uses merged cells heavily for introducer labels.
            # Keep last non-empty introducer per slot column so each investor row
            # still receives its corresponding introducer block.
            carried_intro_name_by_col: dict[int, str] = {}

            # Guard against header drift (reading % instead of amount is a common silent failure).
            # Only check slots we are actually using.
            if commission_slots:
                invested_cols_for_check = tuple(i for _, i, _ in commission_slots if i)
                spread_cols_for_check = tuple(s for _, _, s in commission_slots if s)
                for cidx in invested_cols_for_check:
                    if cidx <= len(headers):
                        h = str(headers[cidx - 1] or "")
                        hn = normalize_text(h)
                        if "subscription fees" not in hn or "%" in h:
                            self.add_warning(
                                "dashboard_commission_header_check",
                                sheet,
                                f"invested_amount_col={cidx} header={h!r}",
                            )
                for cidx in spread_cols_for_check:
                    if cidx <= len(headers):
                        h = str(headers[cidx - 1] or "")
                        hn = normalize_text(h)
                        if "spread pps fees" not in hn:
                            self.add_warning(
                                "dashboard_commission_header_check",
                                sheet,
                                f"spread_col={cidx} header={h!r}",
                            )

            for r_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, max_col=120, values_only=True), start=data_start_row):
                raw_v = row[col["vehicle"] - 1] if len(row) >= col["vehicle"] else None
                if raw_v in (None, ""):
                    continue
                raw_v = str(raw_v).strip()
                mapped_v = alias_by_sheet.get(sheet, alias_map.get(raw_v, raw_v))
                if mapped_v not in scope_codes:
                    continue

                entity = row[col["inv_entity"] - 1] if col["inv_entity"] else None
                last = row[col["inv_last"] - 1] if col["inv_last"] else None
                middle = row[col["inv_mid"] - 1] if col["inv_mid"] else None
                first = row[col["inv_first"] - 1] if col["inv_first"] else None
                if entity not in (None, ""):
                    investor_name = str(entity).strip()
                else:
                    investor_name = " ".join([str(x).strip() for x in (first, middle, last) if x not in (None, "")]).strip()
                if not investor_name:
                    continue

                def cell_float(field: str) -> float:
                    c = col.get(field)
                    if not c:
                        return 0.0
                    return to_float(row[c - 1])

                def cell_raw(field: str) -> Any:
                    c = col.get(field)
                    if not c:
                        return None
                    return row[c - 1]

                shares_val = cell_float("shares")
                ownership_val = cell_float("ownership")
                ownership_raw = row[col["ownership"] - 1] if col.get("ownership") else None
                ownership_missing = ownership_raw in (None, "")
                if col.get("ownership") is None and ownership_fallback_to_shares:
                    ownership_val = shares_val
                    ownership_missing = False

                rec = DashRow(
                    vehicle=mapped_v,
                    sheet=sheet,
                    row_num=r_idx,
                    investor_name=investor_name,
                    investor_key=self.investor_key_for_vehicle(mapped_v, investor_name),
                    currency=normalize_currency_token(row[col["currency"] - 1]) if col.get("currency") else "",
                    cost_per_share=cell_float("cost_per_share"),
                    commitment=cell_float("amount"),
                    price_per_share=cell_float("price_per_share"),
                    shares=shares_val,
                    ownership=ownership_val,
                    contract_date=parse_date(row[col["date"] - 1]) if col.get("date") else None,
                    spread_per_share=cell_float("spread_per_share"),
                    spread_fee=cell_float("spread_fee"),
                    sub_fee_percent=cell_float("sub_fee_percent"),
                    sub_fee=cell_float("sub_fee"),
                    management_fee_percent=cell_float("management_fee_percent"),
                    perf1_percent=cell_float("perf1_percent"),
                    perf1_threshold=cell_float("perf1_threshold"),
                    perf2_percent=cell_float("perf2_percent"),
                    perf2_threshold=cell_float("perf2_threshold"),
                    bd_fee_percent=cell_float("bd_fee_percent"),
                    bd_fee=cell_float("bd_fee"),
                    finra_shares=cell_float("finra_shares"),
                    finra_fee=cell_float("finra_fee"),
                    raw_numeric={
                        "cost_per_share": cell_raw("cost_per_share"),
                        "price_per_share": cell_raw("price_per_share"),
                        "ownership": cell_raw("ownership"),
                        "spread_per_share": cell_raw("spread_per_share"),
                        "spread_fee": cell_raw("spread_fee"),
                        "sub_fee_percent": cell_raw("sub_fee_percent"),
                        "sub_fee": cell_raw("sub_fee"),
                        "management_fee_percent": cell_raw("management_fee_percent"),
                        "perf1_percent": cell_raw("perf1_percent"),
                        "perf1_threshold": cell_raw("perf1_threshold"),
                        "perf2_percent": cell_raw("perf2_percent"),
                        "perf2_threshold": cell_raw("perf2_threshold"),
                        "bd_fee_percent": cell_raw("bd_fee_percent"),
                        "bd_fee": cell_raw("bd_fee"),
                        "finra_shares": cell_raw("finra_shares"),
                        "finra_fee": cell_raw("finra_fee"),
                    },
                )
                if ownership_missing and sheet in blank_ownership_active_by_sheet:
                    # Some sheets (e.g. VC22) legitimately keep subscription rows with blank ownership.
                    # Keep them in row-level subscription checks, but do not count them as "zero ownership".
                    active.append(rec)
                    t = totals[rec.vehicle]
                    t["count"] += 1
                    t["cost_per_share"] += rec.cost_per_share
                    t["commitment"] += rec.commitment
                    t["price_per_share"] += rec.price_per_share
                    t["shares"] += rec.shares
                    t["ownership"] += rec.ownership
                    t["spread_per_share"] += rec.spread_per_share
                    t["spread_fee"] += rec.spread_fee
                    t["sub_fee_percent"] += rec.sub_fee_percent
                    t["sub_fee"] += rec.sub_fee
                    t["management_fee_percent"] += rec.management_fee_percent
                    t["perf1_percent"] += rec.perf1_percent
                    t["perf1_threshold"] += rec.perf1_threshold
                    t["perf2_percent"] += rec.perf2_percent
                    t["perf2_threshold"] += rec.perf2_threshold
                    t["bd_fee_percent"] += rec.bd_fee_percent
                    t["bd_fee"] += rec.bd_fee
                    t["finra_shares"] += rec.finra_shares
                    t["finra_fee"] += rec.finra_fee
                elif rec.ownership <= 0:
                    zero.append(rec)
                else:
                    active.append(rec)
                    t = totals[rec.vehicle]
                    t["count"] += 1
                    t["cost_per_share"] += rec.cost_per_share
                    t["commitment"] += rec.commitment
                    t["price_per_share"] += rec.price_per_share
                    t["shares"] += rec.shares
                    t["ownership"] += rec.ownership
                    t["spread_per_share"] += rec.spread_per_share
                    t["spread_fee"] += rec.spread_fee
                    t["sub_fee_percent"] += rec.sub_fee_percent
                    t["sub_fee"] += rec.sub_fee
                    t["management_fee_percent"] += rec.management_fee_percent
                    t["perf1_percent"] += rec.perf1_percent
                    t["perf1_threshold"] += rec.perf1_threshold
                    t["perf2_percent"] += rec.perf2_percent
                    t["perf2_threshold"] += rec.perf2_threshold
                    t["bd_fee_percent"] += rec.bd_fee_percent
                    t["bd_fee"] += rec.bd_fee
                    t["finra_shares"] += rec.finra_shares
                    t["finra_fee"] += rec.finra_fee

                # Introducer blocks in dashboard (all rows, including ownership=0).
                # Some sheets have shifted commission columns (VC203), so this is rule-driven.
                row_vals = row
                if commission_slots:
                    for _, iidx, sidx in commission_slots:
                        if iidx and iidx <= len(row_vals):
                            intro_totals_all_rows[rec.vehicle]["invested_amount"] += to_float(row_vals[iidx - 1])
                        if sidx and sidx <= len(row_vals):
                            intro_totals_all_rows[rec.vehicle]["spread"] += to_float(row_vals[sidx - 1])
                else:
                    for cidx in invested_cols:
                        if cidx <= len(row_vals):
                            intro_totals_all_rows[rec.vehicle]["invested_amount"] += to_float(row_vals[cidx - 1])
                    for cidx in spread_cols:
                        if cidx <= len(row_vals):
                            intro_totals_all_rows[rec.vehicle]["spread"] += to_float(row_vals[cidx - 1])

                for (nidx, iidx, sidx) in commission_slots:
                    raw_name = row_vals[nidx - 1] if nidx <= len(row_vals) else None
                    if raw_name not in (None, ""):
                        carried_intro_name_by_col[nidx] = str(raw_name).strip()

                    invested_amount = to_float(row_vals[iidx - 1] if (iidx and iidx <= len(row_vals)) else None)
                    spread_amount = to_float(row_vals[sidx - 1] if (sidx and sidx <= len(row_vals)) else None)
                    intro_name = carried_intro_name_by_col.get(nidx, "")
                    if not intro_name or intro_name in {"-", "—", "–"}:
                        if abs(invested_amount) > 0.01 or abs(spread_amount) > 0.01:
                            self.add_warning(
                                "dashboard_commission_amount_without_name",
                                rec.vehicle,
                                (
                                    f"sheet={sheet} row={r_idx} name_col={nidx} "
                                    f"invested={round(invested_amount,2)} spread={round(spread_amount,2)}"
                                ),
                                f"{sheet}:{r_idx}",
                            )
                        continue

                    # Combined labels in one cell (e.g. "Intro A + Intro B") are usually ambiguous.
                    # But if rules explicitly collapse that combined label into a canonical introducer
                    # (e.g. "Rick + Andrew" -> "Altras Capital Financing Broker"), keep it.
                    intro_key = name_key(intro_name, self.aliases)
                    if ("+" in intro_name) or ("\n" in intro_name):
                        combined_key = canonical_name_key(intro_name)
                        mapped_targets = {
                            self.aliases.get(k)
                            for k in alias_key_variants(intro_name)
                            if self.aliases.get(k)
                        }
                        explicit_collapse = any(t and t != combined_key for t in mapped_targets)
                        if not explicit_collapse:
                            if abs(invested_amount) > 0.01 or abs(spread_amount) > 0.01:
                                msg = (
                                    f"sheet={sheet} row={r_idx} introducer={intro_name!r} "
                                    f"invested={round(invested_amount,2)} spread={round(spread_amount,2)}"
                                )
                                if combined_intro_policy == "warn":
                                    self.add_warning("dashboard_combined_introducer_name", rec.vehicle, msg, f"{sheet}:{r_idx}")
                                elif combined_intro_policy == "ignore":
                                    pass
                                else:
                                    self.add_failure("dashboard_combined_introducer_name", rec.vehicle, msg, f"{sheet}:{r_idx}")
                            continue
                    if abs(invested_amount) > 0.01:
                        dash_commissions.append(
                            DashCommission(
                                vehicle=rec.vehicle,
                                sheet=sheet,
                                row_num=r_idx,
                                investor_name=rec.investor_name,
                                investor_key=rec.investor_key,
                                introducer_name=intro_name,
                                introducer_key=intro_key,
                                basis_type="invested_amount",
                                amount=invested_amount,
                            )
                        )
                    if abs(spread_amount) > 0.01:
                        dash_commissions.append(
                            DashCommission(
                                vehicle=rec.vehicle,
                                sheet=sheet,
                                row_num=r_idx,
                                investor_name=rec.investor_name,
                                investor_key=rec.investor_key,
                                introducer_name=intro_name,
                                introducer_key=intro_key,
                                basis_type="spread",
                                amount=spread_amount,
                            )
                        )
        return active, zero, totals, intro_totals_all_rows, dash_commissions, combined_commission_skip_keys

    def load_db_data(self) -> dict[str, Any]:
        env = load_env(ENV_PATH)
        key = env.get("SUPABASE_SERVICE_ROLE_KEY")
        if not key:
            raise RuntimeError("SUPABASE_SERVICE_ROLE_KEY missing in .env.local")

        scope_codes = sorted(self.rules["scope_vehicle_codes"])
        v_filter = "in.(" + ",".join([f'"{x}"' for x in scope_codes]) + ")"
        vehicles = api_get_all("vehicles", "id,entity_code,name", key, {"entity_code": v_filter})
        vid_to_code = {v["id"]: v["entity_code"] for v in vehicles}
        vehicle_ids = list(vid_to_code.keys())
        if not vehicle_ids:
            raise RuntimeError("No scope vehicles found in DB")

        vid_filter = "in.(" + ",".join([f'"{x}"' for x in vehicle_ids]) + ")"
        subs_raw = api_get_all(
            "subscriptions",
            "id,deal_id,investor_id,vehicle_id,commitment,cost_per_share,price_per_share,num_shares,units,contract_date,spread_per_share,spread_fee_amount,subscription_fee_percent,subscription_fee_amount,management_fee_percent,performance_fee_tier1_percent,performance_fee_tier1_threshold,performance_fee_tier2_percent,performance_fee_tier2_threshold,bd_fee_percent,bd_fee_amount,finra_shares,finra_fee_amount,status,funded_amount",
            key,
            {"vehicle_id": vid_filter},
        )
        pos_raw = api_get_all("positions", "id,investor_id,vehicle_id,units,cost_basis", key, {"vehicle_id": vid_filter})
        deals_raw = api_get_all("deals", "id,vehicle_id,name,currency", key, {"vehicle_id": vid_filter})
        deal_ids = [d["id"] for d in deals_raw]
        did_to_vehicle = {d["id"]: vid_to_code.get(d["vehicle_id"]) for d in deals_raw}
        did_to_currency = {d["id"]: normalize_currency_token(d.get("currency")) for d in deals_raw}

        intros = []
        comms = []
        if deal_ids:
            did_filter = "in.(" + ",".join([f'"{x}"' for x in deal_ids]) + ")"
            intros = api_get_all(
                "introductions",
                "id,introducer_id,prospect_investor_id,deal_id,status",
                key,
                {"deal_id": did_filter},
            )
            comms = api_get_all(
                "introducer_commissions",
                "id,introducer_id,deal_id,investor_id,introduction_id,basis_type,rate_bps,base_amount,accrual_amount,status,tier_number",
                key,
                {"deal_id": did_filter},
            )

        inv_ids = sorted(
            {
                x["investor_id"]
                for x in subs_raw
                if x.get("investor_id")
            }
            | {
                x["investor_id"]
                for x in comms
                if x.get("investor_id")
            }
            | {
                x["prospect_investor_id"]
                for x in intros
                if x.get("prospect_investor_id")
            }
        )
        investors = []
        if inv_ids:
            investors = api_get_all(
                "investors",
                "id,legal_name,display_name,first_name,last_name,status,type",
                key,
                {"id": "in.(" + ",".join([f'"{x}"' for x in inv_ids]) + ")"},
            )
        iid_to_name = {i["id"]: (i.get("legal_name") or i.get("display_name") or "") for i in investors}
        iid_to_type = {i["id"]: str(i.get("type") or "").strip().lower() for i in investors}

        introducers = api_get_all("introducers", "id,legal_name,display_name,status,email", key)
        brokers = api_get_all("brokers", "id,legal_name,display_name,status,email,contact_name", key)
        intro_id_to_name = {
            i["id"]: (i.get("legal_name") or i.get("display_name") or "")
            for i in introducers
        }

        subs: list[DbSub] = []
        sub_totals = defaultdict(lambda: defaultdict(float))
        sub_key_counts = Counter()
        sub_null_deal_ids: list[str] = []
        sub_deal_vehicle_mismatch: list[tuple[str, str, str, str]] = []
        for s in subs_raw:
            vc = vid_to_code.get(s["vehicle_id"], "")
            nm = iid_to_name.get(s["investor_id"], "")
            rec = DbSub(
                id=s["id"],
                deal_id=s.get("deal_id"),
                vehicle=vc,
                investor_id=s["investor_id"],
                investor_name=nm,
                investor_key=self.investor_key_for_vehicle(vc, nm),
                currency=did_to_currency.get(s.get("deal_id"), ""),
                cost_per_share=to_float(s.get("cost_per_share")),
                commitment=to_float(s.get("commitment")),
                price_per_share=to_float(s.get("price_per_share")),
                shares=to_float(s.get("num_shares")),
                units=to_float(s.get("units")),
                contract_date=parse_date(s.get("contract_date")),
                spread_per_share=to_float(s.get("spread_per_share")),
                spread_fee=to_float(s.get("spread_fee_amount")),
                sub_fee_percent=to_float(s.get("subscription_fee_percent")),
                sub_fee=to_float(s.get("subscription_fee_amount")),
                management_fee_percent=to_float(s.get("management_fee_percent")),
                perf1_percent=to_float(s.get("performance_fee_tier1_percent")),
                perf1_threshold=to_float(s.get("performance_fee_tier1_threshold")),
                perf2_percent=to_float(s.get("performance_fee_tier2_percent")),
                perf2_threshold=to_float(s.get("performance_fee_tier2_threshold")),
                bd_fee_percent=to_float(s.get("bd_fee_percent")),
                bd_fee=to_float(s.get("bd_fee_amount")),
                finra_shares=to_float(s.get("finra_shares")),
                finra_fee=to_float(s.get("finra_fee_amount")),
                status=str(s.get("status") or ""),
                funded_amount=to_float(s.get("funded_amount")),
            )
            subs.append(rec)
            if not rec.deal_id:
                sub_null_deal_ids.append(rec.id)
            else:
                dvc = did_to_vehicle.get(rec.deal_id, "")
                if dvc and rec.vehicle and dvc != rec.vehicle:
                    sub_deal_vehicle_mismatch.append((rec.id, rec.deal_id, rec.vehicle, dvc))
            k = (rec.vehicle, round(rec.commitment, 2), round(rec.shares, 6), rec.contract_date or "")
            sub_key_counts[k] += 1
            t = sub_totals[rec.vehicle]
            t["count"] += 1
            t["cost_per_share"] += rec.cost_per_share
            t["commitment"] += rec.commitment
            t["price_per_share"] += rec.price_per_share
            t["shares"] += rec.shares
            t["ownership"] += rec.units
            t["spread_per_share"] += rec.spread_per_share
            t["spread_fee"] += rec.spread_fee
            t["sub_fee_percent"] += rec.sub_fee_percent
            t["sub_fee"] += rec.sub_fee
            t["management_fee_percent"] += rec.management_fee_percent
            t["perf1_percent"] += rec.perf1_percent
            t["perf1_threshold"] += rec.perf1_threshold
            t["perf2_percent"] += rec.perf2_percent
            t["perf2_threshold"] += rec.perf2_threshold
            t["bd_fee_percent"] += rec.bd_fee_percent
            t["bd_fee"] += rec.bd_fee
            t["finra_shares"] += rec.finra_shares
            t["finra_fee"] += rec.finra_fee

        pos_units_by_vehicle = defaultdict(float)
        pos_units_by_vehicle_investor_key = defaultdict(float)
        pos_key_counts = Counter()
        pos_dup = Counter()
        zero_pos_rows = []
        for p in pos_raw:
            vc = vid_to_code.get(p["vehicle_id"], "")
            units = to_float(p.get("units"))
            pos_units_by_vehicle[vc] += units
            inv_id = p.get("investor_id")
            inv_name = iid_to_name.get(inv_id, "")
            if inv_name:
                inv_key = self.investor_key_for_vehicle(vc, inv_name)
                pos_units_by_vehicle_investor_key[(vc, inv_key)] += units
                pos_key_counts[(vc, inv_key)] += 1
            pos_dup[(p["investor_id"], p["vehicle_id"])] += 1
            if abs(units) < 1e-9:
                zero_pos_rows.append(p["id"])

        # For IN scope, ownership is tracked in positions table, not subscription.units.
        for vc in vid_to_code.values():
            sub_totals[vc]["ownership"] = pos_units_by_vehicle.get(vc, 0.0)

        db_commission_rows: list[DbCommission] = []
        basis_allow = set(self.rules.get("commission_match_basis_types", ["invested_amount", "spread"]))
        for c in comms:
            btype = str(c.get("basis_type") or "")
            if btype not in basis_allow:
                continue
            vc = did_to_vehicle.get(c.get("deal_id"), "")
            investor_name = iid_to_name.get(c.get("investor_id"), "")
            introducer_name = intro_id_to_name.get(c.get("introducer_id"), "")
            rate_val = c.get("rate_bps")
            rate_bps: int | None = None
            if rate_val not in (None, ""):
                try:
                    rate_bps = int(float(rate_val))
                except Exception:
                    rate_bps = None
            db_commission_rows.append(
                DbCommission(
                    id=str(c.get("id") or ""),
                    vehicle=vc,
                    investor_name=investor_name,
                    investor_key=self.investor_key_for_vehicle(vc, investor_name) if investor_name else "",
                    introducer_name=introducer_name,
                    introducer_key=name_key(introducer_name, self.aliases) if introducer_name else "",
                    basis_type=btype,
                    amount=to_float(c.get("accrual_amount")),
                    rate_bps=rate_bps,
                )
            )

        # Identity-duplicate signal inside scope-relevant data:
        # same raw normalized investor identity mapped to more than one investor_id.
        # NOTE: Use raw canonical_name_key() (not alias/rule mapping), otherwise
        # intentional transfer aliases (e.g. Zandera -> Michael in a vehicle) create false positives.
        investor_scope_refs: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
        for s in subs:
            raw_key = canonical_name_key(s.investor_name)
            if raw_key:
                investor_scope_refs[raw_key][s.investor_id] += 1
        for p in pos_raw:
            inv_id = p.get("investor_id")
            inv_name = iid_to_name.get(inv_id, "")
            ik = canonical_name_key(inv_name) if inv_name else ""
            if ik and inv_id:
                investor_scope_refs[ik][inv_id] += 1
        for i in intros:
            inv_id = i.get("prospect_investor_id")
            inv_name = iid_to_name.get(inv_id, "")
            ik = canonical_name_key(inv_name) if inv_name else ""
            if ik and inv_id:
                investor_scope_refs[ik][inv_id] += 1
        for c in comms:
            inv_id = c.get("investor_id")
            inv_name = iid_to_name.get(inv_id, "")
            ik = canonical_name_key(inv_name) if inv_name else ""
            if ik and inv_id:
                investor_scope_refs[ik][inv_id] += 1

        investor_identity_duplicates = {
            ik: refs for ik, refs in investor_scope_refs.items() if len(refs) > 1
        }

        # Same-individual loose key (first+last) on >1 investor IDs in scope.
        individual_scope_refs: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
        for s in subs:
            if iid_to_type.get(s.investor_id) != "individual":
                continue
            ik = individual_loose_identity_key(s.investor_name)
            if ik:
                individual_scope_refs[ik][s.investor_id] += 1
        for p in pos_raw:
            inv_id = p.get("investor_id")
            if iid_to_type.get(inv_id) != "individual":
                continue
            inv_name = iid_to_name.get(inv_id, "")
            ik = individual_loose_identity_key(inv_name)
            if ik and inv_id:
                individual_scope_refs[ik][inv_id] += 1
        for i in intros:
            inv_id = i.get("prospect_investor_id")
            if iid_to_type.get(inv_id) != "individual":
                continue
            inv_name = iid_to_name.get(inv_id, "")
            ik = individual_loose_identity_key(inv_name)
            if ik and inv_id:
                individual_scope_refs[ik][inv_id] += 1
        for c in comms:
            inv_id = c.get("investor_id")
            if iid_to_type.get(inv_id) != "individual":
                continue
            inv_name = iid_to_name.get(inv_id, "")
            ik = individual_loose_identity_key(inv_name)
            if ik and inv_id:
                individual_scope_refs[ik][inv_id] += 1

        investor_identity_duplicates_individual_loose = {
            ik: refs for ik, refs in individual_scope_refs.items() if len(refs) > 1
        }

        return {
            "vehicles": vehicles,
            "vid_to_code": vid_to_code,
            "did_to_vehicle": did_to_vehicle,
            "subs": subs,
            "sub_totals": sub_totals,
            "sub_key_counts": sub_key_counts,
            "sub_null_deal_ids": sub_null_deal_ids,
            "sub_deal_vehicle_mismatch": sub_deal_vehicle_mismatch,
            "positions_raw": pos_raw,
            "pos_units_by_vehicle": pos_units_by_vehicle,
            "pos_units_by_vehicle_investor_key": pos_units_by_vehicle_investor_key,
            "pos_key_counts": pos_key_counts,
            "pos_dup": pos_dup,
            "zero_pos_rows": zero_pos_rows,
            "introductions": intros,
            "commissions": comms,
            "db_commission_rows": db_commission_rows,
            "introducers": introducers,
            "brokers": brokers,
            "iid_to_name": iid_to_name,
            "iid_to_type": iid_to_type,
            "investor_identity_duplicates": investor_identity_duplicates,
            "investor_identity_duplicates_individual_loose": investor_identity_duplicates_individual_loose,
        }

    def run_checks(self):
        dash_active, dash_zero, dash_totals, dash_intro_totals, dash_commissions, combined_commission_skip_keys = self.load_dashboard_rows()
        db = self.load_db_data()
        tolerance_overrides = self.rules.get("metric_tolerance_overrides", {}) or {}
        percent_compare_rules = self.rules.get("percent_compare_rules", {}) or {}
        ignore_contract_date_if_missing = bool(self.rules.get("ignore_contract_date_if_missing", True))

        configured_scope_codes = list(self.rules["scope_vehicle_codes"])
        if bool(self.rules.get("restrict_to_db_vehicles_with_data", False)):
            db_vehicles_with_data = set()
            db_vehicles_with_data.update(s.vehicle for s in db["subs"] if s.vehicle)
            for p in db["positions_raw"]:
                vc = db["vid_to_code"].get(p.get("vehicle_id"))
                if vc:
                    db_vehicles_with_data.add(vc)
            for i in db["introductions"]:
                vc = db["did_to_vehicle"].get(i.get("deal_id"))
                if vc:
                    db_vehicles_with_data.add(vc)
            for c in db["commissions"]:
                vc = db["did_to_vehicle"].get(c.get("deal_id"))
                if vc:
                    db_vehicles_with_data.add(vc)
            scope_codes = [vc for vc in configured_scope_codes if vc in db_vehicles_with_data]
            scope_set = set(scope_codes)
            dash_active = [r for r in dash_active if r.vehicle in scope_set]
            dash_zero = [r for r in dash_zero if r.vehicle in scope_set]
            dash_commissions = [r for r in dash_commissions if r.vehicle in scope_set]
        else:
            scope_codes = configured_scope_codes
            scope_set = set(scope_codes)

        db_totals = db["sub_totals"]

        # vehicle totals parity
        metrics = [
            "count",
            "cost_per_share",
            "commitment",
            "price_per_share",
            "shares",
            "ownership",
            "spread_per_share",
            "spread_fee",
            "sub_fee_percent",
            "sub_fee",
            "management_fee_percent",
            "perf1_percent",
            "perf1_threshold",
            "perf2_percent",
            "perf2_threshold",
            "bd_fee_percent",
            "bd_fee",
            "finra_shares",
            "finra_fee",
        ]
        skip_vehicle_totals_metrics = set(self.rules.get("vehicle_totals_skip_metrics", []))
        for vc in scope_codes:
            for m in metrics:
                if m in skip_vehicle_totals_metrics:
                    continue
                dv = dash_totals[vc].get(m, 0.0)
                bv = db_totals[vc].get(m, 0.0)
                tol = 0 if m == "count" else 0.01
                if m == "count":
                    mismatch = abs(bv - dv) > tol
                else:
                    mismatch = not values_match(
                        m,
                        dv,
                        bv,
                        tol,
                        tolerance_overrides,
                        percent_compare_rules,
                    )
                if mismatch:
                    self.add_failure("vehicle_totals_mismatch", vc, f"{m}: dashboard={dv} db={bv} delta={bv-dv}")

        # Row-level parity by investor identity (exact first, then deterministic fallbacks)
        def relaxed_key_dash(r: DashRow) -> tuple[str, str, float, float]:
            return (r.vehicle, r.investor_key, round(r.commitment, 2), round(r.shares, 6))

        def relaxed_key_db(r: DbSub) -> tuple[str, str, float, float]:
            return (r.vehicle, r.investor_key, round(r.commitment, 2), round(r.shares, 6))

        def row_sort_dash(r: DashRow):
            return (
                r.contract_date or "",
                round(r.sub_fee, 6),
                round(r.price_per_share, 6),
                round(r.spread_fee, 6),
                round(r.cost_per_share, 6),
                r.row_num,
            )

        def row_sort_db(r: DbSub):
            return (
                r.contract_date or "",
                round(r.sub_fee, 6),
                round(r.price_per_share, 6),
                round(r.spread_fee, 6),
                round(r.cost_per_share, 6),
                r.id,
            )

        dash_groups: dict[tuple[str, str, float, float], list[DashRow]] = defaultdict(list)
        db_groups: dict[tuple[str, str, float, float], list[DbSub]] = defaultdict(list)
        for r in dash_active:
            dash_groups[relaxed_key_dash(r)].append(r)
        for s in db["subs"]:
            db_groups[relaxed_key_db(s)].append(s)

        matched_pairs: list[tuple[DashRow, DbSub, str]] = []
        unruled_fallback_pair_ids: set[tuple[str, int, str]] = set()
        initial_unmatched_dash: list[DashRow] = []
        initial_unmatched_db: list[DbSub] = []
        strict_fallback_fail = bool(self.rules.get("checks", {}).get("fallback_match_is_failure", False))
        warn_safe_fallback = bool(self.rules.get("checks", {}).get("warn_safe_fallback_matches", True))
        warn_ruled_fallback = bool(self.rules.get("checks", {}).get("warn_ruled_fallback_matches", True))

        for k in sorted(set(dash_groups) | set(db_groups)):
            drows = sorted(dash_groups.get(k, []), key=row_sort_dash)
            brows = sorted(db_groups.get(k, []), key=row_sort_db)
            db_by_date: dict[str, list[DbSub]] = defaultdict(list)
            for b in brows:
                db_by_date[b.contract_date or ""].append(b)

            d_remaining: list[DashRow] = []
            for d in drows:
                date_key = d.contract_date or ""
                bucket = db_by_date.get(date_key, [])
                if bucket:
                    matched_pairs.append((d, bucket.pop(0), "exact_date"))
                else:
                    d_remaining.append(d)

            b_remaining = sorted([x for bucket in db_by_date.values() for x in bucket], key=row_sort_db)
            n = min(len(d_remaining), len(b_remaining))
            for i in range(n):
                matched_pairs.append((d_remaining[i], b_remaining[i], "relaxed"))
            initial_unmatched_dash.extend(d_remaining[n:])
            initial_unmatched_db.extend(b_remaining[n:])

        unmatched_dash = initial_unmatched_dash
        unmatched_db = initial_unmatched_db

        def unique_fallback_match(
            dash_rows: list[DashRow],
            db_rows: list[DbSub],
            dash_key_fn,
            db_key_fn,
            mode: str,
        ) -> tuple[list[DashRow], list[DbSub], list[tuple[DashRow, DbSub]]]:
            d_map: dict[tuple[Any, ...], list[int]] = defaultdict(list)
            b_map: dict[tuple[Any, ...], list[int]] = defaultdict(list)
            for i, d in enumerate(dash_rows):
                d_map[dash_key_fn(d)].append(i)
            for i, b in enumerate(db_rows):
                b_map[db_key_fn(b)].append(i)

            matched_d_idx: set[int] = set()
            matched_b_idx: set[int] = set()
            new_pairs: list[tuple[int, int]] = []

            for k in sorted(set(d_map) & set(b_map)):
                di = d_map[k]
                bi = b_map[k]
                # Deterministic pairing when multiplicity is equal:
                # supports split/renamed identity cases where strict-name match fails.
                if len(di) == len(bi):
                    for dx, bx in zip(di, bi):
                        new_pairs.append((dx, bx))
                    continue
                if len(di) == 1 and len(bi) == 1:
                    new_pairs.append((di[0], bi[0]))

            for di, bi in new_pairs:
                matched_d_idx.add(di)
                matched_b_idx.add(bi)
                matched_pairs.append((dash_rows[di], db_rows[bi], mode))

            dash_remaining = [r for i, r in enumerate(dash_rows) if i not in matched_d_idx]
            db_remaining = [r for i, r in enumerate(db_rows) if i not in matched_b_idx]
            pair_rows = [(dash_rows[di], db_rows[bi]) for di, bi in new_pairs]
            return dash_remaining, db_remaining, pair_rows

        def record_fallback_matches(mode: str, pair_rows: list[tuple[DashRow, DbSub]]) -> tuple[int, int]:
            if not pair_rows:
                return (0, 0)
            safe_fallback_modes = {"fallback_loose_name_amounts", "fallback_investor_shares_date"}
            ruled_count = 0
            unruled_count = 0
            for d, b in pair_rows:
                detail = (
                    f"mode={mode} dash_investor={d.investor_name} db_investor={b.investor_name} "
                    f"commitment={round(d.commitment,2)} shares={round(d.shares,6)} "
                    f"dash_date={d.contract_date} db_date={b.contract_date} sub_id={b.id}"
                )
                dkey = canonical_name_key(d.investor_name)
                bkey = canonical_name_key(b.investor_name)
                ruled = (
                    (d.vehicle, dkey, bkey) in self.fallback_ruled_pairs
                    or ("*", dkey, bkey) in self.fallback_ruled_pairs
                )
                if ruled:
                    ruled_count += 1
                    if warn_ruled_fallback:
                        self.add_warning("row_fallback_match_ruled", d.vehicle, detail, f"{d.sheet}:{d.row_num}")
                elif (mode in safe_fallback_modes) and (not strict_fallback_fail):
                    # Safe fallback: same investor identity after loose normalization
                    # or same investor+shares+date with only amount rounding drift.
                    unruled_count += 1
                    if warn_safe_fallback:
                        self.add_warning("row_fallback_match", d.vehicle, detail, f"{d.sheet}:{d.row_num}")
                else:
                    # Unruled fallback means mapping is unresolved; do not trust
                    # numeric comparisons for this pair until a mapping rule exists.
                    unruled_count += 1
                    self.add_failure("row_mapping_unresolved", d.vehicle, detail, f"{d.sheet}:{d.row_num}")
                    unruled_fallback_pair_ids.add((d.sheet, d.row_num, b.id))
            return (ruled_count, unruled_count)

        # Fallback 1: same vehicle + amounts + loose name key (middle-name / initials drift).
        unmatched_dash, unmatched_db, p1 = unique_fallback_match(
            unmatched_dash,
            unmatched_db,
            lambda r: (r.vehicle, loose_name_key(r.investor_name), round(r.commitment, 2), round(r.shares, 6)),
            lambda r: (r.vehicle, loose_name_key(r.investor_name), round(r.commitment, 2), round(r.shares, 6)),
            "fallback_loose_name_amounts",
        )
        _, unruled_p1 = record_fallback_matches("fallback_loose_name_amounts", p1)
        if p1 and warn_safe_fallback and unruled_p1 > 0:
            self.add_warning("row_fallback_match_loose_name_amounts", "", f"matched_pairs={len(p1)}")

        # Fallback 2: same vehicle + amounts + date (name-independent).
        unmatched_dash, unmatched_db, p1 = unique_fallback_match(
            unmatched_dash,
            unmatched_db,
            lambda r: (r.vehicle, round(r.commitment, 2), round(r.shares, 6), r.contract_date or ""),
            lambda r: (r.vehicle, round(r.commitment, 2), round(r.shares, 6), r.contract_date or ""),
            "fallback_amounts_date",
        )
        _, unruled_p1b = record_fallback_matches("fallback_amounts_date", p1)
        if p1 and warn_safe_fallback and unruled_p1b > 0:
            self.add_warning("row_fallback_match_amounts_date", "", f"matched_pairs={len(p1)}")

        # Fallback 3: same vehicle + investor + shares + date (for rounded commitment drifts).
        unmatched_dash, unmatched_db, p2 = unique_fallback_match(
            unmatched_dash,
            unmatched_db,
            lambda r: (r.vehicle, r.investor_key, round(r.shares, 6), r.contract_date or ""),
            lambda r: (r.vehicle, r.investor_key, round(r.shares, 6), r.contract_date or ""),
            "fallback_investor_shares_date",
        )
        _, unruled_p2 = record_fallback_matches("fallback_investor_shares_date", p2)
        if p2 and warn_safe_fallback and unruled_p2 > 0:
            self.add_warning("row_fallback_match_investor_shares_date", "", f"matched_pairs={len(p2)}")

        # Fallback 4: same vehicle + commitment + shares (date-independent for null-date DB rows).
        unmatched_dash, unmatched_db, p3 = unique_fallback_match(
            unmatched_dash,
            unmatched_db,
            lambda r: (r.vehicle, round(r.commitment, 2), round(r.shares, 6)),
            lambda r: (r.vehicle, round(r.commitment, 2), round(r.shares, 6)),
            "fallback_amounts_no_date",
        )
        _, unruled_p3 = record_fallback_matches("fallback_amounts_no_date", p3)
        if p3 and warn_safe_fallback and unruled_p3 > 0:
            self.add_warning("row_fallback_match_amounts_no_date", "", f"matched_pairs={len(p3)}")

        for d in unmatched_dash:
            self.add_failure(
                "row_unmatched_dashboard",
                d.vehicle,
                f"investor={d.investor_name} commitment={d.commitment} shares={d.shares} date={d.contract_date}",
                f"{d.sheet}:{d.row_num}",
            )
        for b in unmatched_db:
            self.add_failure(
                "row_unmatched_db",
                b.vehicle,
                f"investor={b.investor_name} commitment={b.commitment} shares={b.shares} date={b.contract_date} sub_id={b.id}",
            )

        metric_pairs = (
            ("cost_per_share", "cost_per_share"),
            ("price_per_share", "price_per_share"),
            ("ownership", "units"),
            ("spread_per_share", "spread_per_share"),
            ("spread_fee", "spread_fee"),
            ("sub_fee_percent", "sub_fee_percent"),
            ("sub_fee", "sub_fee"),
            ("management_fee_percent", "management_fee_percent"),
            ("perf1_percent", "perf1_percent"),
            ("perf1_threshold", "perf1_threshold"),
            ("perf2_percent", "perf2_percent"),
            ("perf2_threshold", "perf2_threshold"),
            ("bd_fee_percent", "bd_fee_percent"),
            ("bd_fee", "bd_fee"),
            ("finra_shares", "finra_shares"),
            ("finra_fee", "finra_fee"),
        )
        skip_row_metrics = set(self.rules.get("row_compare_skip_metrics", []))
        blank_skip_metrics = set(self.rules.get("row_blank_value_skip_metrics", []))
        for d, b, mode in matched_pairs:
            if mode.startswith("fallback_") and (d.sheet, d.row_num, b.id) in unruled_fallback_pair_ids:
                continue
            if (d.contract_date or "") != (b.contract_date or ""):
                if ignore_contract_date_if_missing and (not d.contract_date or not b.contract_date):
                    pass
                else:
                    self.add_failure(
                        "row_contract_date_mismatch",
                        d.vehicle,
                        f"dash_investor={d.investor_name} db_investor={b.investor_name} dashboard={d.contract_date} db={b.contract_date} sub_id={b.id} mode={mode}",
                        f"{d.sheet}:{d.row_num}",
                    )
            for dm, bm in metric_pairs:
                if dm in skip_row_metrics:
                    continue
                if dm in blank_skip_metrics and is_blank_cell_value(d.raw_numeric.get(dm)):
                    continue
                dv = float(getattr(d, dm))
                bv = float(getattr(b, bm))
                if not values_match(dm, dv, bv, 0.01, tolerance_overrides, percent_compare_rules):
                    self.add_failure(
                        "row_numeric_mismatch",
                        d.vehicle,
                        f"dash_investor={d.investor_name} db_investor={b.investor_name} metric={dm} dashboard={round(dv,6)} db={round(bv,6)} delta={round(bv-dv,6)} sub_id={b.id} mode={mode}",
                        f"{d.sheet}:{d.row_num}",
                    )
            if bool(self.rules["checks"].get("currency_must_match_dashboard_when_present", False)):
                if d.currency and d.currency != b.currency:
                    self.add_failure(
                        "subscription_currency_mismatch",
                        d.vehicle,
                        (
                            f"dash_investor={d.investor_name} db_investor={b.investor_name} "
                            f"dashboard_currency={d.currency} db_currency={b.currency or '<blank>'} sub_id={b.id} mode={mode}"
                        ),
                        f"{d.sheet}:{d.row_num}",
                    )

        # Use row-level matched-pair mapping for downstream checks (name-variant tolerant)
        # without collapsing multiple DB investors into one dashboard key.
        row_investor_map: dict[tuple[str, str, int], str] = {}
        for d, b, _ in matched_pairs:
            row_investor_map[(d.vehicle, d.sheet, d.row_num)] = b.investor_key

        # Ownership parity by investor+vehicle uses positions.
        dash_ownership_by_investor = defaultdict(float)
        for d in dash_active:
            mapped_ik = row_investor_map.get((d.vehicle, d.sheet, d.row_num), d.investor_key)
            dash_ownership_by_investor[(d.vehicle, mapped_ik)] += d.ownership
        db_ownership_by_investor = db["pos_units_by_vehicle_investor_key"]
        all_own_keys = set(dash_ownership_by_investor) | set(db_ownership_by_investor)
        for vc, ik in sorted(all_own_keys):
            dv = dash_ownership_by_investor.get((vc, ik), 0.0)
            bv = db_ownership_by_investor.get((vc, ik), 0.0)
            if abs(bv - dv) > 0.01:
                self.add_failure(
                    "position_vs_dashboard_ownership_by_investor",
                    vc,
                    f"investor_key={ik} dashboard={dv} db={bv} delta={round(bv-dv,6)}",
                )

        if bool(self.rules["checks"].get("position_mapping_coverage", True)):
            pos_key_counts = db.get("pos_key_counts", Counter())
            for (vc, ik), dv in sorted(dash_ownership_by_investor.items()):
                if dv > 0.01 and pos_key_counts.get((vc, ik), 0) == 0:
                    self.add_failure(
                        "position_row_missing_in_db",
                        vc,
                        f"investor_key={ik} dashboard_ownership={dv} db_position_rows=0",
                    )
            for (vc, ik), cnt in sorted(pos_key_counts.items()):
                if cnt <= 0:
                    continue
                if (vc, ik) not in dash_ownership_by_investor and abs(db_ownership_by_investor.get((vc, ik), 0.0)) > 0.01:
                    self.add_failure(
                        "position_row_missing_in_dashboard",
                        vc,
                        f"investor_key={ik} db_position_rows={cnt} db_units={db_ownership_by_investor.get((vc, ik), 0.0)} dashboard_ownership=0",
                    )

        # zero ownership exclusion
        if self.rules["checks"].get("zero_ownership_must_not_be_loaded", False):
            zero_keys = Counter(
                (
                    r.vehicle,
                    # For zero-ownership checks, do not apply row fallback/transfer mapping:
                    # we want the original dashboard investor identity, otherwise valid
                    # transfer aliases (e.g. historical holder -> current holder) produce false positives.
                    name_key(r.investor_name, self.aliases),
                    round(r.commitment, 2),
                    round(r.shares, 6),
                    r.contract_date or "",
                )
                for r in dash_zero
            )
            db_exact_keys = Counter(
                (s.vehicle, s.investor_key, round(s.commitment, 2), round(s.shares, 6), s.contract_date or "")
                for s in db["subs"]
            )
            db_relaxed_keys = Counter(
                (s.vehicle, s.investor_key, round(s.commitment, 2), round(s.shares, 6))
                for s in db["subs"]
            )
            for k, cnt in zero_keys.items():
                exact_cnt = db_exact_keys.get(k, 0)
                relaxed_cnt = db_relaxed_keys.get((k[0], k[1], k[2], k[3]), 0)
                if exact_cnt > 0 or relaxed_cnt > 0:
                    self.add_failure(
                        "zero_ownership_loaded",
                        k[0],
                        f"zero_key={k} dashboard_zero={cnt} db_exact={exact_cnt} db_relaxed={relaxed_cnt}",
                    )

        # status/funded
        if self.rules["checks"].get("status_must_be_funded", False):
            for s in db["subs"]:
                if s.status.lower() != "funded":
                    self.add_failure("status_not_funded", s.vehicle, f"sub={s.id} status={s.status}")
        if self.rules["checks"].get("funded_amount_equals_commitment", False):
            for s in db["subs"]:
                if abs(s.funded_amount - s.commitment) > 0.01:
                    self.add_failure(
                        "funded_amount_mismatch",
                        s.vehicle,
                        f"sub={s.id} commitment={s.commitment} funded={s.funded_amount}",
                    )
        if self.rules["checks"].get("commission_status_must_be_paid", False):
            allowed = {str(x).strip().lower() for x in self.rules.get("allowed_commission_statuses", ["paid"])}
            for c in db["commissions"]:
                st = str(c.get("status") or "").strip().lower()
                if st not in allowed:
                    vc = db["did_to_vehicle"].get(c.get("deal_id"), "")
                    self.add_failure(
                        "commission_status_invalid",
                        vc,
                        f"commission_id={c.get('id')} status={st or '<blank>'} allowed={sorted(allowed)}",
                    )

        # structural checks
        if bool(self.rules["checks"].get("subscription_deal_id_required", True)):
            for sid in db.get("sub_null_deal_ids", []):
                self.add_failure("subscription_missing_deal_id", "", f"subscription_id={sid}")

        if bool(self.rules["checks"].get("subscription_deal_vehicle_consistency", True)):
            for sid, did, svc, dvc in db.get("sub_deal_vehicle_mismatch", []):
                self.add_failure(
                    "subscription_deal_vehicle_mismatch",
                    svc,
                    f"subscription_id={sid} deal_id={did} subscription_vehicle={svc} deal_vehicle={dvc}",
                )

        if bool(self.rules["checks"].get("investor_identity_duplicate_check", True)):
            for ik, refs in sorted(db.get("investor_identity_duplicates", {}).items(), key=lambda x: x[0]):
                detail = ", ".join([f"{iid}:{cnt}" for iid, cnt in sorted(refs.items())])
                self.add_failure("investor_identity_duplicate", "", f"investor_key={ik} ids={detail}")

        if bool(self.rules["checks"].get("investor_identity_duplicate_individual_loose_check", True)):
            for ik, refs in sorted(db.get("investor_identity_duplicates_individual_loose", {}).items(), key=lambda x: x[0]):
                detail = ", ".join([f"{iid}:{cnt}" for iid, cnt in sorted(refs.items())])
                self.add_failure("investor_identity_duplicate_individual_loose", "", f"identity_key={ik} ids={detail}")

        # positions
        if self.rules["checks"].get("positions_units_no_zero_rows", False):
            for pid in db["zero_pos_rows"]:
                self.add_failure("position_zero_units", "", f"position_id={pid}")
        if self.rules["checks"].get("position_unique_per_investor_vehicle", False):
            for (inv_id, veh_id), c in db["pos_dup"].items():
                if c > 1:
                    vc = db["vid_to_code"].get(veh_id, "")
                    self.add_failure("position_duplicate", vc, f"investor_id={inv_id} vehicle_id={veh_id} count={c}")
        for vc in scope_codes:
            db_units = db["pos_units_by_vehicle"].get(vc, 0.0)
            dash_units = dash_totals[vc].get("ownership", 0.0)
            if abs(db_units - dash_units) > 0.01:
                self.add_failure("position_vs_dashboard_ownership", vc, f"positions={db_units} dashboard={dash_units}")

        # commission duplicate exact
        if self.rules["checks"].get("commission_exact_duplicate_check", False):
            dup = Counter(
                (
                    c.get("introducer_id"),
                    c.get("deal_id"),
                    c.get("investor_id"),
                    c.get("introduction_id"),
                    c.get("basis_type"),
                    c.get("tier_number"),
                    c.get("rate_bps"),
                    str(c.get("base_amount")),
                    str(c.get("accrual_amount")),
                )
                for c in db["commissions"]
            )
            for k, c in dup.items():
                if c > 1:
                    did = k[1]
                    vc = db["did_to_vehicle"].get(did, "")
                    if vc not in scope_set:
                        continue
                    self.add_failure("commission_duplicate_exact", vc, f"count={c} key={k}")

        # Dashboard introducer block totals vs DB commission totals (all dashboard rows)
        db_comm_vehicle_basis = defaultdict(lambda: defaultdict(float))
        for c in db["commissions"]:
            vc = db["did_to_vehicle"].get(c.get("deal_id"), "")
            if vc not in scope_set:
                continue
            bt = str(c.get("basis_type") or "")
            db_comm_vehicle_basis[vc][bt] += to_float(c.get("accrual_amount"))
        comm_skip_vcs = set(self.rules.get("commission_totals_exclude_vehicle_codes", []))
        comm_ruled_diffs = self.rules.get("commission_total_ruled_diffs_invested_amount", {})
        for vc in scope_codes:
            if vc in comm_skip_vcs:
                continue
            dash_inv = dash_intro_totals[vc].get("invested_amount", 0.0)
            db_inv = db_comm_vehicle_basis[vc].get("invested_amount", 0.0)
            inv_delta = db_inv - dash_inv
            expected_delta = comm_ruled_diffs.get(vc)
            if abs(inv_delta) > 0.01:
                if expected_delta is not None and abs(inv_delta - float(expected_delta)) <= 0.01:
                    self.add_warning(
                        "commission_totals_ruled_diff_invested_amount",
                        vc,
                        f"dashboard={round(dash_inv,6)} db={round(db_inv,6)} delta={round(inv_delta,6)} expected={round(float(expected_delta),6)}",
                    )
                else:
                    self.add_failure(
                        "commission_totals_mismatch_invested_amount",
                        vc,
                        f"dashboard={round(dash_inv,6)} db={round(db_inv,6)} delta={round(inv_delta,6)}",
                    )
            dash_spread = dash_intro_totals[vc].get("spread", 0.0)
            db_spread = db_comm_vehicle_basis[vc].get("spread", 0.0)
            spread_delta = db_spread - dash_spread
            spread_ruled_diffs = self.rules.get("commission_total_ruled_diffs_spread", {})
            expected_spread_delta = spread_ruled_diffs.get(vc)
            if abs(spread_delta) > 0.01:
                if expected_spread_delta is not None and abs(spread_delta - float(expected_spread_delta)) <= 0.01:
                    self.add_warning(
                        "commission_totals_ruled_diff_spread",
                        vc,
                        f"dashboard={round(dash_spread,6)} db={round(db_spread,6)} delta={round(spread_delta,6)} expected={round(float(expected_spread_delta),6)}",
                    )
                else:
                    self.add_failure(
                        "commission_totals_mismatch_spread",
                        vc,
                        f"dashboard={round(dash_spread,6)} db={round(db_spread,6)} delta={round(spread_delta,6)}",
                    )

        # Row-level commission parity: vehicle + investor + introducer + basis
        basis_allow = set(self.rules.get("commission_match_basis_types", ["invested_amount", "spread"]))
        forbidden_global_keys = {name_key(x, self.aliases) for x in self.rules.get("broker_like_introducers_forbidden_global", [])}
        forbidden_by_vehicle_keys = {
            vc: {name_key(x, self.aliases) for x in names}
            for vc, names in self.rules.get("broker_like_introducers_forbidden_by_vehicle", {}).items()
        }
        warnings_only_keys = {name_key(x, self.aliases) for x in self.rules.get("warnings_only_introducers", [])}
        comm_skip_vcs = set(self.rules.get("commission_totals_exclude_vehicle_codes", []))
        mapped_combined_commission_skip_keys: set[tuple[str, str, str]] = set()
        for vc, sheet, row_num, basis in combined_commission_skip_keys:
            mapped_inv = row_investor_map.get((vc, sheet, row_num), "")
            if mapped_inv:
                mapped_combined_commission_skip_keys.add((vc, mapped_inv, basis))

        dash_comm_agg = defaultdict(float)
        dash_comm_rows_by_key: dict[tuple[str, str, str, str], list[float]] = defaultdict(list)
        dash_meta: dict[tuple[str, str, str, str], dict[str, str]] = {}
        for d in dash_commissions:
            if d.basis_type not in basis_allow:
                continue
            if d.vehicle in comm_skip_vcs:
                continue
            mapped_inv = row_investor_map.get((d.vehicle, d.sheet, d.row_num), "")
            if not mapped_inv:
                # Commission-only rows may not be present in subscription row matching.
                mapped_inv = self.apply_ruled_fallback_investor(d.vehicle, d.investor_key)
            k = (d.vehicle, mapped_inv, d.introducer_key, d.basis_type)
            dash_comm_agg[k] += d.amount
            dash_comm_rows_by_key[k].append(d.amount)
            if k not in dash_meta:
                dash_meta[k] = {"investor": d.investor_name, "introducer": d.introducer_name, "row_ref": f"{d.sheet}:{d.row_num}"}

        db_comm_agg = defaultdict(float)
        db_comm_rows_by_key: dict[tuple[str, str, str, str], list[float]] = defaultdict(list)
        db_meta: dict[tuple[str, str, str, str], dict[str, str]] = {}
        for c in db["db_commission_rows"]:
            if c.basis_type not in basis_allow:
                continue
            if c.vehicle in comm_skip_vcs:
                continue
            k = (c.vehicle, c.investor_key, c.introducer_key, c.basis_type)
            db_comm_agg[k] += c.amount
            # Count/split parity should compare only economically meaningful rows.
            # Historical zero-amount rows are common in DB and should not create
            # synthetic row-count mismatches against dashboard rows.
            if abs(c.amount) > 0.01:
                db_comm_rows_by_key[k].append(c.amount)
            if k not in db_meta:
                db_meta[k] = {"investor": c.investor_name, "introducer": c.introducer_name, "row_ref": c.id}

        for k in sorted(set(dash_comm_agg) | set(db_comm_agg)):
            vc, inv_key, intro_key, basis = k
            if (vc, inv_key, basis) in mapped_combined_commission_skip_keys:
                continue
            dv = dash_comm_agg.get(k, 0.0)
            bv = db_comm_agg.get(k, 0.0)
            meta = dash_meta.get(k) or db_meta.get(k) or {}
            investor = meta.get("investor", inv_key)
            introducer = meta.get("introducer", intro_key)
            row_ref = meta.get("row_ref", "")
            is_forbidden = intro_key in forbidden_global_keys or intro_key in forbidden_by_vehicle_keys.get(vc, set())
            is_warning_only = intro_key in warnings_only_keys

            if abs(dv) > 0.01 and abs(bv) <= 0.01:
                if is_forbidden:
                    self.add_warning(
                        "commission_row_ruled_removed",
                        vc,
                        f"investor={investor} introducer={introducer} basis={basis} dashboard={round(dv,2)} db=0.0",
                        row_ref,
                    )
                else:
                    self.add_failure(
                        "commission_row_missing_in_db",
                        vc,
                        f"investor={investor} introducer={introducer} basis={basis} dashboard={round(dv,2)} db=0.0",
                        row_ref,
                    )
                continue

            if abs(bv) > 0.01 and abs(dv) <= 0.01:
                if is_warning_only:
                    self.add_warning(
                        "commission_row_warning_only_present",
                        vc,
                        f"investor={investor} introducer={introducer} basis={basis} dashboard=0.0 db={round(bv,2)}",
                        row_ref,
                    )
                else:
                    self.add_failure(
                        "commission_row_missing_in_dashboard",
                        vc,
                        f"investor={investor} introducer={introducer} basis={basis} dashboard=0.0 db={round(bv,2)}",
                        row_ref,
                    )
                continue

            if abs(bv - dv) > 0.01:
                self.add_failure(
                    "commission_row_amount_mismatch",
                    vc,
                    f"investor={investor} introducer={introducer} basis={basis} dashboard={round(dv,2)} db={round(bv,2)} delta={round(bv-dv,2)}",
                    row_ref,
                )

            # Multiplicity/split parity: same key should have the same row count and split profile.
            # Aggregate totals can match while per-row split is wrong; this catches that.
            d_rows = sorted([round(x, 6) for x in dash_comm_rows_by_key.get(k, [])])
            b_rows = sorted([round(x, 6) for x in db_comm_rows_by_key.get(k, [])])
            if abs(dv) > 0.01 and abs(bv) > 0.01 and d_rows and b_rows:
                if len(d_rows) != len(b_rows):
                    self.add_failure(
                        "commission_row_count_mismatch",
                        vc,
                        f"investor={investor} introducer={introducer} basis={basis} dashboard_rows={len(d_rows)} db_rows={len(b_rows)}",
                        row_ref,
                    )
                elif len(d_rows) > 1:
                    if any(abs(b_rows[i] - d_rows[i]) > 0.01 for i in range(len(d_rows))):
                        self.add_failure(
                            "commission_row_split_mismatch",
                            vc,
                            f"investor={investor} introducer={introducer} basis={basis} dashboard_split={d_rows} db_split={b_rows}",
                            row_ref,
                        )

        # commission -> introduction FK
        if self.rules["checks"].get("commission_fk_introduction_check", False):
            intro_ids = {x["id"] for x in db["introductions"]}
            for c in db["commissions"]:
                iid = c.get("introduction_id")
                if iid and iid not in intro_ids:
                    vc = db["did_to_vehicle"].get(c.get("deal_id"), "")
                    self.add_failure("commission_broken_introduction_fk", vc, f"commission_id={c.get('id')} intro_id={iid}")

        max_intro = int(self.rules["checks"].get("max_introducers_per_subscription", 0) or 0)
        if max_intro > 0:
            intro_set = defaultdict(set)
            for i in db["introductions"]:
                key_i = (i.get("deal_id"), i.get("prospect_investor_id"))
                intro_set[key_i].add(i.get("introducer_id"))
            for (deal_id, investor_id), ids in intro_set.items():
                if len(ids) > max_intro:
                    vc = db["did_to_vehicle"].get(deal_id, "")
                    inv_name = db["iid_to_name"].get(investor_id, investor_id)
                    self.add_failure(
                        "max_introducers_per_subscription_exceeded",
                        vc,
                        f"investor={inv_name} deal_id={deal_id} introducers={len(ids)} max={max_intro}",
                    )

        # broker rules
        if bool(self.rules.get("run_broker_checks", True)):
            broker_expected = set(self.rules.get("brokers_table_expected", []))
            broker_actual = {str(b.get("legal_name") or "").strip() for b in db["brokers"]}
            missing_brokers = sorted(broker_expected - broker_actual)
            extra_brokers = sorted(broker_actual - broker_expected)
            if missing_brokers:
                self.add_failure("brokers_missing", "", f"{missing_brokers}")
            if extra_brokers:
                self.add_warning("brokers_unexpected", "", f"{extra_brokers}")

        intro_by_id = {i["id"]: str(i.get("legal_name") or i.get("display_name") or "").strip() for i in db["introducers"]}
        forbidden_global = set(self.rules.get("broker_like_introducers_forbidden_global", []))
        forbidden_by_vehicle = self.rules.get("broker_like_introducers_forbidden_by_vehicle", {})
        warnings_only = set(self.rules.get("warnings_only_introducers", []))
        forbidden_master = set(self.rules.get("forbidden_in_introducers_master", []))
        forbidden_in_introductions = set(self.rules.get("forbidden_in_introductions_global", []))

        for i in db["introducers"]:
            iname = str(i.get("legal_name") or i.get("display_name") or "").strip()
            if iname in forbidden_master:
                self.add_failure("forbidden_name_in_introducers_master", "", iname)

        for i in db["introductions"]:
            iname = intro_by_id.get(i.get("introducer_id"), "")
            if iname in forbidden_in_introductions:
                vc = db["did_to_vehicle"].get(i.get("deal_id"), "")
                inv = db["iid_to_name"].get(i.get("prospect_investor_id"), i.get("prospect_investor_id"))
                self.add_failure("forbidden_name_in_introductions", vc, f"introducer={iname} investor={inv}")

        # comm aggregates by introducer + vehicle
        agg = defaultdict(lambda: {"rows": 0, "amount": 0.0})
        for c in db["commissions"]:
            iname = intro_by_id.get(c.get("introducer_id"), "")
            did = c.get("deal_id")
            vc = db["did_to_vehicle"].get(did, "")
            key = (vc, iname)
            agg[key]["rows"] += 1
            agg[key]["amount"] += to_float(c.get("accrual_amount"))

        for (vc, iname), v in sorted(agg.items()):
            if iname in warnings_only:
                self.add_warning(
                    "introducer_warning_only_present",
                    vc,
                    f"{iname} rows={v['rows']} amount={round(v['amount'],2)}",
                )
                continue
            if iname in forbidden_global:
                self.add_failure(
                    "forbidden_broker_name_in_introducer_commissions",
                    vc,
                    f"{iname} rows={v['rows']} amount={round(v['amount'],2)}",
                )
            if vc in forbidden_by_vehicle and iname in set(forbidden_by_vehicle[vc]):
                self.add_failure(
                    "forbidden_vehicle_broker_name_in_introducer_commissions",
                    vc,
                    f"{iname} rows={v['rows']} amount={round(v['amount'],2)}",
                )

        # Explicit forbidden pairs
        forbidden_pairs = self.rules.get("forbidden_commission_pairs", [])
        if forbidden_pairs:
            pair_set = {(x.get("vehicle"), x.get("introducer")) for x in forbidden_pairs}
            for (vc, iname), v in sorted(agg.items()):
                if (vc, iname) in pair_set and v["rows"] > 0:
                    self.add_failure(
                        "forbidden_commission_pair_present",
                        vc,
                        f"introducer={iname} rows={v['rows']} amount={round(v['amount'],2)}",
                    )

        # Specific documented amount/rate checks
        if self.rules["checks"].get("specific_rule_assertions", False):
            comm_by_tuple = defaultdict(lambda: {"amount": 0.0, "rates": set(), "rows": 0})
            for c in db["commissions"]:
                vc = db["did_to_vehicle"].get(c.get("deal_id"), "")
                iid = c.get("investor_id")
                inv_name = db["iid_to_name"].get(iid, "")
                inv_key = name_key(inv_name, self.aliases) if inv_name else ""
                iname = intro_by_id.get(c.get("introducer_id"), "")
                btype = str(c.get("basis_type") or "")
                k = (vc, inv_key, iname, btype)
                comm_by_tuple[k]["amount"] += to_float(c.get("accrual_amount"))
                comm_by_tuple[k]["rows"] += 1
                rb = c.get("rate_bps")
                if rb not in (None, ""):
                    try:
                        comm_by_tuple[k]["rates"].add(int(float(rb)))
                    except Exception:
                        pass

            for rule in self.rules.get("specific_commission_expectations", []):
                vc = rule["vehicle"]
                inv_key = name_key(rule["investor"], self.aliases)
                iname = rule["introducer"]
                btype = rule["basis_type"]
                expected_amount = float(rule["amount"])
                expected_rate = int(rule["rate_bps"])
                got = comm_by_tuple.get((vc, inv_key, iname, btype))
                if not got:
                    self.add_failure(
                        "specific_commission_missing",
                        vc,
                        f"investor={rule['investor']} introducer={iname} basis={btype}",
                    )
                    continue
                got_amount = round(got["amount"], 2)
                if abs(got_amount - round(expected_amount, 2)) > 0.01:
                    self.add_failure(
                        "specific_commission_amount_mismatch",
                        vc,
                        f"investor={rule['investor']} introducer={iname} expected={expected_amount} got={got_amount}",
                    )
                if expected_rate not in got["rates"]:
                    self.add_failure(
                        "specific_commission_rate_mismatch",
                        vc,
                        f"investor={rule['investor']} introducer={iname} expected_rate={expected_rate} got_rates={sorted(got['rates'])}",
                    )

            set_cap_rule = self.rules.get("set_cap_vc215_spread_only_check", {})
            if set_cap_rule:
                vc = set_cap_rule.get("vehicle", "")
                iname = set_cap_rule.get("introducer", "")
                if set_cap_rule.get("invested_amount_must_be_zero"):
                    # agg is across all basis types, so check directly from comm rows for invested_amount
                    invested_total = 0.0
                    for c in db["commissions"]:
                        c_vc = db["did_to_vehicle"].get(c.get("deal_id"), "")
                        c_iname = intro_by_id.get(c.get("introducer_id"), "")
                        if c_vc == vc and c_iname == iname and str(c.get("basis_type") or "") == "invested_amount":
                            invested_total += to_float(c.get("accrual_amount"))
                    if abs(invested_total) > 0.01:
                        self.add_failure(
                            "set_cap_vc215_invested_amount_not_zero",
                            vc,
                            f"introducer={iname} invested_amount_total={round(invested_total,2)}",
                        )

        rows_checked = 0
        red_rows = []
        if bool(self.rules.get("run_contacts_checks", True)):
            # contacts file sanity checks
            contacts_path = ROOT / self.rules["dashboard_files"]["contacts"]
            wb = load_workbook(contacts_path, data_only=True)
            ws = wb[wb.sheetnames[0]]
            scope_set = set(self.rules["scope_vehicle_codes"])
            intro_names = {normalize_text(str(i.get("legal_name") or "")) for i in db["introducers"]}
            broker_names = {normalize_text(str(b.get("legal_name") or "")) for b in db["brokers"]}

            for r in range(2, ws.max_row + 1):
                series = ws.cell(r, 1).value
                legal = ws.cell(r, 2).value
                role = ws.cell(r, 7).value
                intro1 = ws.cell(r, 8).value
                if series in (None, "") and legal in (None, ""):
                    continue
                # simple red row detection on first 2 cells
                is_red = False
                for c in (ws.cell(r, 1), ws.cell(r, 2)):
                    rgb = getattr(getattr(c.fill, "fgColor", None), "rgb", None)
                    if isinstance(rgb, str) and rgb.upper().endswith("FF0000"):
                        is_red = True
                if is_red:
                    red_rows.append(r)
                    continue
                s = str(series).strip() if series not in (None, "") else ""
                if s and s not in scope_set:
                    continue
                rows_checked += 1
                if intro1 not in (None, "", "-"):
                    n = normalize_text(str(intro1))
                    if n not in intro_names and n not in broker_names:
                        self.add_failure(
                            "contacts_intro_not_found_in_db_master",
                            s,
                            f"row={r} intro={intro1}",
                            f"contacts:{r}",
                        )
                if role not in (None, "") and str(role).strip().lower() == "broker" and intro1 not in (None, "", "-"):
                    n = normalize_text(str(intro1))
                    if n not in broker_names:
                        self.add_failure(
                            "contacts_role_broker_missing_in_brokers_table",
                            s,
                            f"row={r} broker={intro1}",
                            f"contacts:{r}",
                        )

        summary = {
            "rules_version": self.rules.get("version"),
            "scope_vehicle_codes": scope_codes,
            "dashboard_active_rows": len(dash_active),
            "dashboard_zero_rows": len(dash_zero),
            "db_subscriptions": len(db["subs"]),
            "db_positions": len(db["positions_raw"]),
            "db_introductions": len(db["introductions"]),
            "db_commissions": len(db["commissions"]),
            "contacts_checked_rows": rows_checked,
            "contacts_red_rows_excluded": red_rows,
            "fail_count": len(self.failures),
            "warning_count": len(self.warnings),
            "fail_by_check": dict(Counter(x["check"] for x in self.failures)),
            "warn_by_check": dict(Counter(x["check"] for x in self.warnings)),
        }
        return summary

    def write_outputs(self, summary: dict[str, Any]):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        run_dir = self.outdir / f"run_{ts}"
        run_dir.mkdir(parents=True, exist_ok=True)

        out_json = run_dir / "audit_report.json"
        out_csv = run_dir / "audit_issues.csv"
        out_md = run_dir / "audit_summary.md"

        report = {
            "summary": summary,
            "failures": self.failures,
            "warnings": self.warnings,
        }
        out_json.write_text(json.dumps(report, indent=2))

        with out_csv.open("w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["severity", "check", "vehicle", "row_ref", "details"])
            w.writeheader()
            for r in self.failures + self.warnings:
                w.writerow(r)

        lines = []
        lines.append("# VC1 Audit Summary")
        lines.append("")
        lines.append(f"- Rules version: `{summary['rules_version']}`")
        lines.append(f"- Dashboard active rows: `{summary['dashboard_active_rows']}`")
        lines.append(f"- Dashboard zero rows: `{summary['dashboard_zero_rows']}`")
        lines.append(f"- DB subscriptions: `{summary['db_subscriptions']}`")
        lines.append(f"- DB positions: `{summary['db_positions']}`")
        lines.append(f"- DB introductions: `{summary['db_introductions']}`")
        lines.append(f"- DB commissions: `{summary['db_commissions']}`")
        lines.append(f"- Contacts checked rows: `{summary['contacts_checked_rows']}`")
        lines.append(f"- Failures: `{summary['fail_count']}`")
        lines.append(f"- Warnings: `{summary['warning_count']}`")
        lines.append("")
        if summary["fail_count"] == 0:
            lines.append("## Result")
            lines.append("PASS (no hard failures)")
        else:
            lines.append("## Result")
            lines.append("FAIL")
            lines.append("")
            lines.append("## Failure breakdown")
            for k, v in sorted(summary["fail_by_check"].items(), key=lambda x: (-x[1], x[0])):
                lines.append(f"- {k}: `{v}`")
        if summary["warning_count"] > 0:
            lines.append("")
            lines.append("## Warning breakdown")
            for k, v in sorted(summary["warn_by_check"].items(), key=lambda x: (-x[1], x[0])):
                lines.append(f"- {k}: `{v}`")
        lines.append("")
        lines.append("## Artifacts")
        lines.append(f"- `{out_json}`")
        lines.append(f"- `{out_csv}`")
        lines.append(f"- `{out_md}`")
        out_md.write_text("\n".join(lines))
        return run_dir, out_json, out_csv, out_md


def parse_args():
    p = argparse.ArgumentParser(description="VC2 deterministic audit")
    p.add_argument("--rules", type=Path, default=DEFAULT_RULES, help="Path to rules JSON")
    p.add_argument("--outdir", type=Path, default=DEFAULT_OUTDIR, help="Output directory")
    return p.parse_args()


def main() -> int:
    args = parse_args()
    if not args.rules.exists():
        print(f"Rules file not found: {args.rules}")
        return 2
    rules = json.loads(args.rules.read_text())
    auditor = Auditor(rules, args.outdir)
    summary = auditor.run_checks()
    run_dir, out_json, out_csv, out_md = auditor.write_outputs(summary)
    print(f"RUN_DIR: {run_dir}")
    print(f"REPORT_JSON: {out_json}")
    print(f"REPORT_CSV: {out_csv}")
    print(f"REPORT_MD: {out_md}")
    print(f"FAIL_COUNT: {summary['fail_count']}")
    print(f"WARN_COUNT: {summary['warning_count']}")
    return 0 if summary["fail_count"] == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
